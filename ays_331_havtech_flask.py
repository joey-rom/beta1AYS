# ays_331_betatwo_flask.py
from flask import Flask, request, jsonify, render_template, url_for, send_from_directory, redirect, send_file, session, flash
from flask_cors import CORS
import os
import spacy
import pandas as pd
import re
import logging
from werkzeug.utils import secure_filename
from PyPDF2 import PdfReader, PdfWriter
from concurrent.futures import ThreadPoolExecutor, as_completed, TimeoutError
import threading
import fitz  # PyMuPDF
from dotenv import load_dotenv
import openai  # Add this for OpenAI API integration
from pathlib import Path
import json  # Import JSON to serialize data for the session
from scripts.ays_314_script import process_pdf_file  # Import the external script
import base64
import tempfile
import io
import os
import io
import base64
import smtplib
import logging
import pandas as pd
from flask import Flask, request, jsonify, after_this_request
from werkzeug.utils import secure_filename
from email.mime.multipart import MIMEMultipart
from email.mime.application import MIMEApplication
from email.mime.text import MIMEText
from datetime import datetime
import os
import pandas as pd
import logging
from collections import Counter
from openpyxl import load_workbook
from openpyxl.styles import numbers
import shutil
import boto3
from helpers.helpers_async_s3_0_9 import (
    AWS_REGION, S3_BUCKET, S3_UPLOAD_PREFIX, S3_RESULTS_PREFIX,
    s3_key, s3_upload_bytes, s3_upload_file, s3_presign_get,
    slugify, make_project_id,
    submit_job, set_job, get_job,
    run_pipeline_to_s3,
    DASHBOARD_XLSX,
    log_completed_job_row,
    CUSTOMER_EXPORT_COLUMNS,
    AWS_REGION, 
    S3_BUCKET, 
    s3_key, 
    s3_presign_get, 
    project_index_from_dashboard, # <-- add this
    get_usage_stats,
    get_next_available_name_local,
)

_s3 = boto3.client("s3", region_name=AWS_REGION)


app = Flask(__name__)
CORS(app)
app.config['UPLOAD_FOLDER'] = 'uploads'
app.config['PROCESSED_FOLDER'] = 'processed'
app.config['ALLOWED_EXTENSIONS'] = {'pdf'}
app.secret_key = 'your_secret_key_here'

os.makedirs(app.config['UPLOAD_FOLDER'], exist_ok=True)
os.makedirs(app.config['PROCESSED_FOLDER'], exist_ok=True)

logging.basicConfig(level=logging.DEBUG)

load_dotenv()

model_path = Path(__file__).parent / 'static/en_core_web_sm'
nlp = spacy.load(model_path)

# ---- Logo base64 loader (global) ----
import os, base64

from functools import wraps

# Routes that don't require login
PUBLIC_ROUTES = {'login', 'static'}

@app.before_request
def require_login():
    # Allow public routes through
    if request.endpoint in PUBLIC_ROUTES:
        return None
    
    # If user is not in session, redirect to login
    if 'user' not in session:
        return redirect(url_for('login'))

def _load_logo_b64(app):
    # 1) env override if you want to supply the image via ENV
    env_val = os.environ.get("AYS_LOGO_BASE64")
    if env_val:
        return env_val

    # 2) fall back to static/logo.png
    try:
        logo_path = os.path.join(app.root_path, "static", "logo.png")
        with open(logo_path, "rb") as f:
            return base64.b64encode(f.read()).decode("utf-8")
    except Exception:
        # if missing, return empty string; the email template will just omit the image
        return ""

LOGO_BASE64 = _load_logo_b64(app)


# Local testing
openai.api_key = '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'

# Global terms lists
global equipment_terms, manufacturer_terms, model_terms, universal_terms, competitor_terms
equipment_terms, manufacturer_terms, model_terms, universal_terms, competitor_terms  = [], [], [], [], []

processing_cancelled = threading.Event()

users = {'havays': 'vT7!qL3#Zp2@Nw8$', 'aysadmin': 'H9^mR4!xK2@dS7&c'}

# Login page route
@app.route('/', methods=['GET', 'POST'])
def login():
    """
    Login page route for user authentication.
    """
    users = {'havays': 'vT7!qL3#Zp2@Nw8$', 'aysadmin': 'H9^mR4!xK2@dS7&c'}
    if request.method == 'POST':
        username = request.form['username']
        password = request.form['password']
        if username in users and users[username] == password:
            session['user'] = username
            return redirect(url_for('explorer_page'))
        else:
            flash('Invalid credentials, please try again.')
    return render_template('login.html')

def allowed_file(filename):
    return '.' in filename and filename.rsplit('.', 1)[1].lower() in app.config['ALLOWED_EXTENSIONS']

@app.route('/analysis', methods=['GET', 'POST'])
def upload_file():
    """
    Handles PDF uploads, processes the file, and renders results.
    """
    if request.method == 'POST':
        # Step 1: Retrieve the uploaded file
        file = request.files.get('file')
        if not file or not allowed_file(file.filename):
            flash("Invalid file type. Please upload a PDF file.")
            return redirect(url_for('upload_file'))

        # Step 2: Secure and save the uploaded file
        filename = secure_filename(file.filename)
        filepath = os.path.join(app.config['UPLOAD_FOLDER'], filename)
        file.save(filepath)

        try:
            # Step 3: Process the PDF using your processing function
            results = process_pdf_file(filepath)

            # Step 4: Verify and unpack results
            if not results:
                flash("Error: Failed to process the PDF. No results found.")
                return redirect(url_for('upload_file'))

            results_data = results.get('results', {})
            sections = results.get('sections', [])
            acceptance_sections = results.get('acceptance_sections', [])
            total_pages = results.get('total_pages', 0)
            processed_filename = results.get('filename')

            # Debugging Logs
            print("DEBUG: Results Data:", results_data)
            print("DEBUG: Sections:", sections)
            print("DEBUG: Acceptance Sections:", acceptance_sections)

            # Step 5: Validate key components
            if not sections and not acceptance_sections:
                flash("No relevant sections or acceptable manufacturers were found in the document.")
                return redirect(url_for('upload_file'))

            # Step 6: Store results in the session
            session['results_data'] = json.dumps(results)

            # Step 7: Render the results page
            return render_template(
                'template1_7.html',
                results=results_data,
                sections=sections,
                acceptance_sections=acceptance_sections,
                filename=processed_filename,
                total_pages=total_pages
            )

        except Exception as e:
            # Step 8: Handle unexpected errors
            logging.error(f"Exception occurred while processing the file: {e}")
            flash("An unexpected error occurred while processing the file. Please try again.")
            return redirect(url_for('upload_file'))

    # Render upload form if GET request
    return render_template('template1_7.html', results=None)


#@app.route('/view_pdf')
#def view_pdf():
#    """
#    Displays the processed PDF with highlights.
#    """
#    pdf = request.args.get('pdf')
#    if not pdf:
#        return "No PDF specified", 400
#    processed_pdf_path = os.path.join(app.config['PROCESSED_FOLDER'], pdf)
#    if not os.path.exists(processed_pdf_path):
#        return "File not found", 404
#    return send_from_directory(app.config['PROCESSED_FOLDER'], pdf)

#@app.route('/download_tables')
#def download_tables():
#    """
#    Allows users to download results tables as an Excel file.
#    """
 #   results_data = json.loads(session.get('results_data', '{}'))
#    if not results_data:
 #       return "No data available to download.", 400
#    output_filepath = os.path.join(app.config['PROCESSED_FOLDER'], 'results_tables.xlsx')
 #   with pd.ExcelWriter(output_filepath) as writer:
 #       for key, data in results_data.get('results', {}).items():
 #           if data:
 #               pd.DataFrame(data).to_excel(writer, sheet_name=key.capitalize(), index=False)
 #   return send_from_directory(app.config['PROCESSED_FOLDER'], 'results_tables.xlsx', as_attachment=True)

@app.route('/nginx-config')
def nginx_config():
    try:
        with open('/etc/nginx/conf.d/proxy.conf', 'r') as file:
            content = file.read()
        return f"<pre>{content}</pre>", 200
    except Exception as e:
        return str(e), 500

@app.route('/terms', methods=['GET', 'POST'])
def terms():
    filepath = 'terms/havtech_Terms.json'

    # Load existing terms from the JSON file
    try:
        with open(filepath, 'r') as file:
            terms_data = json.load(file)
            logging.debug("Loaded terms data successfully.")
    except FileNotFoundError:
        logging.warning(f"File {filepath} not found. Initializing empty terms data.")
        terms_data = []

    if request.method == 'POST':
        term_type = request.form.get('term_type')  # e.g. 'manufacturer', 'competitor', etc.
        action = request.form.get('action', 'add')  # 'add' or 'delete'
        logging.debug(f"POST /terms term_type={term_type}, action={action}")

        if not term_type:
            # No category specified; just go back to page
            return redirect(url_for('terms'))

        # Normalize to match JSON structure (titles: Manufacturer, Competitor, etc.)
        normalized_type = term_type.replace('_', ' ').lower()

        # Find the matching category in terms_data
        category = None
        for cat in terms_data:
            if cat.get('title', '').lower() == normalized_type:
                category = cat
                break

        if not category:
            logging.error(f"Category '{term_type}' not found in terms_data.")
            return jsonify({'status': 'error', 'message': f"Category '{term_type}' not found."}), 400

        if action == 'add':
            # Add a new term
            new_term = request.form.get(f'new_{term_type}_term', '').strip()
            logging.debug(f"Attempting to ADD term '{new_term}' to '{category['title']}'")

            if new_term:
                if new_term not in category['terms']:
                    category['terms'].append(new_term)
                    logging.debug(f"Added new term '{new_term}' to category '{category['title']}'.")
                else:
                    logging.debug(f"Term '{new_term}' already exists in category '{category['title']}'.")
            else:
                logging.debug("No new term provided; nothing to add.")

        elif action == 'delete':
            # Delete an existing term
            delete_term = request.form.get('delete_term', '').strip()
            logging.debug(f"Attempting to DELETE term '{delete_term}' from '{category['title']}'")

            if delete_term and delete_term in category['terms']:
                category['terms'].remove(delete_term)
                logging.debug(f"Deleted term '{delete_term}' from category '{category['title']}'.")
            else:
                logging.debug(f"Term '{delete_term}' not found in category '{category['title']}' – nothing deleted.")

        # Save updated terms back to JSON file
        try:
            with open(filepath, 'w') as file:
                json.dump(terms_data, file, indent=4)
            logging.debug("Successfully saved updated terms to file.")
        except Exception as e:
            logging.error(f"Error saving terms to JSON file: {e}")
            return jsonify({'status': 'error', 'message': str(e)}), 500

        # Redirect back to the appropriate section anchor
        return redirect(url_for('terms') + '#' + term_type)

    # Prepare terms for rendering
    categorized_terms = {
        'equipment_terms': next((x['terms'] for x in terms_data if x['title'].lower() == 'equipment'), []),
        'manufacturer_terms': next((x['terms'] for x in terms_data if x['title'].lower() == 'manufacturer'), []),
        'model_terms': next((x['terms'] for x in terms_data if x['title'].lower() == 'model'), []),
        'universal_terms': next((x['terms'] for x in terms_data if x['title'].lower() == 'universal'), []),
        'competitor_terms': next((x['terms'] for x in terms_data if x['title'].lower() == 'competitor'), []),
    }

    logging.debug(f"Categorized terms before rendering: {categorized_terms}")

    return render_template('terms4.html', **categorized_terms)


@app.route('/view_pdf')
def view_pdf():
    pdf = request.args.get('pdf')
    page = request.args.get('page', 1, type=int)
    word = request.args.get('word', '')

    if not pdf:
        return "No PDF specified", 400

    processed_pdf_path = os.path.join('processed', pdf)
    if not os.path.exists(processed_pdf_path):
        logging.error(f"File not found: {processed_pdf_path}")
        return "File not found", 404

    # Pass required information to the template
    return render_template(
        'pdf_viewer1.html',
        pdf_url=url_for('download_file', filename=pdf),
        page_number=page,
        filename=os.path.splitext(pdf)[0],
        word=word
    )


@app.route('/download_section')
def download_section():
    start_page = request.args.get('start_page', type=int)
    stop_page = request.args.get('stop_page', type=int)
    section_name = request.args.get('section_name')
    filename = request.args.get('filename')
    pdf_path = os.path.join('processed', filename)

    if not os.path.exists(pdf_path):
        logging.error(f"File not found: {pdf_path}")
        return "File not found", 404

    try:
        # Create a new PDF with only the desired pages
        reader = PdfReader(pdf_path)
        writer = PdfWriter()

        for page_num in range(start_page - 1, stop_page):
            writer.add_page(reader.pages[page_num])

        output_file = os.path.join('processed', f"{section_name}_Section.pdf")
        with open(output_file, 'wb') as f:
            writer.write(f)

        # Return the file for download
        return send_file(output_file, as_attachment=True)
    except Exception as e:
        logging.error(f"Error generating section PDF: {e}")
        return "Error processing the section.", 500



@app.route('/download_tables')
def download_tables():
    # Retrieve results_data from the session
    results_data = json.loads(session.get('results_data', '{}'))  # Deserialize JSON from session

    if not results_data or not any(results_data.values()):
        return "No data available to download.", 400

    # Create a new Excel file with pandas to hold the tables
    output_file_path = os.path.join(app.config['PROCESSED_FOLDER'], 'results_tables.xlsx')

    # Write data to the Excel file if present
    with pd.ExcelWriter(output_file_path) as writer:
        if results_data['equipment_tables']:
            equipment_df = pd.DataFrame(results_data['equipment_tables'])
            equipment_df.to_excel(writer, sheet_name='Equipment', index=False)
        if results_data['manufacturer_tables']:
            manufacturer_df = pd.DataFrame(results_data['manufacturer_tables'])
            manufacturer_df.to_excel(writer, sheet_name='Manufacturer', index=False)
        if results_data['model_tables']:
            model_df = pd.DataFrame(results_data['model_tables'])
            model_df.to_excel(writer, sheet_name='Model', index=False)
        if results_data['universal_tables']:
            universal_df = pd.DataFrame(results_data['universal_tables'])
            universal_df.to_excel(writer, sheet_name='Universal', index=False)
        if results_data['competitor_tables']:
            competitor_df = pd.DataFrame(results_data['competitor_tables'])
            competitor_df.to_excel(writer, sheet_name='Competitors', index=False)


    # Send the generated file to the user
    return send_file(output_file_path, as_attachment=True)

@app.route('/processed/<path:filename>')
def download_file(filename):
    return send_from_directory(app.config['PROCESSED_FOLDER'], filename)

@app.route('/cancel', methods=['POST'])
def cancel_processing():
    processing_cancelled.set()
    logging.info("Processing has been cancelled by the user.")
    return jsonify({"status": "cancelled"})

# List of valid project types (names only)
VALID_PROJECT_TYPES = [
    "New Development",
    "Rehabilitation (Rehab)",
    "Renovation",
    "Remodeling",
    "Restoration",
    "Expansion (Addition)",
    "Adaptive Reuse",
    "Demolition",
    "Infrastructure Development",
    "Fit-Out (Interior Build-Out)",
    "Green Building (Sustainable Construction)",
    "Tenant Improvement",
    "Historic Preservation",
    "Civil Engineering Works",
    "Seismic Retrofitting",
    "Brownfield Redevelopment",
    "Commercial Development",
    "Residential Development",
    "Industrial Construction",
    "Public Sector Projects"
]

# Flask route for summarizing PDF files
@app.route('/summarize', methods=['GET', 'POST'])
def summarize_pdf():
    summary = None
    info_table = None
    
    if request.method == 'POST':
        file = request.files['file']
        if file and allowed_file(file.filename):
            filename = secure_filename(file.filename)
            filepath = os.path.join(app.config['UPLOAD_FOLDER'], filename)
            file.save(filepath)

            # Extract and summarize the PDF, including extracting information for the table
            summary = extract_and_summarize(filepath)
            
            # Extract information for the table using the combined summary
            info_table = extract_project_information_chunked(summary)
            
            # Debug: Print info_table to check its content
            print("Extracted Project Information:", info_table)

    return render_template('summarize_upload1.html', summary=summary, info_table=info_table)


def generate_summary(text, max_words):
    """Generates a summary of the given text, limited to a specified number of words."""
    MAX_TOKENS_OUTPUT = 750  # Adjust based on word limit needs

    # Create a prompt for generating the summary
    prompt_content = (
        f"Summarize the following text into a coherent and concise final summary of approximately {max_words} words. "
        "Ensure that the summary captures all key aspects and presents them in a clear, conversational style. "
        "Avoid unnecessary details, bullet points, section headers, or lists. Provide a single, cohesive narrative that flows naturally.\n\n" 
        + text
    )

    print(f"Calling OpenAI API to generate a summary of {max_words} words...")

    response = openai.ChatCompletion.create(
        model="gpt-4o-mini",
        messages=[
            {"role": "system", "content": "You are a helpful assistant specialized in summarizing documents."},
            {"role": "user", "content": prompt_content}
        ],
        max_tokens=MAX_TOKENS_OUTPUT,
        temperature=0.5
    )

    # Check if the response is valid
    if 'choices' in response and len(response['choices']) > 0:
        summary_text = response['choices'][0]['message']['content']
        print(f"Summary generated successfully. Length: {len(summary_text.split())} words.")
        return summary_text
    else:
        print("Error: Invalid response from OpenAI API.")
        return "Error: Failed to summarize the document."


def extract_and_summarize(pdf_path):
    """Extracts and summarizes text from a PDF document following a step-by-step iterative process."""
    doc = fitz.open(pdf_path)
    total_pages = len(doc)

    # Restrict pages to review based on total page count
    if total_pages <= 100:
        pages_to_review = 20
    else:
        pages_to_review = 50

    combined_summary = ""  # This will store the progressively built summary

    # Process pages in chunks of pages_to_review
    print(f"Total pages to process: {total_pages}")
    for start_page in range(0, pages_to_review, 10):  # Process in chunks of 10 pages
        end_page = min(start_page + 10, pages_to_review)
        extracted_text = extract_pages(doc, start_page, end_page)

        if start_page == 0:
            # If this is the first chunk, summarize to 250 words
            print(f"Summarizing first chunk from pages {start_page} to {end_page}")
            combined_summary = generate_summary(extracted_text, max_words=250)
        else:
            # Combine the previous summary with the new extracted text
            combined_text = combined_summary + "\n\n" + extracted_text
            # Generate a new summary of 500 words or less
            print(f"Combining and summarizing chunk from pages {start_page} to {end_page}")
            combined_summary = generate_summary(combined_text, max_words=500)

    doc.close()

    # Final summary is now stored in combined_summary
    print("Final summary generated.")
    return combined_summary or "Summary not available"


def extract_pages(doc, start_page, end_page):
    """Extracts text from a range of pages in a PDF document."""
    extracted_text = ""
    for page_num in range(start_page, end_page):
        if page_num >= 0 and page_num < len(doc):  # Check bounds
            page = doc.load_page(page_num)
            extracted_text += page.get_text()
    return extracted_text


def extract_project_information_chunked(text):
    """
    Extracts the project information by splitting the text into chunks and sending it to GPT.
    """
    MAX_TOKENS_INPUT = 1200  # Set a safe limit for tokens to prevent exceeding context length
    MAX_TOKENS_OUTPUT = 250  # Limit output tokens to keep it concise
    chunk_size = 1000  # Smaller chunk size to ensure safety within token limits
    project_info_chunks = []

    # Split text into manageable chunks
    chunks = [text[i:i + chunk_size] for i in range(0, len(text), chunk_size)]

    # Process chunks concurrently
    with ThreadPoolExecutor(max_workers=5) as executor:  # Use directly without `concurrent.`
        futures = [executor.submit(process_chunk, chunk, MAX_TOKENS_OUTPUT) for chunk in chunks]
        for future in as_completed(futures):  # Use directly without `concurrent.`
            result = future.result()
            if "Error" not in result:  # Only add valid results
                project_info_chunks.append(result)
            else:
                print(f"Error in processing chunk: {result}")  # Debugging information

    # Combine project info results
    combined_project_info = "\n".join(project_info_chunks)
    extracted_info = extract_info_from_response(combined_project_info)

    # Debug: Print the extracted_info to verify the correct format
    print("Extracted Info from Response:", extracted_info)

    return extracted_info

def process_chunk(chunk, max_tokens_output):
    """Processes a single chunk by sending it to the OpenAI API and returning the response."""
    try:
        # Use GPT-4o Mini to infer project details from each chunk
        response = openai.ChatCompletion.create(
            model="gpt-4o-mini",
            messages=[
                {"role": "system", "content": (
                    "You are an assistant specialized in analyzing construction documents. Your task is to identify "
                    "specific roles and their associated entities from the text. Look for and extract the following information:\n"
                    "- **Owner**: The entity or organization that owns the project.\n"
                    "- **Owner's Representative**: The person or organization representing the owner in the project.\n"
                    "- **Engineering Firm**: The engineering firm involved in planning or specifications.\n"
                    "- **Architect**: The architect or architectural firm associated with the project.\n"
                    "- **Project Type**: The type of project (e.g., New Development, Renovation, Demolition).\n\n"
                    "Provide the extracted information clearly in this format:\n"
                    "- Owner: [Owner Name]\n"
                    "- Owner's Representative: [Representative Name]\n"
                    "- Engineering Firm: [Engineering Firm Name]\n"
                    "- Architect: [Architect Name]\n"
                    "- Project Type: [Valid Project Type]\n\n"
                    "If not mentioned, respond with 'Not Specified'. Analyze the following text carefully:\n" + chunk
                )},
                {"role": "user", "content": f"The following is a text chunk extracted from a construction document:\n{chunk}"}
            ],
            max_tokens=max_tokens_output,
            temperature=0.5,
            timeout=120  # Set a higher timeout value
        )

        # Check if response is valid and return the result
        if 'choices' in response and len(response['choices']) > 0:
            return response.choices[0].message['content']
        else:
            print("Error: Invalid response from OpenAI API.")
            return "Error: Failed to extract project information."

    except Exception as e:
        print(f"Exception occurred while processing chunk: {e}")
        return "Error: An exception occurred during the extraction process."

def extract_info_from_response(response_text):
    """
    Extracts the owner, owner's representative, engineering firm, architect, and project type
    from the GPT-4o Mini response. Ensures Project Type is only selected from the predefined list.
    """
    extracted_info = {
        "Owner": "Not Specified",
        "Owner's Representative": "Not Specified",
        "Engineering Firm": "Not Specified",
        "Architect": "Not Specified",
        "Project Type": "Not Specified"
    }

    # Enhanced parsing logic to match lines and extract the relevant information
    for line in response_text.split('\n'):
        if 'Owner:' in line and 'Not Specified' not in line:
            extracted_info["Owner"] = line.split(':', 1)[1].strip()
        elif "Owner's Representative:" in line and 'Not Specified' not in line:
            extracted_info["Owner's Representative"] = line.split(':', 1)[1].strip()
        elif 'Engineering Firm:' in line and 'Not Specified' not in line:
            extracted_info["Engineering Firm"] = line.split(':', 1)[1].strip()
        elif 'Architect:' in line and 'Not Specified' not in line:
            extracted_info["Architect"] = line.split(':', 1)[1].strip()
        elif 'Project Type:' in line and 'Unknown' not in line:
            project_type = line.split(':', 1)[1].strip()
            # Validate Project Type against the list of valid types
            if project_type in VALID_PROJECT_TYPES:
                extracted_info["Project Type"] = project_type
            else:
                extracted_info["Project Type"] = "Unknown"  # Set to 'Unknown' if not in list

    # Debug: Check if all required keys have valid data
    print("Extracted Information Details:", extracted_info)

    return extracted_info



import os
import io
import base64
import smtplib
import logging
import pandas as pd
from flask import Flask, request, jsonify, after_this_request
from werkzeug.utils import secure_filename
from email.mime.multipart import MIMEMultipart
from email.mime.application import MIMEApplication
from email.mime.text import MIMEText

# SMTP Config for Office365
SMTP_SERVER = "smtp.office365.com"
SMTP_PORT = 587
SMTP_USERNAME = "havtech@areyouspecified.com"
SMTP_PASSWORD = "HH#d3t@12345%%"
SENDER_EMAIL = "havtech@areyouspecified.com"

# ---------------------------------------
# HELPER: Send email with attachments
# ---------------------------------------
from email.mime.image import MIMEImage


@app.get("/s3/ping")
def s3_ping():
    key = s3_key(S3_UPLOAD_PREFIX, "app-ping.txt")
    s3_upload_bytes(b"hello from flask", key)
    url = s3_presign_get(key, expires=600)
    return jsonify({"key": key, "presigned_url": url})

def send_email_with_attachments(to_address, subject, body_text, attachments, logo_path=None, original_message_id=None):
    msg = MIMEMultipart('related')
    msg['From'] = SENDER_EMAIL
    msg['To'] = to_address
    msg['Subject'] = subject

    # Add reply headers if available
    if original_message_id:
        msg['In-Reply-To'] = original_message_id
        msg['References'] = original_message_id

    # Create alternative part for HTML body
    alt_part = MIMEMultipart('alternative')
    msg.attach(alt_part)

    # Attach the HTML body
    alt_part.attach(MIMEText(body_text, 'html'))

    # Attach inline logo image
    if logo_path and os.path.isfile(logo_path):
        with open(logo_path, 'rb') as img_file:
            img = MIMEImage(img_file.read(), name=os.path.basename(logo_path))
            img.add_header('Content-ID', '<logo_cid>')
            img.add_header('Content-Disposition', 'inline', filename='logo.png')
            msg.attach(img)

    # Attach other files
    for path in attachments:
        with open(path, 'rb') as f:
            part = MIMEApplication(f.read())
            part.add_header('Content-Disposition', 'attachment', filename=os.path.basename(path))
            msg.attach(part)

    # Send the message
    with smtplib.SMTP(SMTP_SERVER, SMTP_PORT) as server:
        server.starttls()
        server.login(SMTP_USERNAME, SMTP_PASSWORD)
        server.sendmail(SENDER_EMAIL, to_address, msg.as_string())


# ---------------------------------------
# HELPER: Write Excel with all 7 tabs
# ---------------------------------------
import pandas as pd
import logging

def write_results_to_excel(results, excel_path):
    try:
        with pd.ExcelWriter(excel_path, engine='xlsxwriter') as writer:
            workbook = writer.book
            header_format = workbook.add_format({
                'bold': True,
                'bg_color': '#F58220',
                'font_color': '#FFFFFF',
                'border': 1
            })
            content_format = workbook.add_format({
                'border': 1
            })

            def write_df_to_sheet(df, sheet_name):
                df = df[[col for col in df.columns if 'link' not in col.lower()]]

                if df.empty:
                    df = pd.DataFrame(columns=[''])

                df.to_excel(writer, sheet_name=sheet_name, index=False, startrow=1, header=False)
                worksheet = writer.sheets[sheet_name]

                for col_num, value in enumerate(df.columns.values):
                    worksheet.write(0, col_num, value, header_format)
                    worksheet.set_column(col_num, col_num, max(15, len(str(value)) + 2))

                for row_num in range(1, len(df) + 1):
                    for col_num in range(len(df.columns)):
                        worksheet.write(row_num, col_num, df.iloc[row_num - 1, col_num], content_format)

                worksheet.freeze_panes(1, 0)

            # Sections
            df_sections = pd.DataFrame(results.get('sections', []))
            write_df_to_sheet(df_sections, 'Sections')

            # Acceptance Sections (renamed to 'Specified Manufacturers')
            df_accept = pd.DataFrame(results.get('acceptance_sections', []))
            write_df_to_sheet(df_accept, 'Specified Manufacturers')

            # Keyword Sheets with updated sheet names
            category_mapping = {
                'manufacturer': 'Your Manufacturers',
                'competitor': 'Comp Manufacturers',
                'equipment': 'Equipment',
                'model': 'Model',
                'universal': 'BOD by Section',
            }

            results_data = results.get('results', {})
            for key, sheet_name in category_mapping.items():
                records = results_data.get(key, [])
                df = pd.DataFrame(records)
                if df.empty:
                    df = pd.DataFrame(columns=['Word', 'Page', 'Section', 'Section Name'])
                else:
                    df = df[['Word', 'Page', 'Section', 'Section Name']]
                write_df_to_sheet(df, sheet_name)

        logging.info(f"✅ Excel workbook written to {excel_path}")

    except Exception as e:
        logging.error(f"Error writing Excel workbook: {e}", exc_info=True)

# ---------------------------------------
# HELPER: Build nicely formatted email body
# ---------------------------------------


import fitz  # PyMuPDF

def get_highlighted_pages(pdf_path):
    doc = fitz.open(pdf_path)
    highlighted_pages = set()
    for page_number in range(len(doc)):
        page = doc[page_number]
        for annot in page.annots():
            if annot.type[0] == 8:  # 8 = Highlight
                highlighted_pages.add(page_number)
                break
    return sorted(highlighted_pages)


import logging
from collections import defaultdict
import fitz  # PyMuPDF

def create_highlighted_only_pdf(highlighted_pdf_path, results, output_path):
    src_doc = fitz.open(highlighted_pdf_path)
    new_doc = fitz.open()
    added_pages = 0

    for i in range(len(src_doc)):  # Check every page
        page = src_doc[i]
        has_highlight = False

        for annot in page.annots():
            if annot.type[0] == 8:  # Highlight
                has_highlight = True
                break

        if has_highlight:
            new_doc.insert_pdf(src_doc, from_page=i, to_page=i)
            logging.debug(f"✅ Included page {i} with highlight(s).")
            added_pages += 1
        else:
            logging.debug(f"⛔ Skipped page {i} (no highlights).")

    if added_pages > 0:
        new_doc.save(output_path)
        logging.info(f"✅ Highlighted-only PDF saved: {output_path} with {added_pages} pages")
    else:
        logging.warning("⚠️ No valid highlighted pages found — no output generated.")

    src_doc.close()
    new_doc.close()

# ---------------------------------------
# Format Email Tables
# ---------------------------------------
def format_table(rows):
    if not rows:
        return "None found."

    col_widths = [15, 6, 8, 25]  # Adjust as needed
    headers = ["Word", "Page", "Section", "Section Name"]
    output = []

    # Header row
    header_line = " | ".join(h.ljust(col_widths[i]) for i, h in enumerate(headers))
    separator = "-+-".join("-" * col_widths[i] for i in range(len(headers)))
    output.append(header_line)
    output.append(separator)

    # Data rows
    for row in rows:
        line = " | ".join(str(row.get(h, "")).ljust(col_widths[i]) for i, h in enumerate(headers))
        output.append(line)

    return "\n".join(output)



from xhtml2pdf import pisa
import io
import logging

def html_to_pdf(html_content, output_path):
    """
    Convert HTML to PDF using xhtml2pdf.

    If xhtml2pdf fails (e.g. due to a malformed table that causes
    'must have at least a row and column'), we log the error and
    write a simple fallback PDF instead of crashing the job.
    """
    try:
        with open(output_path, "wb") as f:
            result = pisa.CreatePDF(io.StringIO(html_content), dest=f)

        # pisa.CreatePDF returns an object with .err count
        if result.err:
            logging.error("xhtml2pdf reported %s errors while rendering %s",
                          result.err, output_path)

    except Exception as e:
        logging.exception("html_to_pdf failed, writing fallback PDF instead")

        # Fallback: write a minimal PDF so the pipeline can continue
        try:
            from reportlab.pdfgen import canvas

            buf = io.BytesIO()
            c = canvas.Canvas(buf)
            c.drawString(72, 720, "Email summary could not be rendered as PDF.")
            c.drawString(72, 700, "Check the HTML view or dashboard for details.")
            c.save()
            buf.seek(0)

            with open(output_path, "wb") as f:
                f.write(buf.read())

            logging.info("Fallback PDF written to %s", output_path)
        except Exception:
            # If even the fallback fails, log but don't re-raise
            logging.exception("Fallback PDF generation also failed")



# ---------------------------------------
# FLASK ROUTE: /api/process
# ---------------------------------------
import shutil

@app.route('/api/process', methods=['POST'])
def api_process_pdf():
    try:
        data = request.get_json()

        # basics
        filename = data.get('AttachmentName')
        content_b64 = data.get('AttachmentContent')
        original_receiver = data.get('From')
        original_subject = data.get('Subject')
        original_message_id = data.get('MessageID')

        # ✅ NEW: meta fields passed from /project-process payload (or Power Automate)
        meta_fields = data.get('Meta') or {}

        if not all([filename, content_b64, original_receiver, original_subject]):
            return jsonify({"error": "Missing required fields."}), 400

        # create AYS id
        timestamp = datetime.now().strftime("%Y%m%d-%H%M%S")
        ays_id = f"AYS-{timestamp}"

        # save original
        secure_name = secure_filename(filename)
        upload_path = os.path.join(app.config['UPLOAD_FOLDER'], secure_name)
        with open(upload_path, 'wb') as f:
            f.write(base64.b64decode(content_b64))

        # process
        results = process_pdf_file(upload_path)
        if not results or not results.get('results'):
            return jsonify({"error": "Processing failed."}), 500

        results_data = results['results']
        total_keywords = sum(len(t) for t in results_data.values() if t)

        manufacturer_rows = [
            {"Word": r.get('Word',''), "Page": r.get('Page',''), "Section": r.get('Section',''), "Section Name": r.get('Section Name','')}
            for r in results_data.get('manufacturer', [])
        ]
        competitor_rows = [
            {"Word": r.get('Word',''), "Page": r.get('Page',''), "Section": r.get('Section',''), "Section Name": r.get('Section Name','')}
            for r in results_data.get('competitor', [])
        ]

        # create Excel (unchanged)
        excel_filename = f"tables_{secure_name.rsplit('.', 1)[0]}.xlsx"
        excel_path = os.path.join(app.config['PROCESSED_FOLDER'], excel_filename)
        write_results_to_excel(results, excel_path)

        # load highlighted pdf
        highlighted_pdf_filename = results.get('filename')
        if not highlighted_pdf_filename:
            return jsonify({"error": "Highlighted PDF missing."}), 500
        highlighted_pdf_path = os.path.join(app.config['PROCESSED_FOLDER'], highlighted_pdf_filename)
        if not os.path.isfile(highlighted_pdf_path):
            return jsonify({"error": "Highlighted PDF not found."}), 500

        # highlights-only
        original_base = os.path.splitext(filename)[0]
        highlighted_only_pdf_path = os.path.join(app.config['PROCESSED_FOLDER'], f"only_highlights_{original_base}.pdf")
        create_highlighted_only_pdf(highlighted_pdf_path, results, highlighted_only_pdf_path)

        # recommendation logic (unchanged)
        has_mfg = bool(manufacturer_rows)
        has_comp = bool(competitor_rows)
        if has_mfg and has_comp:
            recommendation = "You and your competitor are specified. Bid this opportunity!"
            subject_summary = "Specified - Bid!"
        elif has_mfg:
            recommendation = "You are Specified! Bid this opportunity!"
            subject_summary = "Specified - Bid!"
        elif has_comp:
            recommendation = "Your competitor is specified - Review this opportunity."
            subject_summary = "Competitor Specified"
        else:
            recommendation = "You are not specified. Pass on this opportunity"
            subject_summary = "Not Specified - Do not Bid!"

       
        # build email body with customer project metadata
        email_body = generate_email_body(
            original_subject,
            total_keywords,
            manufacturer_rows,
            competitor_rows,
            recommendation,
            LOGO_BASE64,
            ays_id,
            highlighted_pdf_key=None,          # no S3 key in this path
            meta=meta_fields,
            sections=results.get("sections")   # safe even if None/absent
        )

        
        # render the email summary PDF from the final HTML
        email_pdf_path = os.path.join(app.config['PROCESSED_FOLDER'], f"{ays_id}_email_summary.pdf")
        html_to_pdf(email_body, email_pdf_path)



        # log dashboard (unchanged)
        log_results_to_excel(
            ays_id=ays_id,
            from_email=original_receiver,
            project_name=original_subject,
            manufacturer_terms=[row["Word"] for row in manufacturer_rows],
            recommendation=subject_summary
        )

        # rename artifacts to project name (unchanged)
        project_name_clean = secure_filename(original_subject)
        renamed_files = []
        rename_map = {
            highlighted_pdf_path: f"{project_name_clean}_highlighted.pdf",
            highlighted_only_pdf_path: f"{project_name_clean}_only_highlights.pdf",
            excel_path: f"{project_name_clean}_tables.xlsx",
            email_pdf_path: f"{project_name_clean}_email_summary.pdf"
        }
        for old_path, new_filename in rename_map.items():
            new_path = os.path.join(app.config['PROCESSED_FOLDER'], new_filename)
            os.rename(old_path, new_path)
            renamed_files.append(new_path)

        # zip and return (unchanged)
        import zipfile, tempfile
        with tempfile.NamedTemporaryFile(delete=False, suffix=".zip") as tmp_zip:
            with zipfile.ZipFile(tmp_zip.name, 'w') as zipf:
                for file_path in renamed_files:
                    zipf.write(file_path, arcname=os.path.basename(file_path))
            zip_download_path = tmp_zip.name

        zip_name_for_download = f"{project_name_clean}_AYS_Report.zip"
        return send_file(zip_download_path, as_attachment=True, download_name=zip_name_for_download)

    except Exception as e:
        logging.error(f"API Exception: {e}", exc_info=True)
        return jsonify({"error": "Internal Server Error"}), 500


@app.route('/dashboard1')
def view_dashboard():
    dashboard_path = DASHBOARD_XLSX
    if not os.path.isfile(dashboard_path):
        return "<p>No dashboard data yet.</p>"

    df = pd.read_excel(dashboard_path, dtype=str).fillna("")
    needed = [
        "Date", "AYS ID", "Email", "Project Name",
        "Manufacturer Terms", "Recommendation", "Download URL"
    ]
    for c in needed:
        if c not in df.columns:
            df[c] = ""

    # Sort newest first by Date
    try:
        df["Date"] = pd.to_datetime(df["Date"], errors="coerce")
        df = df.sort_values(by="Date", ascending=False)
        df["Date"] = df["Date"].dt.strftime("%m/%d/%Y")
    except Exception:
        pass

    table_columns = needed
    table_data = df[table_columns].to_dict(orient="records")

    counts = df["Recommendation"].value_counts()
    labels = counts.index.tolist()
    values = counts.values.tolist()

    return render_template(
        "dashboard1.html",
        table_columns=table_columns,
        table_data=table_data,
        chart_labels=labels,
        chart_values=values,
    )


# === UPDATED: /dashboard/download (uses the shared path) ===
from helpers.helpers_async_s3_0_9 import DASHBOARD_XLSX
import tempfile
from datetime import datetime
from flask import after_this_request
try:
    from helpers.helpers_async_s3_0_9 import CUSTOMER_EXPORT_COLUMNS
except ImportError:
    CUSTOMER_EXPORT_COLUMNS = [
        "Date", "AYS ID", "Email", "Project Name", "Manufacturer Terms", "Recommendation"
    ]


@app.route('/dashboard1/download')
def download_dashboard_excel():
    if not os.path.isfile(DASHBOARD_XLSX):
        return "No dashboard data yet.", 404

    df = pd.read_excel(DASHBOARD_XLSX, dtype=str).fillna("")

    # Keep only customer-safe columns (ignore any missing gracefully)
    cols = [c for c in CUSTOMER_EXPORT_COLUMNS if c in df.columns]
    if not cols:
        return "No customer-visible data to export.", 404

    df = df[cols]

    # Keep only rows that have basic customer info
    if "Email" in df.columns:
        df = df[df["Email"].str.strip() != ""]
    if "Project Name" in df.columns:
        df = df[df["Project Name"].str.strip() != ""]

    if df.empty:
        return "No customer rows available to export.", 404

    # Write to a temp file and send it; then delete it
    tmp = tempfile.NamedTemporaryFile(delete=False, suffix=".xlsx")
    tmp_path = tmp.name
    tmp.close()
    df.to_excel(tmp_path, index=False)

    @after_this_request
    def _cleanup(response):
        try:
            os.unlink(tmp_path)
        except Exception:
            pass
        return response

    filename = f"ays_dashboard_{datetime.now().strftime('%Y%m%d')}.xlsx"
    return send_file(tmp_path, as_attachment=True, download_name=filename)



@app.route('/project')
def start_project():
    return render_template('project_submitted5.html')

# ---- imports (clean) ----
import os
import base64
import logging
from datetime import datetime

import pandas as pd
from flask import (
    request, render_template, redirect, url_for,
    jsonify, abort
)

# pull everything we use from your helpers module
from helpers.helpers_async_s3_0_9 import (
    # job + pipeline
    make_project_id, submit_job, set_job, get_job, run_pipeline_to_s3,
    # dashboard write + read
    DASHBOARD_XLSX, log_completed_job_row,
    list_projects_from_dashboard, list_project_docs, get_project_meta,
    # s3 presign + constants
    s3_presign_get, S3_RESULTS_PREFIX, S3_UPLOAD_PREFIX,
)

# Optional: “From” identity shown in the Explorer email pane
AYS_FROM_NAME  = os.getenv("AYS_FROM_NAME",  "AYS Reports")
AYS_FROM_EMAIL = os.getenv("AYS_FROM_EMAIL", "noreply@areyouspecified.com")



# === Helpers to render meta fields into the email ===
def _render_meta_html(meta: dict) -> str:
    """Return an HTML table for meta_* fields (labels are meta_* without the prefix)."""
    if not meta:
        return ""
    rows = []
    for k, v in meta.items():
        if not v:
            continue
        label = k.replace("meta_", "").replace("_", " ").title()
        rows.append(
            "<tr>"
            f"<td style='padding:6px 10px;border:1px solid #ddd;white-space:nowrap;'><b>{label}</b></td>"
            f"<td style='padding:6px 10px;border:1px solid #ddd;'>{v}</td>"
            "</tr>"
        )
    if not rows:
        return ""
    return (
        "<h3 style='margin:16px 0 8px 0;'>Project Details</h3>"
        "<table cellpadding='0' cellspacing='0' style='border-collapse:collapse;"
        "font-family:Arial,sans-serif;font-size:13px;'>"
        f"{''.join(rows)}"
        "</table><br>"
    )

def _prepend_meta_to_email(html_body: str, meta: dict) -> str:
    """Prepend the Project Details table to the email HTML."""
    meta_html = _render_meta_html(meta or {})
    return (meta_html + html_body) if meta_html else html_body



# =========================
# Project submission (background jobs)
# =========================
from datetime import datetime
import base64
import logging
from flask import jsonify, render_template, request

@app.post('/project-process')
def project_process():
    """
    Handles project submissions (uploads multiple PDFs + project info)
    and queues them for background processing.
    """
    try:
        subject = request.form.get('Subject') or "Untitled_Project"
        email   = request.form.get('From') or "unknown@example.com"
        files   = request.files.getlist('files')

        if not files:
            return jsonify({"error": "No files uploaded"}), 400

        # Capture project metadata
        meta_fields = {
            k: v for k, v in request.form.items()
            if k.startswith("meta_")
        }

        project_id = make_project_id(subject)
        job_ids = []

        callbacks = dict(
            process_pdf_file=process_pdf_file,
            create_highlighted_only_pdf=create_highlighted_only_pdf,
            generate_email_body=generate_email_body,
            logo_base64=LOGO_BASE64,
            write_results_to_excel=write_results_to_excel,
        )

        # One timestamp for this batch (each file still gets its own doc_folder via this)
        submitted_at_iso = datetime.utcnow().isoformat(timespec='seconds') + "Z"

        for file in files:
            if not file or not file.filename:
                continue

            # --------------------------
            # DUPLICATE-SAFE NAMING (LOCAL /uploads)
            # --------------------------
            original = file.filename
            base, ext = os.path.splitext(original)
            if not ext:
                ext = ".pdf"

            # Clean the base filename (avoid weird symbols)
            safe_base = re.sub(r'[^A-Za-z0-9._ -]+', '_', base).strip() or "file"

            # Ask local upload folder for the next available name
            unique_name = get_next_available_name_local(
                safe_base,
                ext,
                app.config['UPLOAD_FOLDER'],
            )
            # --------------------------

            payload = {
                "AttachmentName": unique_name,  # what we actually write to disk
                "DisplayName": unique_name,     # what we log/display on dashboard
                "AttachmentContent": base64.b64encode(file.read()).decode('utf-8'),
                "From": email,
                "Subject": subject,
                "MessageID": None,
                "ProjectID": project_id,
                "SubmittedAt": submitted_at_iso,
                "Meta": meta_fields,
            }

            def job_fn(job_id, payload_inner):
                try:
                    set_job(job_id, state="STARTED", project_id=project_id)

                    res = run_pipeline_to_s3(
                        job_id=job_id,
                        payload=payload_inner,
                        callbacks=callbacks,
                        upload_folder=app.config["UPLOAD_FOLDER"],
                        processed_folder=app.config["PROCESSED_FOLDER"],
                    )

                    set_job(job_id, state="SUCCESS", info=res, project_id=project_id)

                    log_completed_job_row(
                        ays_id=res["ays_id"],
                        from_email=payload_inner.get("From") or "",
                        project_name=payload_inner.get("Subject") or "",
                        manufacturer_terms=res.get("manufacturer_terms"),
                        recommendation=res.get("recommendation"),
                        project_id=res["project_id"],
                        doc_folder=res["doc_folder"],
                        zip_key=res["zip_key"],
                        job_id=job_id,
                        submitted_at=payload_inner.get("SubmittedAt"),
                        meta_fields=payload_inner.get("Meta") or {},
                        pages_processed=res.get("pages_processed"),
                        attachment_name=res.get("display_name"),  # NEW: show on dashboard
                    )
                except Exception as e:
                    logging.exception("Background job failed")
                    set_job(
                        job_id,
                        state="FAILURE",
                        info={"error": str(e)},
                        project_id=payload_inner["ProjectID"],
                    )

            job_id = submit_job(job_fn, payload)
            job_ids.append(job_id)

        # project-level job bundle marker
        set_job(project_id, state="PROJECT", jobs=job_ids, subject=subject)

        wants_json = (
            request.accept_mimetypes.best == 'application/json'
            or request.headers.get('X-Requested-With') == 'XMLHttpRequest'
        )

        return (
            jsonify({'project_id': project_id, 'job_ids': job_ids})
            if wants_json
            else render_template(
                'project_submitted5.html',
                project_id=project_id,
                job_ids=job_ids
            )
        )

    except Exception as e:
        logging.exception("Error in project_process")
        return jsonify({"error": str(e)}), 500



# =========================
# Lightweight status endpoints
# =========================
@app.get('/status/<job_id>')
def job_status(job_id):
    rec = get_job(job_id)
    if not rec:
        return jsonify({"error": "unknown job"}), 404
    return jsonify({
        "job_id": job_id,
        "state": rec.get("state", "QUEUED"),
        "info": rec.get("info", {}),
    })


@app.get("/project-status/<project_id>")
def project_status(project_id):
    rec = get_job(project_id)
    if not rec:
        return jsonify({"error":"unknown project"}), 404
    job_ids = rec.get("jobs", [])
    statuses = []
    for jid in job_ids:
        j = get_job(jid) or {}
        statuses.append({
            "job_id": jid,
            "state": j.get("state"),
            "doc_folder": (j.get("info") or {}).get("doc_folder"),
        })
    return jsonify({
        "project_id": project_id,
        "subject": rec.get("subject"),
        "jobs": statuses,
        # optionally set elsewhere if you add it later:
        "index_url": rec.get("index_url")
    })


# =========================
# Permanent download by job_id (works after restarts)
# =========================
@app.get("/dl/<job_id>")
def dl_always(job_id):
    # 1) Try in-memory job store (if still around)
    rec = get_job(job_id)
    zip_key = None
    if rec and rec.get("state") == "SUCCESS":
        zip_key = (rec.get("info") or {}).get("zip_key")

    # 2) Fallback to dashboard workbook (survives restarts)
    if not zip_key:
        try:
            df = pd.read_excel(DASHBOARD_XLSX, dtype=str).fillna("")
            row = df.loc[df.get("Job ID", "") == job_id]
            if not row.empty:
                zip_key = row.iloc[0].get("S3 Zip Key", "")
        except Exception as e:
            logging.error(f"dl_always: failed reading dashboard for {job_id}: {e}")

    if not zip_key:
        return jsonify({"error": "unknown or incomplete job"}), 404

    # 3) Presign fresh every click
    url = s3_presign_get(zip_key, expires=3600)  # 1-hour URL, refreshed each click
    return redirect(url, code=302)


# =========================
# Chat/email-style Explorer
# =========================

@app.get("/api/explorer")
def api_explorer():
    """
    Returns:
      - For ?path=results/ : a list of projects (the 'root' of the explorer)
      - For ?path=results/<PROJECT_ID>/ or deeper: directory + file listing under that prefix
    """
    raw = (request.args.get("path") or "").strip()
    path = "results/" if raw in ("", "/") else raw
    if not path.endswith("/"):
        path += "/"

    # Root: list projects (newest first is handled in list_projects_from_dashboard)
    if path.lower() == "results/":
        projects = list_projects_from_dashboard()  # [{'project_id','project_name','email','date','sort_key'}, ...]
        items = [{
            "type": "project",
            "name": p.get("project_name") or p["project_id"].split("_AYS-")[0].replace("_", " "),
            "email": p.get("email", ""),
            "date": p.get("date", ""),
            "key":  f"results/{p['project_id']}/",
        } for p in projects]
        return jsonify({"path": "results/", "items": items})

    # Children: delegate to S3 lister (should return {"path":..., "items":[{"type":"dir"/"file", ...}]})
    return jsonify(s3_list_dir(path))


from helpers.helpers_async_s3_0_9 import (
    list_projects_from_dashboard,
    get_project_meta,
    list_project_docs,
    s3_presign_get,
    S3_RESULTS_PREFIX,
    S3_UPLOAD_PREFIX,
    generate_email_body,
    s3_list_dir,
    update_project_meta_row,   # <-- NEW
)

# --- Flask app: Explorer endpoints (REPLACE) ---
from flask import jsonify, render_template, redirect, abort
import urllib.parse
import pandas as pd

# Page
@app.route('/dashboard')
def explorer_page():
    return render_template('explorer_chat_email15.html')  # use the old look template name

from data.bid_status_store import get_all_bid_status

@app.get('/api/explorer/projects')
def api_explorer_projects():
    items = list_projects_from_dashboard()

    # Load once (fast) so we don't lock/read the json file for every project row
    bid_map = get_all_bid_status()

    # Attach status to each project item
    for p in items:
        pid = str(p.get("project_id", "")).strip()
        rec = bid_map.get(pid) or {}
        p["bid_status"] = rec.get("bid_status", "")
        p["bid_updated_by"] = rec.get("updated_by", "")
        p["bid_updated_at"] = rec.get("updated_at", "")

    return jsonify(items)

# Docs for a project, with preview (inline) and attachments (download)
@app.get('/api/explorer/<project_id>/docs')
def api_explorer_docs(project_id):
    """
    Return:
      {
        "project": {
           "project_id": ...,
           "project_name": ...,
           "email": ...,
           "date": ...,
           "meta_bid_date": ...,
           "meta_drawing_date": ...,
           "meta_address": ...,
           "meta_engineer": ...,
           "meta_general_contractor": ...,
           "meta_notes": ...
        },
        "docs": [ ... ]
      }
    This mirrors the meta fields used by /confirm-send.
    """
    # ----- 1) Build base project meta (fallbacks) -----
    meta = {
        "project_id": project_id,
        "project_name": project_id.split("_AYS-")[0].replace("_", " "),
        "email": "",
        "date": "",
        "meta_bid_date": "",
        "meta_drawing_date": "",
        "meta_address": "",
        "meta_engineer": "",
        "meta_general_contractor": "",
        "meta_notes": "",
    }

    # ----- 2) Hydrate from DASHBOARD_XLSX, same as /confirm-send -----
    try:
        if project_id and os.path.isfile(DASHBOARD_XLSX):
            df = pd.read_excel(DASHBOARD_XLSX, dtype=str).fillna("")

            if "Project ID" in df.columns:
                df = df[df["Project ID"].astype(str) == str(project_id)]

            if not df.empty:
                row = df.tail(1).iloc[0]

                meta["project_name"] = row.get("Project Name", meta["project_name"])
                meta["email"]       = row.get("Email", "")
                meta["date"]        = row.get("Date", "")

                # same column → meta_* mapping as /confirm-send
                meta["meta_bid_date"]           = row.get("Bid Date", "")
                meta["meta_drawing_date"]       = row.get("Drawing Date", "")
                meta["meta_address"]            = row.get("Address", "")
                meta["meta_engineer"]           = row.get("Engineer", "")
                meta["meta_general_contractor"] = row.get("General Contractor", "")
                meta["meta_notes"]              = row.get("Notes", "")
    except Exception:
        logging.exception("api_explorer_docs: failed to read dashboard workbook")

    # ----- 3) List docs from S3 (unchanged logic) -----
    raw_docs = list_project_docs(project_id)  # [{doc_folder, artifacts:{...}, ...}]

    def wrap_dl(key):
        return f"/get?key={urllib.parse.quote(key)}" if key else None

    def wrap_view(key):
        return f"/view/by-key?key={urllib.parse.quote(key)}" if key else None

    docs = []
    for d in raw_docs:
        folder = d["doc_folder"]
        a = d.get("artifacts") or {}

        # Optional extra info from list_project_docs, if present
        submitted_at_display = d.get("submitted_at_display") or d.get("submitted_at") or ""
        attachment_name      = d.get("attachment_name") or folder

        # Attachments
        attachments = []
        if a.get("zip"):
            attachments.append({"label": "Report ZIP", "href": wrap_dl(a["zip"]), "ext": "zip"})
        if a.get("tables"):
            attachments.append({"label": "Tables.xlsx", "href": wrap_dl(a["tables"]), "ext": "xlsx"})
        if a.get("highlighted"):
            attachments.append({"label": "Highlighted.pdf", "href": wrap_dl(a["highlighted"]), "ext": "pdf"})
        if a.get("only_highlights"):
            attachments.append({"label": "Highlights-only.pdf", "href": wrap_dl(a["only_highlights"]), "ext": "pdf"})
        if a.get("email_pdf"):
            attachments.append({"label": "Email Summary.pdf", "href": wrap_dl(a["email_pdf"]), "ext": "pdf"})

        docs.append({
            "doc_folder": folder,
            "attachment_name": attachment_name,
            "submitted_at_display": submitted_at_display,
            "subject": f"{meta['project_name'] or 'Project'} — {folder} summary",
            "email_meta": {
                "from_name": AYS_FROM_NAME,
                "from_email": AYS_FROM_EMAIL,
                "to_email": meta.get("email", ""),
                "date": meta.get("date", ""),
            },
            "preview": {
                "html": wrap_view(a.get("email_html")),
                "pdf":  wrap_view(a.get("email_pdf")),
            },
            "attachments": attachments,
        })
        
    bid_rec = get_bid_status(project_id) or {}
    meta["bid_status"] = bid_rec.get("bid_status", "")
    meta["bid_updated_by"] = bid_rec.get("updated_by", "")
    meta["bid_updated_at"] = bid_rec.get("updated_at", "")

    return jsonify({"project": meta, "docs": docs})


# Always-inline view (PDF/HTML) for iframe
@app.get("/view/by-key")
def view_by_key():
    """
    302 to a presigned S3 URL and, if ?page=N is provided, append #page=N so the PDF viewer jumps to that page.
    This route must be called with absolute origin from the email HTML (APP_PUBLIC_BASE).
    """
    key = (request.args.get("key") or "").lstrip("/")
    page = (request.args.get("page") or "").strip()

    if not key or ".." in key:
        return jsonify({"error": "bad key"}), 400

    # Pick a reasonable content type when we know it's a PDF/HTML
    low = key.lower()
    if low.endswith(".pdf"):
        ct = "application/pdf"
    elif low.endswith(".html") or low.endswith(".htm"):
        ct = "text/html; charset=utf-8"
    else:
        ct = None

    extra = {}
    if ct:
        extra["ResponseContentType"] = ct
        # important: inline so it opens in browser
        extra["ResponseContentDisposition"] = "inline"

    try:
        url = s3_presign_get(key, expires=3600, extra=extra)
        if page.isdigit():
            # append fragment so the PDF viewer opens to that page
            url = f"{url}#page={page}"
        return redirect(url, code=302)
    except Exception:
        logging.exception("presign failed")
        return jsonify({"error": "presign failed"}), 500



    
@app.get("/confirm-send")
def job_confirm_send():
    """
    Loads the confirm_send_popup.html form in an iframe so the user
    can review/edit the job info, manufacturer terms, and project meta
    (bid date, address, etc.) for a single doc.
    """
    project_id = (request.args.get("project_id") or "").strip()
    doc_folder = (request.args.get("doc_folder") or "").strip()

    # Base context with all fields your popup template can use
    ctx = {
        "project_id": project_id,
        "doc_folder": doc_folder,
        "project_name": "",
        "email": "",
        "date": "",
        "ays_id": "",
        "manufacturer_terms": "",
        "recommendation": "",
        # NEW: meta_* fields for project details
        "meta_bid_date": "",
        "meta_drawing_date": "",
        "meta_address": "",
        "meta_engineer": "",
        "meta_general_contractor": "",
        "meta_notes": "",
    }

    try:
        if project_id and os.path.isfile(DASHBOARD_XLSX):
            df = pd.read_excel(DASHBOARD_XLSX, dtype=str).fillna("")

            # filter by project
            if "Project ID" in df.columns:
                df = df[df["Project ID"].astype(str) == project_id]

            # further filter by doc folder if present
            if doc_folder and "Doc Folder" in df.columns:
                df = df[df["Doc Folder"].astype(str) == doc_folder]

            if not df.empty:
                # last matching row for this project/doc_folder
                row = df.tail(1).iloc[0]

                ctx.update(
                    {
                        "project_name": row.get("Project Name", ""),
                        "email": row.get("Email", ""),
                        "date": row.get("Date", ""),
                        "ays_id": row.get("AYS ID", ""),
                        "manufacturer_terms": row.get("Manufacturer Terms", ""),
                        "recommendation": row.get("Recommendation", ""),
                        # Map Excel columns → template variables (meta_*)
                        # These columns will exist once you log them from log_completed_job_row
                        "meta_bid_date": row.get("Bid Date", ""),
                        "meta_drawing_date": row.get("Drawing Date", ""),
                        "meta_address": row.get("Address", ""),
                        "meta_engineer": row.get("Engineer", ""),
                        "meta_general_contractor": row.get("General Contractor", ""),
                        "meta_notes": row.get("Notes", ""),
                    }
                )
    except Exception:
        logging.exception("Confirm and Send: failed to read dashboard workbook")

    return render_template("confirm_send_popup.html", **ctx)



# Generic download by exact S3 key (used by attachment buttons)
@app.get('/get')
def get_by_key():
    key = (request.args.get("key") or "").strip()
    if not key:
        abort(400, "missing key")
    if not (key.startswith(f"{S3_RESULTS_PREFIX}/") or key.startswith(f"{S3_UPLOAD_PREFIX}/")):
        abort(403, "forbidden key")
    url = s3_presign_get(key, expires=3600)
    return redirect(url, code=302)



# (Optional helper for generic download from explorer tables/grids)
@app.get("/dl/by-key")
def dl_by_key():
    key = (request.args.get("key") or "").lstrip("/")
    if not key or ".." in key:
        return jsonify({"error": "bad key"}), 400
    try:
        url = s3_presign_get(key, expires=3600)
        return redirect(url, code=302)
    except Exception:
        logging.exception("presign failed")
        return jsonify({"error": "presign failed"}), 500

@app.get("/section/dl")
def section_dl_by_key():
    """
    Dynamically slice a section from the highlighted PDF in S3 and download it.

    Query params:
      - key:   S3 key of the highlighted PDF (must be under results/)
      - start: 1-based start page (inclusive)
      - stop:  1-based stop page (inclusive)
    """
    key = (request.args.get("key") or "").lstrip("/")
    start = request.args.get("start", type=int)
    stop  = request.args.get("stop",  type=int)

    if not key or ".." in key:
        return jsonify({"error": "bad key"}), 400
    if not key.startswith(f"{S3_RESULTS_PREFIX}/"):
        # Only allow section slicing on results artifacts
        return jsonify({"error": "forbidden key"}), 403

    if not start or start < 1:
        return jsonify({"error": "invalid start page"}), 400
    if not stop or stop < start:
        stop = start  # treat as single-page section if stop is missing/bad

    try:
        # Fetch highlighted PDF from S3
        obj = _s3.get_object(Bucket=S3_BUCKET, Key=key)
        pdf_bytes = obj["Body"].read()

        reader = PdfReader(io.BytesIO(pdf_bytes))
        writer = PdfWriter()

        num_pages = len(reader.pages)
        s = max(1, start)
        e = min(stop, num_pages)

        for p in range(s - 1, e):
            writer.add_page(reader.pages[p])

        buf = io.BytesIO()
        writer.write(buf)
        buf.seek(0)

        filename = f"Section_p{s}_to_p{e}.pdf"
        return send_file(
            buf,
            mimetype="application/pdf",
            as_attachment=True,
            download_name=filename,
        )

    except Exception as e:
        logging.exception("section_dl_by_key failed")
        return jsonify({"error": "failed to slice section"}), 500


@app.post("/project/<project_id>/update-meta")
def project_update_meta(project_id):
    """
    Update a project's job name + meta fields (address, engineer, etc.)
    across all rows for that Project ID in the dashboard.

    Accepts either JSON or form-encoded data.

    Expected keys:
      - Subject or project_name  (new Job Name)
      - Email or email
      - any number of meta_* fields (meta_bid_date, meta_address, meta_engineer, ...)
    """
    try:
        if request.is_json:
            data = request.get_json() or {}
            project_name = data.get("Subject") or data.get("project_name")
            email = data.get("Email") or data.get("email")
            meta_fields = {k: v for k, v in data.items() if k.startswith("meta_")}
        else:
            project_name = request.form.get("Subject") or request.form.get("project_name")
            email = request.form.get("Email") or request.form.get("email")
            meta_fields = {k: v for k, v in request.form.items() if k.startswith("meta_")}

        updated = update_project_meta_row(
            project_id=project_id,
            project_name=project_name,
            email=email,
            meta_fields=meta_fields,
        )

        # Also keep the in-memory job record subject in sync, if present
        rec = get_job(project_id)
        if rec and project_name:
            set_job(project_id, subject=project_name)

        if not updated:
            return jsonify({"ok": False, "message": "No rows updated for this Project ID"}), 404

        return jsonify({"ok": True, "project_id": project_id})

    except Exception as e:
        logging.exception("project_update_meta failed")
        return jsonify({"ok": False, "error": str(e)}), 500

@app.get("/project/<project_id>/edit")
def edit_project(project_id):
    """
    Load an existing project's info + meta fields from the dashboard
    and render an edit form (reusing project_submitted3.html).
    """
    ctx = {
        "project_id": project_id,
        "project_name": "",
        "email": "",
        "date": "",
        # same meta_* names you already use elsewhere
        "meta_bid_date": "",
        "meta_drawing_date": "",
        "meta_address": "",
        "meta_engineer": "",
        "meta_general_contractor": "",
        "meta_notes": "",
        "mode": "edit",   # let the template know this is EDIT mode
    }

    try:
        if os.path.isfile(DASHBOARD_XLSX):
            df = pd.read_excel(DASHBOARD_XLSX, dtype=str).fillna("")
            if "Project ID" in df.columns:
                rows = df[df["Project ID"].astype(str) == str(project_id)]
                if not rows.empty:
                    row = rows.tail(1).iloc[0]

                    ctx["project_name"] = row.get("Project Name", "")
                    ctx["email"] = row.get("Email", "")
                    ctx["date"] = row.get("Date", "")

                    # map visible columns -> meta_* fields
                    ctx["meta_bid_date"] = row.get("Bid Date", "")
                    ctx["meta_drawing_date"] = row.get("Drawing Date", "")
                    ctx["meta_address"] = row.get("Address", "")
                    ctx["meta_engineer"] = row.get("Engineer", "")
                    ctx["meta_general_contractor"] = row.get("General Contractor", "")
                    ctx["meta_notes"] = row.get("Notes", "")
    except Exception:
        logging.exception("edit_project: failed to read dashboard workbook")

    return render_template("project_submitted5.html", **ctx)


from helpers.helpers_async_s3_0_9 import get_next_available_name_local

@app.post("/project/<project_id>/add-doc")
def add_doc_to_project(project_id):
    """
    Upload one or more PDFs into an existing project.
    This mirrors /project-process so:
      - duplicate filenames get auto-renamed
      - pages get counted
      - dashboard logs correctly
    """
    try:
        files = request.files.getlist("files")
        if not files:
            return jsonify({"error": "No files uploaded"}), 400

        # Reuse/override project details from form
        subject = request.form.get("Subject") or "Untitled_Project"
        email   = request.form.get("From") or "unknown@example.com"

        # Project metadata
        meta_fields = {
            k: v for k, v in request.form.items()
            if k.startswith("meta_")
        }

        job_ids = []

        callbacks = dict(
            process_pdf_file=process_pdf_file,
            create_highlighted_only_pdf=create_highlighted_only_pdf,
            generate_email_body=generate_email_body,
            logo_base64=LOGO_BASE64,
            write_results_to_excel=write_results_to_excel,
        )

        submitted_at_iso = datetime.utcnow().isoformat(timespec="seconds") + "Z"

        for file in files:
            if not file or not file.filename:
                continue

            # --------------------------
            # SAME NAMING LOGIC AS /project-process
            # --------------------------
            original = file.filename
            base, ext = os.path.splitext(original)
            if not ext:
                ext = ".pdf"

            safe_base = re.sub(r'[^A-Za-z0-9._ -]+', '_', base).strip() or "file"

            unique_name = get_next_available_name_local(
                safe_base,
                ext,
                app.config["UPLOAD_FOLDER"],
            )
            # --------------------------

            payload = {
                "AttachmentName": unique_name,
                "DisplayName": unique_name,
                "AttachmentContent": base64.b64encode(file.read()).decode("utf-8"),
                "From": email,
                "Subject": subject,
                "MessageID": None,
                "ProjectID": project_id,
                "SubmittedAt": submitted_at_iso,
                "Meta": meta_fields,
            }

            def job_fn(job_id, payload_inner):
                try:
                    set_job(job_id, state="STARTED", project_id=project_id)

                    res = run_pipeline_to_s3(
                        job_id=job_id,
                        payload=payload_inner,
                        callbacks=callbacks,
                        upload_folder=app.config["UPLOAD_FOLDER"],
                        processed_folder=app.config["PROCESSED_FOLDER"],
                    )

                    set_job(job_id, state="SUCCESS", info=res, project_id=project_id)

                    log_completed_job_row(
                        ays_id=res["ays_id"],
                        from_email=payload_inner.get("From") or "",
                        project_name=payload_inner.get("Subject") or "",
                        manufacturer_terms=res.get("manufacturer_terms"),
                        recommendation=res.get("recommendation"),
                        project_id=res["project_id"],
                        doc_folder=res["doc_folder"],
                        zip_key=res["zip_key"],
                        job_id=job_id,
                        submitted_at=payload_inner.get("SubmittedAt"),
                        meta_fields=payload_inner.get("Meta") or {},
                        pages_processed=res.get("pages_processed"),
                        attachment_name=res.get("display_name"),  # NEW
                    )
                except Exception as e:
                    logging.exception("Background job failed")
                    set_job(
                        job_id,
                        state="FAILURE",
                        info={"error": str(e)},
                        project_id=project_id,
                    )

            job_id = submit_job(job_fn, payload)
            job_ids.append(job_id)

        # Mark project as having new jobs
        set_job(project_id, state="PROJECT", jobs=job_ids)

        return jsonify({"ok": True, "project_id": project_id, "job_ids": job_ids})

    except Exception as e:
        logging.exception("Error in add_doc_to_project")
        return jsonify({"error": str(e)}), 500



from helpers.helpers_async_s3_0_9 import delete_project_doc_s3

from flask import jsonify, current_app  # if not already imported

@app.post("/api/project/<project_id>/docs/<doc_folder>/delete")
def api_delete_project_doc(project_id, doc_folder):
    """
    Delete a project's document folder from S3:
      results/<project_id>/<doc_folder>/...

    Called by the Explorer UI when the user clicks the trash icon
    on a document row.
    """
    try:
        result = delete_project_doc_s3(project_id, doc_folder)
        return jsonify({
            "ok": True,
            "project_id": project_id,
            "doc_folder": doc_folder,
            **result,
        })
    except Exception as exc:
        current_app.logger.exception(
            "api_delete_project_doc: failed for project_id=%s doc_folder=%s",
            project_id, doc_folder,
        )
        return jsonify({
            "ok": False,
            "project_id": project_id,
            "doc_folder": doc_folder,
            "error": str(exc),
        }), 500
    
@app.route("/api/usage_stats")
def api_usage_stats():
    try:
        stats = get_usage_stats()
        return jsonify(stats)
    except Exception as e:
        current_app.logger.exception("Usage stats failed")
        return jsonify({"error": "failed to compute usage stats"}), 500

from flask import request, jsonify
from data.bid_status_store import set_bid_status, get_bid_status

@app.route("/api/bid-status", methods=["POST"])
def api_set_bid_status():
    payload = request.get_json(force=True) or {}

    job_id = str(payload.get("job_id", "")).strip()
    bid_status = str(payload.get("bid_status", "")).strip()
    project_name = str(payload.get("project_name", "")).strip()

    if not job_id:
        return jsonify({"success": False, "error": "Missing job_id"}), 400

    # Pick the best "who" you have available.
    # If you have a login system, replace this with your user identity.
    updated_by = session.get("user") or payload.get("updated_by") or request.headers.get("X-User") or "user"


    record = set_bid_status(
        job_id=job_id,
        project_name=project_name,
        bid_status=bid_status,
        updated_by=str(updated_by),
    )

    return jsonify({"success": True, "record": record})


@app.route("/api/bid-status/<job_id>", methods=["GET"])
def api_get_bid_status(job_id):
    rec = get_bid_status(str(job_id).strip())
    return jsonify({"success": True, "record": rec})

@app.get("/api/version")
def api_version():
    return jsonify({"version": os.environ.get("AYS_VERSION", "3.4.27")})

# =========================
# CONFIG UPLOAD + ROUTER
# =========================

from io import BytesIO, StringIO
import csv
from flask import request, send_file, Response, render_template, flash, redirect, url_for
import openpyxl

from parsers.kcc_parser1 import convert_kcc_pdf_to_xlsx_bytes
from parsers.valent_parser1 import convert_valent_pdf_to_xlsx_bytes
from parsers.innovent_parser2 import convert_innovent_pdf_to_xlsx_bytes
from parsers.multistack_parser1 import convert_multistack_pdf_to_xlsx_bytes
from parsers.aaon_parser1 import convert_aaon_pdf_to_xlsx_bytes
from parsers.superior_parser1 import convert_superior_to_xlsx_bytes
from parsers.weishaupt_parser1 import convert_weishaupt_to_xlsx_bytes
from parsers.daikin_parser import convert_daikin_to_xlsx_bytes

CONFIG_CONVERTERS = {
    "kcc": convert_kcc_pdf_to_xlsx_bytes,
    "valent": convert_valent_pdf_to_xlsx_bytes,
    "innovent": convert_innovent_pdf_to_xlsx_bytes,
    "multistack": convert_multistack_pdf_to_xlsx_bytes,
    "aaon": convert_aaon_pdf_to_xlsx_bytes,
    "superior": convert_superior_to_xlsx_bytes,
    "weishaupt": convert_weishaupt_to_xlsx_bytes,
    "daikin": convert_daikin_to_xlsx_bytes,
}


def _xlsx_bytes_to_csv(xlsx_bytes: bytes) -> bytes:
    """Convert the first sheet of an xlsx workbook to CSV bytes."""
    wb = openpyxl.load_workbook(BytesIO(xlsx_bytes), data_only=True)
    ws = wb.active

    buf = StringIO()
    writer = csv.writer(buf)
    for row in ws.iter_rows(values_only=True):
        writer.writerow(['' if v is None else v for v in row])

    return buf.getvalue().encode('utf-8')


@app.route("/config", methods=["GET"])
def config_page():
    return render_template("config_upload2.html")


@app.route("/config", methods=["POST"])
def config_router():
    manufacturer = (request.form.get("manufacturer") or "").strip().lower()

    if manufacturer in CONFIG_CONVERTERS:
        return config_run()

    flash(f"'{manufacturer or 'unknown'}' is not implemented yet.", "error")
    return redirect(url_for("config_page"))


@app.route("/config/run", methods=["POST"])
def config_run():
    app.logger.info(
        "CONFIG/RUN HIT ✅ method=%s content_type=%s",
        request.method,
        request.content_type
    )
    app.logger.info(
        "CONFIG/RUN form keys=%s files=%s",
        list(request.form.keys()),
        list(request.files.keys())
    )

    try:
        job_name     = (request.form.get("job_name") or "").strip()
        manufacturer = (request.form.get("manufacturer") or "").strip().lower()
        output_type  = (request.form.get("output_type") or "all_in_one").strip().lower()
        pdf          = request.files.get("pdf_file")

        app.logger.info(
            "CONFIG/RUN parsed job_name=%r manufacturer=%r output_type=%r pdf=%s",
            job_name, manufacturer, output_type, getattr(pdf, "filename", None)
        )

        if manufacturer not in CONFIG_CONVERTERS:
            app.logger.warning("CONFIG/RUN manufacturer not supported: %r", manufacturer)
            return Response(f"Manufacturer '{manufacturer}' not supported.", status=400)

        if output_type not in ("all_in_one", "shopping_list"):
            app.logger.warning("CONFIG/RUN unsupported output_type: %r", output_type)
            return Response("Unsupported output type.", status=400)

        if not pdf or not pdf.filename:
            app.logger.warning("CONFIG/RUN missing pdf_file upload")
            return Response("Missing PDF upload.", status=400)

        pdf_bytes = pdf.read()
        app.logger.info("CONFIG/RUN read pdf bytes=%d", len(pdf_bytes))

        converter_fn = CONFIG_CONVERTERS[manufacturer]
        app.logger.info("CONFIG/RUN calling converter=%s...", getattr(converter_fn, "__name__", str(converter_fn)))

        xlsx_bytes, filename_from_func = converter_fn(
            pdf_bytes=pdf_bytes,
            job_name=job_name,
            output_type=output_type,
        )

        if hasattr(xlsx_bytes, "getvalue"):
            xlsx_bytes = xlsx_bytes.getvalue()

        app.logger.info(
            "CONFIG/RUN convert returned bytes=%d filename=%s",
            len(xlsx_bytes), filename_from_func
        )

        # Convert xlsx → csv
        csv_bytes = _xlsx_bytes_to_csv(xlsx_bytes)

        safe_job = "".join(c for c in job_name if c.isalnum() or c in (" ", "-", "_")).strip() or "job"
        filename = f"{safe_job}_{manufacturer}_{output_type}_template_output.csv".replace(" ", "_")

        out = BytesIO(csv_bytes)
        out.seek(0)

        resp = send_file(
            out,
            as_attachment=True,
            download_name=filename,
            mimetype="text/csv",
            max_age=0
        )
        resp.headers["Cache-Control"] = "no-store"
        resp.headers["X-AYS-Config"] = manufacturer
        resp.headers["X-AYS-OutputType"] = output_type

        app.logger.info("CONFIG/RUN returning download filename=%s", filename)
        return resp

    except Exception:
        app.logger.exception("CONFIG/RUN FAILED")
        return Response("Conversion failed (see server logs).", status=500)


from routes.routes_schedule import schedule_bp
app.register_blueprint(schedule_bp)


if __name__ == "__main__":
    os.makedirs(app.config['UPLOAD_FOLDER'], exist_ok=True)
    os.makedirs(app.config['PROCESSED_FOLDER'], exist_ok=True)
    app.run(host='0.0.0.0', port=8000)


