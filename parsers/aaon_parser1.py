# aaon_parser1.py  (DROP-IN MODULE)
from __future__ import annotations

import io
import re
import logging
from typing import Any, Optional

import fitz  # PyMuPDF

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

logger = logging.getLogger(__name__)


# ============================================================
# OCR (optional fallback)
#   - Works only if pytesseract + pdf2image + system deps exist
# ============================================================

def _ocr_pdf_to_text(pdf_bytes: bytes) -> str:
    """
    OCR fallback. Requires:
      - apt: tesseract-ocr, poppler-utils
      - pip: pytesseract, pdf2image, pillow
    """
    try:
        from pdf2image import convert_from_bytes
        import pytesseract
    except Exception as e:
        raise RuntimeError(
            "OCR fallback required but OCR libs not installed. "
            "Install: pip install pytesseract pdf2image pillow "
            "and apt install tesseract-ocr poppler-utils"
        ) from e

    images = convert_from_bytes(pdf_bytes, dpi=300)
    pages: list[str] = []
    for img in images:
        # psm 6 = assume a uniform block of text (good for many quote forms)
        txt = pytesseract.image_to_string(img, config="--psm 6")
        pages.append(txt or "")
    return "\n\n".join(pages).strip()


# ============================================================
# Helpers
# ============================================================

def _norm(s: str) -> str:
    return re.sub(r"\s+", " ", (s or "")).strip()

def _safe_sheet_name(name: str) -> str:
    bad = r'[:\\/?*\[\]]'
    name = re.sub(bad, "-", (name or "")).strip()
    return (name[:31] or "Sheet1")

def _looks_like_empty_text(text: str) -> bool:
    """
    Heuristic: treat as "empty" if it has very few alphanumerics.
    Image PDFs often return empty or junk.
    """
    t = text or ""
    alnum = sum(ch.isalnum() for ch in t)
    # tune threshold as needed
    return alnum < 80

def _extract_first(pattern: str, text: str, flags: int = 0) -> str:
    m = re.search(pattern, text or "", flags)
    return m.group(1).strip() if m else ""

def _extract_model(text: str) -> str:
    # You gave example: RNA-015-B-A-3-HJB0C-A03NA
    # Keep this broad: AAON model strings are usually uppercase with dashes.
    m = re.search(r"\b([A-Z]{2,5}-[A-Z0-9]{2,}(?:-[A-Z0-9]{1,})+)\b", text or "")
    return m.group(1).strip() if m else ""

def _extract_tag(text: str) -> str:
    # Common patterns: "Tag: XYZ" or "Tag XYZ" or header columns "Tag" then value below.
    tag = _extract_first(r"\bTag\s*[:#]?\s*([A-Za-z0-9_\-\/]+)\b", text, flags=re.IGNORECASE)
    if tag:
        return tag

    # Try table-ish pattern: line contains "Tag" and next non-empty token line is value
    lines = [ln.strip() for ln in (text or "").splitlines()]
    for i, ln in enumerate(lines):
        if re.fullmatch(r"Tag", ln, flags=re.IGNORECASE):
            # scan forward
            for j in range(i + 1, min(i + 6, len(lines))):
                v = lines[j].strip()
                if v and v.lower() != "qty" and v.lower() != "model":
                    # first token
                    return v.split()[0]
    return ""

def _extract_equipment(text: str) -> str:
    """
    You asked: Equipment (can you find it?).
    We try to infer from AAON doc keywords. If not found, default to RTU.
    """
    t = (text or "").lower()
    # RN Series is a rooftop (often DOAS/heat pump variants)
    if "rn series" in t:
        return "RTU"
    if "doas" in t:
        return "DOAS"
    if "make up air" in t or "mua" in t:
        return "Make Up Air"
    if "air handler" in t:
        return "Air Handler"
    return "RTU"

def _extract_quote_number(text: str) -> str:
    # Many AAON quotes include "Quote" or "Quotation" or "Proposal"
    q = _extract_first(r"\bQuote\s*(?:No\.?|#)?\s*[:#]?\s*([A-Za-z0-9\-]+)", text, flags=re.IGNORECASE)
    if q:
        return q
    q2 = _extract_first(r"\bQuotation\s*(?:No\.?|#)?\s*[:#]?\s*([A-Za-z0-9\-]+)", text, flags=re.IGNORECASE)
    return q2

def _extract_part_number_after_model(text: str, model: str) -> str:
    """
    Requirement: Part Number = "Everything after the model number" (if present).
    If we find a line that contains the model and additional trailing tokens,
    return that trailing remainder. Otherwise blank.
    """
    if not model:
        return ""

    for ln in (text or "").splitlines():
        if model in ln:
            idx = ln.find(model)
            tail = ln[idx + len(model):].strip()
            # Clean separators
            tail = re.sub(r"^[\s:\-–|]+", "", tail).strip()
            # If tail is just punctuation or empty, ignore
            if tail and re.search(r"[A-Za-z0-9]", tail):
                return tail
    return ""

def _extract_key_value_options(text: str, model: str) -> list[str]:
    """
    Build description lines like:
      RN Series
      Unit Size Fifteen
      Voltage 460V/3q/60Hz
      ...

    This tries to extract "Label  Value" lines robustly (works for OCR too).

    Strategy:
      - Take lines after the model appears (if found), else whole doc.
      - Keep lines that look like "some words ... some value"
      - Remove obvious pricing totals / $ lines
    """
    raw_lines = [ln.rstrip() for ln in (text or "").splitlines()]
    lines = [ln.strip() for ln in raw_lines if ln.strip()]

    # Start after model line if possible
    start_idx = 0
    if model:
        for i, ln in enumerate(lines):
            if model in ln:
                start_idx = i
                break
    lines = lines[start_idx:]

    out: list[str] = []
    for ln in lines:
        # skip money/prices
        if "$" in ln:
            continue
        if re.search(r"\b(total|subtotal|tax|freight|price)\b", ln, flags=re.IGNORECASE):
            # AAON docs sometimes have a big pricing section; exclude most of it
            continue

        # keep "RN Series" etc
        if re.fullmatch(r"[A-Za-z0-9][A-Za-z0-9 \-\/&]+", ln) and len(ln) <= 80:
            # We'll keep short-ish meaningful lines
            pass

        # Try to convert "Label: Value" to "Label Value"
        ln = re.sub(r"\s*:\s*", " ", ln).strip()

        # Many OCR layouts come in with big spaces between columns; normalize
        ln = re.sub(r"\s{2,}", " ", ln).strip()

        # Avoid pure headers
        if ln.lower() in ("tag", "qty", "model", "part number", "options", "features"):
            continue

        # Keep lines that have at least 2 words OR look like "Voltage 460V/..."
        if len(ln.split()) >= 2:
            out.append(ln)

    # De-dupe while preserving order
    seen = set()
    uniq: list[str] = []
    for x in out:
        k = _norm(x).lower()
        if k in seen:
            continue
        seen.add(k)
        uniq.append(x)

    return uniq


# ============================================================
# Main parse entry (PDF bytes -> structured dict)
# ============================================================

def parse_aaon_pdf_bytes(pdf_bytes: bytes) -> dict[str, Any]:
    """
    Returns the same schema as your other converters:
      {
        "manufacturer": "AAON",
        "header": {...},
        "lines": [
          {
            "model_code": "...",
            "tagging": "...",
            "tags": [...],
            "qty": 1,
            "options": [
              {"feature": "Equipment", "desc": "...", "add_price": 0.0},
              ... (optional include rows)
            ]
          }
        ]
      }
    """

    # 1) Attempt standard PDF text extraction
    doc = fitz.open(stream=pdf_bytes, filetype="pdf")
    text_pages: list[str] = []
    for i in range(len(doc)):
        try:
            text_pages.append(doc[i].get_text("text") or "")
        except Exception:
            text_pages.append("")
    text = "\n\n".join(text_pages).strip()

    used_ocr = False
    if _looks_like_empty_text(text):
        # 2) OCR fallback
        logger.info("AAON: text extraction looked empty; falling back to OCR")
        text = _ocr_pdf_to_text(pdf_bytes)
        used_ocr = True

    # Extract fields
    model = _extract_model(text)
    tag = _extract_tag(text)
    equipment = _extract_equipment(text)
    quote_no = _extract_quote_number(text)
    part_number = _extract_part_number_after_model(text, model)

    # Options/description lines
    option_lines = _extract_key_value_options(text, model=model)

    # Build description: one per line (as you like)
    desc = "\n".join(option_lines).strip()

    header = {
        "quote_number": quote_no,
        "used_ocr": used_ocr,
        "part_number": part_number,
    }

    # Build “options” payload for writer
    options: list[dict[str, Any]] = []
    options.append({
        "feature": "Equipment",
        "desc": desc or _norm(f"{model}").strip(),
        "add_price": 0.0,
    })

    # In this generic AAON parser we treat every option line as an Include in shopping_list mode,
    # but for all_in_one we keep them in E2 only.
    # We'll still store them as include-like rows in case you want to use them later.
    for ln in option_lines:
        options.append({
            "feature": "Include",
            "desc": ln,
            "add_price": 0.0,
        })

    return {
        "manufacturer": "AAON",
        "header": header,
        "lines": [{
            "model_code": model,
            "tagging": tag,
            "tags": [tag] if tag else [""],
            "qty": 1,
            "part_number": part_number,
            "equipment": equipment,
            "options": options,
        }]
    }


# ============================================================
# Workbook writer (same vibe as your others)
# ============================================================

def _write_one_sheet(
    ws,
    *,
    manufacturer: str,
    equipment: str,
    model_code: str,
    part_number: str,
    tags: list[str],
    header: dict[str, Any],
    options: list[dict[str, Any]],
    output_type: str = "all_in_one",
) -> None:
    output_type = (output_type or "all_in_one").strip().lower()

    # Row 1 headers A-F
    top_headers = [
        "Equipment", "Manufacturer", "Model", "Part Number",
        "Description (Not Overwritten)", "Notes (Not Overwritten)"
    ]
    for col, h in enumerate(top_headers, start=1):
        ws.cell(row=1, column=col, value=h)

    # Row 2 values
    ws.cell(row=2, column=1, value=equipment or "")
    ws.cell(row=2, column=2, value=manufacturer or "")
    ws.cell(row=2, column=3, value=model_code or "")
    ws.cell(row=2, column=4, value=part_number or "")

    tags_label = ", ".join(t for t in (tags or []) if t).strip()

    # E2: all_in_one gets the Equipment description only (clean), shopping_list can be lighter
    main_desc = ""
    include_descs: list[str] = []
    for opt in (options or []):
        if not isinstance(opt, dict):
            continue
        feat = (opt.get("feature") or "").strip().lower()
        desc = (opt.get("desc") or "").strip()
        if not desc:
            continue
        if feat == "equipment" and not main_desc:
            main_desc = desc
        elif feat == "include":
            include_descs.append(desc)

    if output_type == "all_in_one":
        e2_parts: list[str] = []
        # Keep tag visible but not in notes unless you want it there too
        if tags_label:
            e2_parts.append(f"Tag: {tags_label}")
            e2_parts.append("")
        if main_desc:
            e2_parts.append(main_desc)
        ws.cell(row=2, column=5, value="\n".join(e2_parts).strip())
    else:
        # shopping_list: can keep E2 simpler
        ws.cell(row=2, column=5, value=(f"Tag: {tags_label}" if tags_label else ""))

    # F2 Notes: quote number + OCR status (optional)
    quote_no = _norm(str((header or {}).get("quote_number") or ""))
    used_ocr = bool((header or {}).get("used_ocr"))
    notes_lines: list[str] = []
    if quote_no:
        notes_lines.append(f"Quote: {quote_no}")
    if used_ocr:
        notes_lines.append("Note: OCR used")
    ws.cell(row=2, column=6, value="\n".join(notes_lines).strip())

    # Row 3 option headers starting at C
    opt_headers = [
        "Tag", "Part Number", "Feature", "Description",
        "Qty", "List Price", "LP Ext.", "Buy Mult.", "Net Price", "Markup", "Margin", "Sell Price",
        "Weight", "Freight", "Fr. Multi.", "Alignment", "Subtotal", "Option Price"
    ]
    start_col = 3
    for i, h in enumerate(opt_headers):
        ws.cell(row=3, column=start_col + i, value=h)

    def z() -> int:
        return 0

    row_idx = 4
    option_qty = 1

    # shopping_list: write include rows (price 0)
    # all_in_one: leave lower table empty (your current “premise”)
    if output_type == "shopping_list":
        for tag in (tags or [""]):
            for d in include_descs:
                ws.cell(row=row_idx, column=3, value=tag)
                ws.cell(row=row_idx, column=4, value="")               # Part Number col in table
                ws.cell(row=row_idx, column=5, value="Include")        # Feature
                ws.cell(row=row_idx, column=6, value=d)                # Desc

                ws.cell(row=row_idx, column=7, value=option_qty)
                ws.cell(row=row_idx, column=8, value=0.0)
                ws.cell(row=row_idx, column=9, value=0.0)

                ws.cell(row=row_idx, column=10, value=z())
                ws.cell(row=row_idx, column=11, value=z())
                ws.cell(row=row_idx, column=12, value=z())
                ws.cell(row=row_idx, column=13, value=z())
                ws.cell(row=row_idx, column=14, value=z())
                ws.cell(row=row_idx, column=15, value=z())
                ws.cell(row=row_idx, column=16, value=z())
                ws.cell(row=row_idx, column=17, value=z())
                ws.cell(row=row_idx, column=18, value=z())
                ws.cell(row=row_idx, column=19, value=z())
                ws.cell(row=row_idx, column=20, value=0.0)

                row_idx += 1

    # Formatting (same vibe)
    header_fill = PatternFill("solid", fgColor="D9EAD3")
    top_header_font = Font(bold=True, italic=True)
    opt_header_font = Font(bold=True)

    for c in range(1, 7):
        cell = ws.cell(row=1, column=c)
        cell.font = top_header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    for c in range(3, 21):
        cell = ws.cell(row=3, column=c)
        cell.font = opt_header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    for r in range(2, max(row_idx, 4)):
        for c in range(1, 21):
            ws.cell(row=r, column=c).alignment = Alignment(vertical="top", wrap_text=True)

    widths = {
        1: 18, 2: 18, 3: 30, 4: 35, 5: 65, 6: 35,
        7: 8, 8: 12, 9: 12, 10: 10, 11: 12, 12: 10, 13: 10, 14: 12,
        15: 10, 16: 10, 17: 10, 18: 10, 19: 12, 20: 12
    }
    for col, w in widths.items():
        ws.column_dimensions[get_column_letter(col)].width = w

    ws.freeze_panes = "A4"


def write_aaon_template_workbook(parsed: dict[str, Any], output_type: str = "all_in_one") -> Workbook:
    wb = Workbook()
    wb.remove(wb.active)

    manufacturer = parsed.get("manufacturer", "AAON")
    header = parsed.get("header", {}) or {}
    lines = parsed.get("lines", []) or []

    if not lines:
        ws = wb.create_sheet("AAON Import")
        _write_one_sheet(
            ws,
            manufacturer=manufacturer,
            equipment="RTU",
            model_code="",
            part_number="",
            tags=[],
            header=header,
            options=[],
            output_type=output_type,
        )
        return wb

    # One sheet per model_code (same as your other converters)
    used_names: set[str] = set()
    for ln in lines:
        model_code = ln.get("model_code", "") or ""
        equipment = ln.get("equipment", "") or "RTU"
        tags = ln.get("tags", [""]) or [""]
        part_number = ln.get("part_number", "") or ""
        options = [o for o in (ln.get("options", []) or []) if isinstance(o, dict) and o]

        base_name = _safe_sheet_name(model_code or "AAON Import")
        name = base_name
        n = 2
        while name in used_names:
            suffix = f"_{n}"
            name = _safe_sheet_name(base_name[: 31 - len(suffix)] + suffix)
            n += 1
        used_names.add(name)

        ws = wb.create_sheet(title=name)
        _write_one_sheet(
            ws,
            manufacturer=manufacturer,
            equipment=equipment,
            model_code=model_code,
            part_number=part_number,
            tags=tags,
            header=header,
            options=options,
            output_type=output_type,
        )

    return wb


# ============================================================
# ROUTE-COMPATIBLE RUNNER
# ============================================================

def convert_aaon_pdf_to_xlsx_bytes(
    pdf_bytes: bytes,
    job_name: str | None = None,
    output_type: str = "all_in_one",
) -> tuple[bytes, str]:
    """
    Signature matches your unified Flask route premise:
      converter_fn(pdf_bytes=..., job_name=..., output_type=...)
    """
    job = (job_name or "").strip() or "job"
    out_type = (output_type or "all_in_one").strip().lower()
    if out_type not in ("all_in_one", "shopping_list"):
        raise ValueError(f"Unsupported output_type: {output_type}")

    parsed = parse_aaon_pdf_bytes(pdf_bytes)
    wb = write_aaon_template_workbook(parsed, output_type=out_type)

    out = io.BytesIO()
    wb.save(out)
    xlsx_bytes = out.getvalue()
    out.close()

    filename = f"{job}_aaon_{out_type}_template_output.xlsx".replace(" ", "_")
    return xlsx_bytes, filename
