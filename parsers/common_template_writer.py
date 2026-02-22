from openpyxl.styles import Font, Alignment, PatternFill
from openpyxl.utils import get_column_letter

def safe_sheet_name(name: str) -> str:
    import re
    bad = r'[:\\/?*\[\]]'
    name = re.sub(bad, "-", (name or "")).strip()
    return (name[:31] or "Sheet1")

def write_template_sheet(
    ws,
    *,
    equipment: str,
    manufacturer: str,
    model: str,
    part_number: str,
    top_description: str,
    notes: str,
    include_rows: list[dict],
):
    """
    include_rows dicts require:
      - tag (str)
      - description (str)   # already combined "Category - Item"
      - option_price (float)
    Rules enforced:
      - NO borders
      - NO blank gap row (options header is row 3)
      - Feature always "Include"
      - Qty always 1
      - All numeric cols populated (0 default)
    """

    # Row 1: headers A-F
    top_headers = [
        "Equipment",
        "Manufacturer",
        "Model",
        "Part Number",
        "Description (Not Overwritten)",
        "Notes (Not Overwritten)",
    ]
    for col, h in enumerate(top_headers, start=1):
        ws.cell(row=1, column=col, value=h)

    # Row 2: top values
    ws.cell(row=2, column=1, value=equipment)
    ws.cell(row=2, column=2, value=manufacturer)
    ws.cell(row=2, column=3, value=model)
    ws.cell(row=2, column=4, value=part_number)
    ws.cell(row=2, column=5, value=top_description)
    ws.cell(row=2, column=6, value=notes)

    # Row 3: options headers (start at column C)
    opt_headers = [
        "Tag", "Part Number", "Feature", "Description",
        "Qty", "List Price", "LP Ext.", "Buy Mult.", "Net Price", "Markup", "Margin", "Sell Price",
        "Weight", "Freight", "Fr. Multi.", "Alignment", "Subtotal", "Option Price"
    ]
    start_col = 3  # C
    for i, h in enumerate(opt_headers):
        ws.cell(row=3, column=start_col + i, value=h)

    # All numeric columns must be populated
    def z():
        return 0

    # Data rows from row 4
    r = 4
    for row in include_rows:
        tag = row.get("tag", "")
        desc = row.get("description", "")
        price = float(row.get("option_price") or 0)

        qty = 1

        ws.cell(row=r, column=3, value=tag)         # Tag
        ws.cell(row=r, column=4, value="")          # Part Number (line-level)
        ws.cell(row=r, column=5, value="Include")   # Feature
        ws.cell(row=r, column=6, value=desc)        # Description

        # Numeric columns (G..T)
        ws.cell(row=r, column=7, value=qty)         # Qty (always 1)
        ws.cell(row=r, column=8, value=price)       # List Price
        ws.cell(row=r, column=9, value=price*qty)   # LP Ext.
        ws.cell(row=r, column=10, value=z())        # Buy Mult.
        ws.cell(row=r, column=11, value=z())        # Net Price
        ws.cell(row=r, column=12, value=z())        # Markup
        ws.cell(row=r, column=13, value=z())        # Margin
        ws.cell(row=r, column=14, value=z())        # Sell Price
        ws.cell(row=r, column=15, value=z())        # Weight
        ws.cell(row=r, column=16, value=z())        # Freight
        ws.cell(row=r, column=17, value=z())        # Fr. Multi.
        ws.cell(row=r, column=18, value=z())        # Alignment
        ws.cell(row=r, column=19, value=z())        # Subtotal
        ws.cell(row=r, column=20, value=price)      # Option Price

        r += 1

    # Minimal formatting (NO borders)
    header_fill = PatternFill("solid", fgColor="D9EAD3")
    top_header_font = Font(bold=True, italic=True)
    opt_header_font = Font(bold=True)

    for c in range(1, 7):
        cell = ws.cell(row=1, column=c)
        cell.font = top_header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    for c in range(3, 21):
        cell = ws.cell(row=3, column=c)
        cell.font = opt_header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    for rr in range(2, r):
        for cc in range(1, 21):
            ws.cell(row=rr, column=cc).alignment = Alignment(vertical="top", wrap_text=True)

    # Column widths
    widths = {
        1: 18, 2: 18, 3: 16, 4: 28, 5: 40, 6: 55,
        7: 8, 8: 12, 9: 12, 10: 10, 11: 12, 12: 10, 13: 10, 14: 12,
        15: 10, 16: 10, 17: 10, 18: 10, 19: 12, 20: 12
    }
    for col, w in widths.items():
        ws.column_dimensions[get_column_letter(col)].width = w

    ws.freeze_panes = "A4"
