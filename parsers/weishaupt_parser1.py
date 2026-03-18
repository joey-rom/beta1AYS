from __future__ import annotations

import io
import re
from typing import Any

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

import logging
logger = logging.getLogger(__name__)

# ─────────────────────────────────────────────
# Shared helpers
# ─────────────────────────────────────────────

# Matches a full item line (combined/old format):
# 1.1  1  21810514502  Flame tube complete WM-G(L)10/4  548.81  548.81
ITEM_RE = re.compile(
    r"^(\d+\.\d+)\s+"
    r"(\d+)\s+"
    r"(\S+)\s+"
    r"(.+?)\s+"
    r"([\d,]+\.\d{2})\s+"
    r"([\d,]+\.\d{2})\s*$"
)

SKIP_RE = re.compile(
    r"^(Quotation:|Item\s+Qty|Weishaupt America|1320 Ellsworth|Atlanta|Phone:|Page \d|"
    r"WE ARE PLEASED|TO SIMPLIFY|Terms|PLEASE NOTE|TERMS AND|Our prices|All applicable|"
    r"Prices are firm|For burner|Start up|Warranty|All labor|This quotation|If the combustion|"
    r"Shop Drawings|Delivery times|Confirmation|Any orders|If no appliance|Orders are|"
    r"Due to the nature|Please note that|Installation of|If you would|Once we have|"
    r"By issuing|Tariff Surcharge|Grand Total|MINIMUM ORDER|\d+\.\s)",
    re.IGNORECASE
)

# Additional skip patterns used in split-format parsing
SPLIT_SKIP_RE = re.compile(
    r"^(Quotation:|Item$|Part Number$|Description$|Total$|Price/Unit$|Qty$|"
    r"Weishaupt America|1320 Ellsworth|Atlanta,|Phone:|Page \d|Visit us|"
    r"WE ARE PLEASED|TO SIMPLIFY|PLEASE NOTE|TERMS AND|Our prices|All applicable|"
    r"Prices are firm|For burner|Start up|Warranty|All labor|This quotation|If the combustion|"
    r"Shop Drawings|Delivery times|Confirmation|Any orders|If no appliance|Orders are|"
    r"Due to the nature|Please note that|Installation of|If you would|Once we have|"
    r"By issuing|Tariff Surcharge|Grand Total|MINIMUM ORDER|USD\s)",
    re.IGNORECASE
)

SECTION_NAMES = {
    "PARTS", "GAS BUTTERFLY PARTS", "GAS PARTS",
    "BURNER PARTS", "ELECTRICAL PARTS", "MECHANICAL PARTS",
}

TAG_RE      = re.compile(r"^\s*(\d+\.\d+)\s*$")
PART_NUM_RE = re.compile(r"^\s*(\d{5,})\s*$")
PRICE_RE    = re.compile(r"^\s*([\d,]+\.\d{2})\s*$")
QTY_RE      = re.compile(r"^\s*(\d+)\s*$")
STOCK_RE    = re.compile(r"^\(Stock Quantity:")
REPLACES_RE = re.compile(r"^\(replaces")


def _money(s: str) -> float:
    return float(s.replace(",", ""))


def _safe_sheet_name(name: str) -> str:
    name = re.sub(r'[:\\/?*\[\]]', "-", (name or "")).strip()
    return (name[:31] or "Sheet1")


def _is_section(line: str) -> bool:
    u = line.strip().upper()
    if u in SECTION_NAMES:
        return True
    if (re.match(r"^[A-Z][A-Z\s]{4,}$", line.strip())
            and not SKIP_RE.match(line.strip())
            and not ITEM_RE.match(line.strip())):
        return True
    return False


# ─────────────────────────────────────────────
# Format detection
# ─────────────────────────────────────────────

def _detect_format(lines: list[str]) -> str:
    """
    'combined' -> old format: tag/qty/part/desc/price/price all on one line
    'split'    -> new format: each field on its own line
    """
    for line in lines:
        if ITEM_RE.match(line.strip()):
            return "combined"
    return "split"


# ─────────────────────────────────────────────
# Combined-format parser (original/demo server)
# ─────────────────────────────────────────────

def _parse_combined(lines: list[str]) -> list[dict[str, Any]]:
    items = []
    current_section = "PARTS"
    i = 0
    while i < len(lines):
        line = lines[i].strip()
        i += 1
        if not line:
            continue
        if _is_section(line):
            current_section = line
            continue
        m = ITEM_RE.match(line)
        if m:
            item_num, qty, part_num, desc, unit_price, total = m.groups()
            desc = desc.strip()
            while i < len(lines):
                nxt = lines[i].strip()
                if not nxt: break
                if ITEM_RE.match(nxt): break
                if SKIP_RE.match(nxt): break
                if re.match(r"^\(Stock Quantity:", nxt): break
                if re.match(r"^\(replaces", nxt): break
                if re.match(r"^[A-Z][A-Z\s]{4,}$", nxt): break
                nxt_clean = nxt.lower().strip()
                desc_clean = desc.lower().strip()
                if nxt_clean not in desc_clean and not desc_clean.startswith(nxt_clean):
                    desc = desc + " " + nxt
                i += 1
            items.append({
                "tag": item_num,
                "qty": int(qty),
                "part_number": part_num,
                "desc": desc.strip(),
                "unit_price": _money(unit_price),
                "total_price": _money(total),
                "section": current_section,
            })
    return items


# ─────────────────────────────────────────────
# Split-format parser (new server)
# ─────────────────────────────────────────────

def _find_blocks(lines: list[str]) -> list[tuple]:
    """
    Find all item blocks in split format.
    Each block is: PART_NUM / TAG / QTY / PRICE / PRICE on consecutive lines.
    Returns list of (part_num_idx, tag_idx, qty_idx, p1_idx, p2_idx).
    """
    blocks = []
    n = len(lines)
    for i in range(n - 4):
        if (PART_NUM_RE.match(lines[i]) and
                TAG_RE.match(lines[i + 1]) and
                QTY_RE.match(lines[i + 2]) and
                PRICE_RE.match(lines[i + 3]) and
                PRICE_RE.match(lines[i + 4])):
            blocks.append((i, i + 1, i + 2, i + 3, i + 4))
    return blocks


def _block_tail_end(lines: list[str], p2_i: int, desc: str) -> int:
    """
    Return the index just after a block's trailing junk lines:
    (Stock Quantity: ...), (replaces ...), and one optional duplicate-desc line.
    """
    n = len(lines)
    ptr = p2_i + 1
    while ptr < n and (STOCK_RE.match(lines[ptr].strip()) or REPLACES_RE.match(lines[ptr].strip())):
        ptr += 1
    # Skip one optional duplicate desc line
    if ptr < n:
        candidate = lines[ptr].strip()
        if candidate and desc and candidate.lower() in desc.lower():
            ptr += 1
    return ptr


_HEADER_LINE_RE = re.compile(
    r"^(Page \d|1320|Weishaupt America|Atlanta|Phone|Fax|Visit|"
    r"Item$|Part Number|Description$|Total$|Qty$|Price|Quotation:)",
    re.IGNORECASE
)


def _parse_split(lines: list[str]) -> list[dict[str, Any]]:
    blocks = _find_blocks(lines)
    if not blocks:
        return []

    # Find the first PARTS/section header to anchor where item descriptions begin
    first_section_line = 0
    for i, l in enumerate(lines):
        if l.strip().upper() in SECTION_NAMES:
            first_section_line = i
            break

    items: list[dict[str, Any]] = []
    n = len(lines)

    for b_idx, (pn_i, tag_i, qty_i, p1_i, p2_i) in enumerate(blocks):

        # Determine section for this block by scanning backwards
        section = "PARTS"
        for i in range(pn_i, -1, -1):
            if _is_section(lines[i].strip()) and lines[i].strip().upper() in SECTION_NAMES:
                section = lines[i].strip().upper()
                break

        # Determine window start for description lines
        if b_idx == 0:
            window_start = first_section_line + 1
        else:
            prev_p2 = blocks[b_idx - 1][4]
            prev_desc = items[-1]["desc"] if items else ""
            window_start = _block_tail_end(lines, prev_p2, prev_desc)
            # Skip page headers, section names, and other junk in the gap
            while window_start < pn_i:
                l = lines[window_start].strip()
                if (not l or SPLIT_SKIP_RE.match(l) or _is_section(l)
                        or _HEADER_LINE_RE.match(l)):
                    window_start += 1
                else:
                    break

        # Collect description lines from window_start up to part number
        desc_lines = []
        for k in range(window_start, pn_i):
            l = lines[k].strip()
            if not l:
                continue
            if SPLIT_SKIP_RE.match(l):
                continue
            if STOCK_RE.match(l) or REPLACES_RE.match(l):
                continue
            if _is_section(l):
                continue
            if _HEADER_LINE_RE.match(l):
                continue
            desc_lines.append(l)

        desc       = " ".join(desc_lines).strip()
        part_num   = lines[pn_i].strip()
        tag        = lines[tag_i].strip()
        qty        = int(lines[qty_i].strip())
        unit_price = _money(lines[p1_i].strip())
        total      = _money(lines[p2_i].strip())

        items.append({
            "tag": tag,
            "qty": qty,
            "part_number": part_num,
            "desc": desc,
            "unit_price": unit_price,
            "total_price": total,
            "section": section,
        })

    return items


# ─────────────────────────────────────────────
# PDF PARSER (uses PyMuPDF / fitz)
# ─────────────────────────────────────────────

def parse_pdf_bytes(pdf_bytes: bytes, filename: str = "") -> dict[str, Any]:
    import fitz

    doc = fitz.open(stream=pdf_bytes, filetype="pdf")

    p1_text = doc[0].get_text("text") if len(doc) > 0 else ""

    quote_number = ""
    m = re.search(r"Quotation:\s*(\S+)", p1_text)
    if m: quote_number = m.group(1).strip()

    customer = ""
    m = re.search(r"To:\s*([A-Z][A-Z\s&,\.]+?)(?:\s{2,}|\n)", p1_text)
    if m: customer = m.group(1).strip()

    quote_date = ""
    m = re.search(r"Issued:\s*\w+,\s*(.+)", p1_text)
    if m: quote_date = m.group(1).strip()

    expiry_date = ""
    m = re.search(r"Expiry Date:\s*\w+,\s*(.+)", p1_text)
    if m: expiry_date = m.group(1).strip()

    project = ""
    m = re.search(r"Project:\s*(.+)", p1_text)
    if m: project = m.group(1).strip()

    sales_contact = ""
    m = re.search(r"Sales Contact:\s*(.+)", p1_text)
    if m: sales_contact = m.group(1).strip()

    header = {
        "quote_number": quote_number,
        "quote_date": quote_date,
        "expiry_date": expiry_date,
        "requested_ship_date": "",
        "job_name": customer,
        "customer": sales_contact,
        "project": project,
    }

    # Collect all lines from all pages
    all_lines: list[str] = []
    for page in doc:
        all_lines.extend((page.get_text("text") or "").splitlines())

    # Detect format and parse accordingly
    fmt = _detect_format(all_lines)
    logger.info(f"weishaupt_parser: detected format='{fmt}'")

    if fmt == "combined":
        items = _parse_combined(all_lines)
    else:
        items = _parse_split(all_lines)

    logger.info(f"weishaupt_parser: parsed {len(items)} items")

    if not items:
        return {"manufacturer": "Weishaupt", "header": header, "lines": []}

    equipment = items[0]["desc"] if items else (project or quote_number)

    options = [
        {
            "part_number": it["part_number"],
            "desc": it["desc"],
            "qty": it["qty"],
            "unit_price": it["unit_price"],
            "total_price": it["total_price"],
            "tag": it["tag"],
            "section": it["section"],
        }
        for it in items
    ]

    line = {
        "model_code": quote_number,
        "description": f"Weishaupt Spare Parts - {project}" if project else quote_number,
        "equipment": equipment,
        "tagging": "",
        "tags": [],
        "qty": 1,
        "list_each": 0.0,
        "total_list": 0.0,
        "options": options,
    }

    return {"manufacturer": "Weishaupt", "header": header, "lines": [line]}


# ─────────────────────────────────────────────
# XLSX WRITER
# ─────────────────────────────────────────────

def _write_one_sheet(ws, *, manufacturer, equipment, model_code,
                     description, tags, header, options) -> None:
    top_headers = ["Equipment", "Manufacturer", "Model", "Part Number",
                   "Description (Not Overwritten)", "Notes (Not Overwritten)"]
    for col, h in enumerate(top_headers, start=1):
        ws.cell(row=1, column=col, value=h)

    quote_no    = header.get("quote_number", "")
    job_name    = header.get("job_name", "")
    quote_date  = header.get("quote_date", "")
    expiry_date = header.get("expiry_date", "")
    customer    = header.get("customer", "")
    project     = header.get("project", "")

    ws.cell(row=2, column=1, value=equipment)
    ws.cell(row=2, column=2, value=manufacturer)
    ws.cell(row=2, column=3, value=None)
    ws.cell(row=2, column=4, value=None)
    ws.cell(row=2, column=5, value=description)

    notes_parts = []
    if quote_no:    notes_parts.append(f"Quote #: {quote_no}")
    if project:     notes_parts.append(f"Project: {project}")
    if job_name:    notes_parts.append(f"Customer: {job_name}")
    if customer:    notes_parts.append(f"Sales Contact: {customer}")
    if quote_date:  notes_parts.append(f"Issued: {quote_date}")
    if expiry_date: notes_parts.append(f"Expiry: {expiry_date}")
    ws.cell(row=2, column=6, value="\n".join(notes_parts))

    opt_headers = [
        "Tag", "Part Number", "Feature", "Description",
        "Qty", "List Price", "LP Ext.", "Buy Mult.", "Net Price",
        "Markup", "Margin", "Sell Price", "Weight", "Freight",
        "Fr. Multi.", "Alignment", "Subtotal", "Option Price",
    ]
    for i, h in enumerate(opt_headers):
        ws.cell(row=3, column=3 + i, value=h)

    row_idx = 4
    for opt_idx, opt in enumerate(options or []):
        part_num   = opt.get("part_number", "")
        desc       = opt.get("desc", "")
        unit_price = float(opt.get("unit_price") or 0)
        opt_qty    = int(opt.get("qty") or 1)
        lp_ext     = unit_price * opt_qty
        tag        = opt.get("tag") or (", ".join(tags) if tags else "")
        feature    = "Equipment" if opt_idx == 0 else "Accessory"

        ws.cell(row=row_idx, column=3,  value=tag)
        ws.cell(row=row_idx, column=4,  value=part_num)
        ws.cell(row=row_idx, column=5,  value=feature)
        ws.cell(row=row_idx, column=6,  value=desc)
        ws.cell(row=row_idx, column=7,  value=opt_qty)
        ws.cell(row=row_idx, column=8,  value=unit_price)
        ws.cell(row=row_idx, column=9,  value=lp_ext)
        for c in range(10, 21):
            ws.cell(row=row_idx, column=c, value=0)
        ws.cell(row=row_idx, column=20, value=unit_price)
        row_idx += 1

    header_fill = PatternFill("solid", fgColor="D9EAD3")
    top_font    = Font(bold=True, italic=True, name="Arial")
    opt_font    = Font(bold=True, name="Arial")
    body_font   = Font(name="Arial")

    for c in range(1, 7):
        cell = ws.cell(row=1, column=c)
        cell.font = top_font; cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    for c in range(3, 21):
        cell = ws.cell(row=3, column=c)
        cell.font = opt_font; cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    for r in range(2, max(row_idx, 4)):
        for c in range(1, 21):
            cell = ws.cell(row=r, column=c)
            cell.alignment = Alignment(vertical="top", wrap_text=True)
            cell.font = body_font

    widths = {1:28, 2:18, 3:16, 4:20, 5:55, 6:45,
              7:8, 8:12, 9:12, 10:10, 11:12, 12:10,
              13:10, 14:12, 15:10, 16:10, 17:10, 18:10, 19:12, 20:12}
    for col, w in widths.items():
        ws.column_dimensions[get_column_letter(col)].width = w

    ws.freeze_panes = "A4"


def write_template_workbook(parsed: dict[str, Any]) -> Workbook:
    wb = Workbook()
    wb.remove(wb.active)

    manufacturer = parsed.get("manufacturer", "")
    header       = parsed.get("header", {})
    lines        = parsed.get("lines", [])

    if not lines:
        ws = wb.create_sheet("Import")
        _write_one_sheet(ws, manufacturer=manufacturer, equipment="",
                         model_code="", description="", tags=[],
                         header=header, options=[])
        return wb

    sheet_names_seen: set[str] = set()
    for ln in lines:
        model_code  = ln.get("model_code", "") or ""
        description = ln.get("description", "") or ""
        tagging     = ln.get("tagging", "") or ""
        tags        = ln.get("tags") or ([tagging] if tagging else [])
        options     = ln.get("options", []) or []
        equipment   = ln.get("equipment", "")

        base_name = _safe_sheet_name(model_code or "Import")
        name = base_name
        n = 2
        while name in sheet_names_seen:
            suffix = f"_{n}"
            name = _safe_sheet_name(base_name[:31 - len(suffix)] + suffix)
            n += 1
        sheet_names_seen.add(name)

        ws = wb.create_sheet(title=name)
        _write_one_sheet(ws, manufacturer=manufacturer, equipment=equipment,
                         model_code=model_code, description=description,
                         tags=tags, header=header, options=options)

    return wb


# ─────────────────────────────────────────────
# PUBLIC API
# ─────────────────────────────────────────────

def convert_weishaupt_to_xlsx_bytes(
    pdf_bytes: bytes,
    filename: str = "",
    job_name: str | None = None,
    output_type: str | None = None,
    **kwargs,
) -> tuple[bytes, str]:
    """
    Main entry point. Matches the standard converter signature used by all parsers.
    Requires PyMuPDF (PyMuPDF==1.26.3 already in your environment).
    Handles both combined (old server) and split (new server) PDF text extraction formats.
    """
    parsed = parse_pdf_bytes(pdf_bytes, filename=filename)

    if job_name:
        parsed["header"]["job_name"] = job_name.strip()

    wb = write_template_workbook(parsed)

    out = io.BytesIO()
    wb.save(out)
    xlsx_bytes = out.getvalue()
    out.close()

    job = (parsed["header"].get("job_name") or
           re.sub(r'\.pdf$', '', filename, flags=re.IGNORECASE) or
           "output")
    return xlsx_bytes, f"{job}_template_output.xlsx"


# ─────────────────────────────────────────────
# Quick test when run directly
# ─────────────────────────────────────────────
if __name__ == "__main__":
    import sys
    path = sys.argv[1] if len(sys.argv) > 1 else "/mnt/user-data/uploads/P11773926-00-Wren_Industries.pdf"
    with open(path, "rb") as f:
        data = f.read()
    fname = path.split("/")[-1]
    xlsx_bytes, out_name = convert_weishaupt_to_xlsx_bytes(data, filename=fname)
    out_path = f"/mnt/user-data/outputs/{out_name}"
    with open(out_path, "wb") as f:
        f.write(xlsx_bytes)
    print(f"Written: {out_path}")
