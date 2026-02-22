# multistack_parser1.py  (DROP-IN MODULE)
from __future__ import annotations

import io
import re
import logging
from typing import Any

import fitz  # PyMuPDF
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

logger = logging.getLogger(__name__)


# -----------------------------
# Helpers
# -----------------------------

def _norm(s: str) -> str:
    return re.sub(r"\s+", " ", (s or "")).strip()

def _safe_sheet_name(name: str) -> str:
    bad = r'[:\\/?*\[\]]'
    name = re.sub(bad, "-", (name or "")).strip()
    return (name[:31] or "Sheet1")


# -----------------------------
# PDF text extraction
# -----------------------------

def _pdf_text(pdf_bytes: bytes) -> str:
    doc = fitz.open(stream=pdf_bytes, filetype="pdf")
    parts: list[str] = []
    for i in range(len(doc)):
        t = doc[i].get_text("text") or ""
        if t.strip():
            parts.append(t)
    return "\n".join(parts)

def _grab_quote(text: str) -> str:
    m = re.search(r"\bQSD\d+\b", text)
    return m.group(0) if m else ""

def _grab_modules_and_model(text: str) -> tuple[int | None, str]:
    # Prefer explicit MULTISTACK (2) MS50Z6H2W
    m = re.search(r"\bMULTISTACK\s*\((\d+)\)\s*([A-Z0-9\-]+)\b", text, flags=re.IGNORECASE)
    if m:
        return int(m.group(1)), m.group(2).strip()

    # Fallback: just model
    m2 = re.search(r"\b(MS\d+[A-Z0-9\-]*)\b", text)
    if m2:
        return None, m2.group(1).strip()

    return None, ""

def _grab_line_after_label(text: str, label: str) -> str:
    # Handles cases where value is on the next line
    m = re.search(rf"{re.escape(label)}\s*\n\s*([^\n]+)", text, flags=re.IGNORECASE)
    return _norm(m.group(1)) if m else ""

def _grab_inline_after_colon(text: str, label: str) -> str:
    m = re.search(rf"{re.escape(label)}\s*:\s*([^\n]+)", text, flags=re.IGNORECASE)
    return _norm(m.group(1)) if m else ""

def _grab_dims_from_header_area(text: str) -> tuple[str, str, str]:
    """
    This PDF interleaves the '58 / 49 1/4 / 64' values near the top header area.
    Heuristic: find the first Month-name date line, then take the next 3 numeric-ish lines.
    """
    lines = [ln.strip() for ln in (text or "").splitlines() if ln.strip()]
    months = r"(January|February|March|April|May|June|July|August|September|October|November|December)"
    date_i = None
    for i, ln in enumerate(lines):
        if re.search(months, ln) and re.search(r"\b\d{4}\b", ln):
            date_i = i
            break
    if date_i is None:
        return "", "", ""

    nums: list[str] = []
    for ln in lines[date_i + 1 : date_i + 20]:
        # allow: 58, 49 1/4, 64, 49-1/4, etc.
        if re.fullmatch(r"\d+(?:[\s\-]+\d/\d)?", ln):
            nums.append(ln.replace("-", " "))
            if len(nums) == 3:
                break

    if len(nums) == 3:
        return nums[0], nums[1], nums[2]
    return "", "", ""

def _extract_bullets(text: str) -> list[str]:
    """
    Multistack 'features' are bullet-driven with "•" markers.
    Pull those lines and wrap continuations until next bullet / header.
    """
    raw_lines = [ln.rstrip() for ln in (text or "").splitlines()]
    lines = [ln.strip() for ln in raw_lines]

    skip_prefixes = [
        "Length:", "Width:", "Height:", "Weight:", "Dimensions (Inches):",
        "(Dimensions Do Not Include Junction Boxes)", "Total Shipping Weight:",
        "Total Operating Weight:", "Cooling Capacity:", "Sound Pressure:",
        "ELECTRICAL DATA", "MAIN POWER SUPPLY", "Minimum Circuit Ampacity",
        "Maximum Over Current Protection",
    ]

    bullets: list[str] = []
    i = 0
    while i < len(lines):
        ln = lines[i]
        if ln == "•" or ln.startswith("•"):
            cur = ln.lstrip("•").strip()
            if not cur:
                i += 1
                cur = lines[i].strip() if i < len(lines) else ""

            parts: list[str] = [cur] if cur else []
            j = i + 1
            while j < len(lines):
                nxt = lines[j].strip()
                if not nxt:
                    j += 1
                    continue
                if nxt == "•" or nxt.startswith("•"):
                    break
                if nxt.isupper() and len(nxt.split()) <= 5:
                    break
                if any(nxt.startswith(p) for p in skip_prefixes):
                    j += 1
                    continue

                parts.append(nxt)
                j += 1

            bullet = " ".join([p for p in parts if p]).strip()
            bullet = re.sub(r"\s+", " ", bullet)
            bullet = re.sub(r"Dimensions \(Inches\).*", "", bullet, flags=re.IGNORECASE).strip()
            if bullet and bullet not in bullets:
                bullets.append(bullet)

            i = j
            continue

        i += 1

    return bullets


# -----------------------------
# Parser -> structured dict
# -----------------------------

def parse_multistack_pdf_bytes(pdf_bytes: bytes) -> dict[str, Any]:
    text = _pdf_text(pdf_bytes)

    quote_no = _grab_quote(text)
    modules, model = _grab_modules_and_model(text)

    # Dimensions & Weight
    length, width, height = _grab_dims_from_header_area(text)
    ship_w = _grab_inline_after_colon(text, "Total Shipping Weight")
    op_w = _grab_inline_after_colon(text, "Total Operating Weight")

    cooling = _grab_inline_after_colon(text, "Cooling Capacity")
    sound = _grab_inline_after_colon(text, "Sound Pressure")

    # Electrical (values often on next line)
    main_power = _grab_line_after_label(text, "MAIN POWER SUPPLY")
    mca = _grab_line_after_label(text, "Minimum Circuit Ampacity (amps)")
    mop = _grab_line_after_label(text, "Maximum Over Current Protection (MOP)")

    bullets = _extract_bullets(text)

    # Build description exactly in the “single big block” style you showed
    desc_lines: list[str] = []

    # Title line
    title_bits = []
    if modules:
        title_bits.append(f"MULTISTACK ({modules}) Modules")
    else:
        title_bits.append("MULTISTACK Modules")
    if model:
        title_bits.append(f"- model {model}")
    title = " ".join(title_bits).strip()
    if model:
        title += " with the following standard and optional features:"
    desc_lines.append(title)

    # Dimensions
    dim_lines: list[str] = []
    if any([length, width, height]):
        dim_lines.append("Dimensions (Inches): (Dimensions Do Not Include Junction Boxes)")
        if length:
            dim_lines.append(f"Length: {length}")
        if width:
            dim_lines.append(f"Width: {width}")
        if height:
            dim_lines.append(f"Height: {height}")

    # Weight
    weight_lines: list[str] = []
    if ship_w or op_w:
        weight_lines.append("Weight:")
        if ship_w:
            weight_lines.append(f"Total Shipping Weight: {ship_w}")
        if op_w:
            weight_lines.append(f"Total Operating Weight: {op_w}")

    # Performance
    perf_lines: list[str] = []
    if cooling:
        perf_lines.append(f"Cooling Capacity: {cooling}")
    if sound:
        perf_lines.append(f"Sound Pressure: {sound}")

    # Electrical
    elec_lines: list[str] = []
    if main_power or mca or mop:
        elec_lines.append("ELECTRICAL DATA")
        if main_power:
            elec_lines.append(f"MAIN POWER SUPPLY {main_power}")
        if mca:
            elec_lines.append(f"Minimum Circuit Ampacity (amps) {mca}")
        if mop:
            elec_lines.append(f"Maximum Over Current Protection (MOP) {mop}")

    # Features (bullets)
    feat_lines: list[str] = []
    if bullets:
        feat_lines.extend(bullets)

    # Append sections with blank spacing like your example
    for block in [dim_lines, weight_lines, perf_lines, elec_lines, feat_lines]:
        if block:
            desc_lines.append("")
            desc_lines.extend(block)

    description = "\n".join(desc_lines).strip()

    return {
        "manufacturer": "Multistack",
        "header": {"quote_number": quote_no},
        "lines": [{
            "model_code": model,
            "tagging": "",
            "tags": [""],
            "qty": 1,
            "list_each": 0.0,
            "total_list": 0.0,
            "options": [
                {"feature": "Equipment", "desc": description, "add_price": 0.0},
            ]
        }]
    }


# -----------------------------
# Workbook writer (same vibe as your others)
# -----------------------------

def _write_one_sheet(
    ws,
    *,
    manufacturer: str,
    equipment: str,
    model_code: str,
    header: dict[str, Any],
    options: list[dict[str, Any]],
) -> None:
    # Row 1 headers A-F
    top_headers = [
        "Equipment", "Manufacturer", "Model", "Part Number",
        "Description (Not Overwritten)", "Notes (Not Overwritten)"
    ]
    for col, h in enumerate(top_headers, start=1):
        ws.cell(row=1, column=col, value=h)

    # Row 2 values per your requirement:
    # C2 = full model, D2 blank, notes = quote #
    ws.cell(row=2, column=1, value=equipment)         # Chiller
    ws.cell(row=2, column=2, value=manufacturer)      # Multistack
    ws.cell(row=2, column=3, value=model_code or "")  # MS50Z6H2W
    ws.cell(row=2, column=4, value="")                # blank part number

    main_desc = ""
    for opt in (options or []):
        if isinstance(opt, dict) and (opt.get("feature") or "").strip().lower() == "equipment":
            main_desc = (opt.get("desc") or "").strip()
            break

    ws.cell(row=2, column=5, value=main_desc)  # big description block

    quote_no = _norm(str((header or {}).get("quote_number") or ""))
    ws.cell(row=2, column=6, value=(f"Quote #: {quote_no}" if quote_no else ""))

    # Row 3 option headers starting at C (kept for template consistency; left empty)
    opt_headers = [
        "Tag","Part Number","Feature","Description","Qty","List Price","LP Ext.","Buy Mult.","Net Price","Markup","Margin","Sell Price",
        "Weight","Freight","Fr. Multi.","Alignment","Subtotal","Option Price"
    ]
    start_col = 3
    for i, h in enumerate(opt_headers):
        ws.cell(row=3, column=start_col + i, value=h)

    # Formatting
    header_fill = PatternFill("solid", fgColor="D9EAD3")
    top_header_font = Font(bold=True, italic=True)
    opt_header_font = Font(bold=True)

    for c in range(1, 7):
        cell = ws.cell(row=1, column=c)
        cell.font = top_header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    for c in range(3, 21):
        cell = ws.cell(row=3, column=c)
        cell.font = opt_header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    for r in range(2, 6):
        for c in range(1, 21):
            ws.cell(row=r, column=c).alignment = Alignment(vertical="top", wrap_text=True)

    widths = {
        1: 18, 2: 18, 3: 22, 4: 28, 5: 60, 6: 28,
        7: 8, 8: 12, 9: 12, 10: 10, 11: 12, 12: 10, 13: 10, 14: 12,
        15: 10, 16: 10, 17: 10, 18: 10, 19: 12, 20: 12
    }
    for col, w in widths.items():
        ws.column_dimensions[get_column_letter(col)].width = w

    ws.freeze_panes = "A4"


def write_multistack_workbook(parsed: dict[str, Any]) -> Workbook:
    wb = Workbook()
    wb.remove(wb.active)

    manufacturer = parsed.get("manufacturer", "Multistack")
    header = parsed.get("header", {}) or {}
    lines = parsed.get("lines", []) or []

    if not lines:
        ws = wb.create_sheet("Multistack Import")
        _write_one_sheet(
            ws,
            manufacturer=manufacturer,
            equipment="Chiller",
            model_code="",
            header=header,
            options=[],
        )
        return wb

    # One sheet per model (typical for these)
    for ln in lines:
        model_code = (ln.get("model_code") or "").strip()
        options = [o for o in (ln.get("options", []) or []) if isinstance(o, dict)]
        ws = wb.create_sheet(_safe_sheet_name(model_code or "Multistack Import"))
        _write_one_sheet(
            ws,
            manufacturer=manufacturer,
            equipment="Chiller",
            model_code=model_code,
            header=header,
            options=options,
        )

    return wb


# -----------------------------
# Route-compatible runner
# -----------------------------

def convert_multistack_pdf_to_xlsx_bytes(
    pdf_bytes: bytes,
    job_name: str | None = None,
    output_type: str = "all_in_one",
) -> tuple[bytes, str]:
    """
    Same signature as your existing converters.
    output_type is accepted for compatibility, but Multistack output is an all-in-one style sheet.
    """
    job = (job_name or "").strip() or "job"
    parsed = parse_multistack_pdf_bytes(pdf_bytes)

    wb = write_multistack_workbook(parsed)

    out = io.BytesIO()
    wb.save(out)
    xlsx_bytes = out.getvalue()
    out.close()

    filename = f"{job}_multistack_{(output_type or 'all_in_one')}_template_output.xlsx".replace(" ", "_")
    return xlsx_bytes, filename

