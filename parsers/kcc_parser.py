# kcc_parser.py  (DROP-IN REPLACEMENT)
from __future__ import annotations

import io
import re
from typing import Any

import fitz  # PyMuPDF
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter


# -----------------------------
# Helpers / regex
# -----------------------------

DATE_RE = r"[0-9]{1,2}/[0-9]{1,2}/[0-9]{4}"
MONEY_RE = re.compile(r"\$([0-9]{1,3}(?:,[0-9]{3})*(?:\.[0-9]{2}))")


def _grab(pattern: str, text: str, default: str = "") -> str:
    m = re.search(pattern, text, flags=re.MULTILINE)
    return m.group(1).strip() if m else default


def _money_to_float(s: str) -> float:
    return float(s.replace(",", ""))


def _split_tags(tagging: str) -> list[str]:
    """
    Example: "RTU-CORR-A; RTU-CORR-B, RTU-CORR-C"
    Split on ';' and ',' and trim. Remove empties.
    """
    if not tagging:
        return []
    parts = re.split(r"[;,]", tagging)
    return [p.strip() for p in parts if p.strip()]


def _safe_sheet_name(name: str) -> str:
    # Excel: <=31 chars, no : \ / ? * [ ]
    bad = r'[:\\/?*\[\]]'
    name = re.sub(bad, "-", (name or "")).strip()
    return (name[:31] or "Sheet1")


# ✅ robust warranty detection (contains)
def _is_warranty_category(category: str) -> bool:
    if not category:
        return False
    return "warranty" in category.strip().lower()


def _build_includes_description(options: list[dict[str, Any]]) -> str:
    """
    Builds the Description cell (E2) as:

    Base Unit

    Category: Desc
    Category: Desc
    ...

    Warranty

    Desc
    Desc
    ...
    """
    base_lines: list[str] = []
    warranty_lines: list[str] = []

    for opt in (options or []):
        cat = (opt.get("category") or "").strip()
        desc = (opt.get("desc") or "").strip()
        if not cat and not desc:
            continue

        if _is_warranty_category(cat):
            if desc:
                warranty_lines.append(desc)
        else:
            if cat and desc:
                base_lines.append(f"{cat}: {desc}")
            elif desc:
                base_lines.append(desc)
            elif cat:
                base_lines.append(cat)

    chunks: list[str] = []
    chunks.append("Base Unit")
    if base_lines:
        chunks.append("")
        chunks.extend(base_lines)

    if warranty_lines:
        chunks.append("")
        chunks.append("")
        chunks.append("Warranty")
        chunks.append("")
        chunks.extend(warranty_lines)

    return "\n".join(chunks).strip()


def _build_notes(header: dict[str, Any]) -> str:
    """
    Builds the Notes cell (F2) as quote metadata.
    """
    quote_no = (header.get("quote_number") or "").strip()
    job_name = (header.get("job_name") or "").strip()
    quote_date = (header.get("quote_date") or "").strip()
    ship_date = (header.get("requested_ship_date") or "").strip()

    lines: list[str] = []
    if quote_no:
        lines.append(f"Quote #: {quote_no}")
    if job_name:
        lines.append(f"Job Name: {job_name}")
    if quote_date:
        lines.append(f"Quote Date: {quote_date}")
    if ship_date:
        lines.append(f"Requested Ship Date: {ship_date}")

    return "\n".join(lines)


# -----------------------------
# KCC PARSER (PDF bytes -> structured dict)
# -----------------------------

def parse_kcc_pdf_bytes(pdf_bytes: bytes) -> dict[str, Any]:
    """
    Returns:
      {
        "manufacturer": "KCC",
        "header": {...},
        "lines": [
           {
             "model_code": "...",
             "tagging": "...",
             "tags": [...],
             "qty": <int>,
             "list_each": <float>,
             "total_list": <float>,
             "options": [{"category":..., "desc":..., "add_price":...}, ...]
           },
           ...
        ]
      }
    """
    doc = fitz.open(stream=pdf_bytes, filetype="pdf")

    # Page 1 header
    p1 = doc[0].get_text("text") if len(doc) > 0 else ""
    header = {
        "quote_number": _grab(r"Quote\s*#\s*([0-9A-Za-z\-]+)", p1),
        "quote_date": _grab(r"Date:\s*(" + DATE_RE + r")", p1),
        "requested_ship_date": _grab(r"Requested Ship Date\s*(" + DATE_RE + r")", p1),
        "job_name": _grab(r"Job Name:\s*([^\n]+)", p1),
    }

    # Page 2 blocks
    p2_blocks = doc[1].get_text("blocks") if len(doc) > 1 else []

    # --- Find ALL order blocks (future-proof: multiple model lines) ---
    order_blocks: list[str] = []
    for b in p2_blocks:
        t = (b[4] or "").strip()
        # KCC sample: model code begins with OADD... and order block includes at least 2 money values
        if "OADD" in t and t.count("$") >= 2:
            order_blocks.append(t)

    lines_out: list[dict[str, Any]] = []
    for order_block_text in order_blocks:
        model_code = ""
        tagging = ""
        qty = 0
        list_each = 0.0
        total_list = 0.0

        block_lines = [
            ln.strip()
            for ln in order_block_text.splitlines()
            if ln.strip() and ln.strip() != "..."
        ]

        if block_lines:
            # Model code may span 2 lines if first ends with '-'
            if len(block_lines) >= 2 and block_lines[0].endswith("-"):
                model_code = (block_lines[0] + block_lines[1]).replace(" ", "")
                rest = block_lines[2:]
            else:
                model_code = block_lines[0].replace(" ", "")
                rest = block_lines[1:]

            # Tagging + Qty line typically ends with qty
            if rest:
                m = re.search(r"(.+?)\s+(\d+)\s*$", rest[0])
                if m:
                    tagging = " ".join(m.group(1).split())
                    qty = int(m.group(2))

            # Money lines
            monies: list[float] = []
            for ln in block_lines:
                mm = MONEY_RE.search(ln)
                if mm:
                    monies.append(_money_to_float(mm.group(1)))
            if len(monies) >= 2:
                list_each = monies[0]
                total_list = monies[1]

        lines_out.append({
            "model_code": model_code,
            "tagging": tagging,
            "tags": _split_tags(tagging),
            "qty": int(qty or 0),
            "list_each": float(list_each or 0),
            "total_list": float(total_list or 0),
            "options": []  # filled after we parse options globally
        })

    # --- Parse option blocks (category line, desc lines, $ price line) ---
    options: list[dict[str, Any]] = []
    for b in p2_blocks:
        t = (b[4] or "").strip()
        if not t or t == "..." or "$" not in t:
            continue
        if "OADD" in t:  # skip main order block(s)
            continue

        parts = [ln.strip() for ln in t.splitlines() if ln.strip() and ln.strip() != "..."]
        if len(parts) < 3:
            continue

        category = parts[0]
        price_line = parts[-1]
        desc = " ".join(parts[1:-1]).strip()

        mm = MONEY_RE.search(price_line)
        if not mm:
            continue

        add_price = _money_to_float(mm.group(1))
        options.append({"category": category, "desc": desc, "add_price": add_price})

    # Apply parsed options to each line (works for your current KCC format)
    for ln in lines_out:
        ln["options"] = options

    return {
        "manufacturer": "KCC",
        "header": header,
        "lines": lines_out
    }


# -----------------------------
# TEMPLATE WRITER (per model sheet, NOT per tag)
# -----------------------------

def _write_one_sheet(
    ws,
    *,
    manufacturer: str,
    equipment: str,
    model_code: str,
    tags: list[str],
    header: dict[str, Any],
    options: list[dict[str, Any]],
) -> None:
    """
    Write ONE sheet for a single (equipment, model_code),
    containing ALL tags for that model.

    Each tag gets its own block of option rows in the same sheet.
    """
    # Row 1: top headers (A-F)
    top_headers = [
        "Equipment",
        "Manufacturer",
        "Model",
        "Part Number",
        "Description (Not Overwritten)",
        "Notes (Not Overwritten)",
    ]
    for col, h in enumerate(top_headers, start=1):
        ws.cell(row=1, column=col, value=h)

    # Row 2: top values (shared for the whole model)
    model_short = model_code.split("-")[0] if model_code else ""

    ws.cell(row=2, column=1, value=equipment)
    ws.cell(row=2, column=2, value=manufacturer)
    ws.cell(row=2, column=3, value=model_short)
    ws.cell(row=2, column=4, value=model_code)

    # ✅ NEW behavior:
    #   Description = Base Unit/Warranty includes
    #   Notes = Quote # / Job Name / Quote Date / Ship Date
    ws.cell(row=2, column=5, value=_build_includes_description(options))
    ws.cell(row=2, column=6, value=_build_notes(header))

    # Row 3: options table headers starting at column C (NO GAP ROW)
    opt_headers = [
        "Tag",
        "Part Number",
        "Feature",
        "Description",
        "Qty",
        "List Price",
        "LP Ext.",
        "Buy Mult.",
        "Net Price",
        "Markup",
        "Margin",
        "Sell Price",
        "Weight",
        "Freight",
        "Fr. Multi.",
        "Alignment",
        "Subtotal",
        "Option Price",
    ]
    start_col = 3  # C
    for i, h in enumerate(opt_headers):
        ws.cell(row=3, column=start_col + i, value=h)

    def z() -> int:
        return 0

    option_qty = 1  # ALWAYS 1 for upload rows
    row_idx = 4

    # loop over ALL tags for this model, and write the same options for each tag
    for tag in (tags or [""]):
        for opt in options:
            combined_desc = f"{(opt.get('category') or '').strip()} - {(opt.get('desc') or '').strip()}".strip(" -")
            option_price = float(opt.get("add_price") or 0)
            lp_ext = option_price * option_qty

            ws.cell(row=row_idx, column=3, value=tag)           # Tag
            ws.cell(row=row_idx, column=4, value="")            # Part Number
            ws.cell(row=row_idx, column=5, value="Include")     # Feature
            ws.cell(row=row_idx, column=6, value=combined_desc) # Description

            ws.cell(row=row_idx, column=7, value=option_qty)    # Qty
            ws.cell(row=row_idx, column=8, value=option_price)  # List Price
            ws.cell(row=row_idx, column=9, value=lp_ext)        # LP Ext.
            ws.cell(row=row_idx, column=10, value=z())          # Buy Mult.
            ws.cell(row=row_idx, column=11, value=z())          # Net Price
            ws.cell(row=row_idx, column=12, value=z())          # Markup
            ws.cell(row=row_idx, column=13, value=z())          # Margin
            ws.cell(row=row_idx, column=14, value=z())          # Sell Price
            ws.cell(row=row_idx, column=15, value=z())          # Weight
            ws.cell(row=row_idx, column=16, value=z())          # Freight
            ws.cell(row=row_idx, column=17, value=z())          # Fr. Multi.
            ws.cell(row=row_idx, column=18, value=z())          # Alignment
            ws.cell(row=row_idx, column=19, value=z())          # Subtotal
            ws.cell(row=row_idx, column=20, value=option_price) # Option Price

            row_idx += 1

    # Minimal formatting (NO borders)
    header_fill = PatternFill("solid", fgColor="D9EAD3")
    top_header_font = Font(bold=True, italic=True)
    opt_header_font = Font(bold=True)

    for c in range(1, 7):
        cell = ws.cell(row=1, column=c)
        cell.font = top_header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    for c in range(3, 21):
        cell = ws.cell(row=3, column=c)
        cell.font = opt_header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    for r in range(2, row_idx):
        for c in range(1, 21):
            ws.cell(row=r, column=c).alignment = Alignment(vertical="top", wrap_text=True)

    widths = {
        1: 18,
        2: 18,
        3: 16,
        4: 28,
        5: 40,  # Description column is multi-line
        6: 40,  # Notes column is multi-line now too
        7: 8,
        8: 12,
        9: 12,
        10: 10,
        11: 12,
        12: 10,
        13: 10,
        14: 12,
        15: 10,
        16: 10,
        17: 10,
        18: 10,
        19: 12,
        20: 12,
    }
    for col, w in widths.items():
        ws.column_dimensions[get_column_letter(col)].width = w

    ws.freeze_panes = "A4"


def write_kcc_template_workbook(parsed: dict[str, Any]) -> Workbook:
    """
    Creates ONE workbook and ONE sheet per (equipment, model_code).

    - Sheet name = model code (fallbacks if blank).
    - Each sheet contains ALL tags for that model/equipment combo.
    """
    wb = Workbook()
    wb.remove(wb.active)

    manufacturer = parsed.get("manufacturer", "KCC")
    header = parsed.get("header", {})
    lines = parsed.get("lines", [])

    # If nothing parsed, just create a single blank-ish sheet
    if not lines:
        ws = wb.create_sheet("KCC Import")
        _write_one_sheet(
            ws,
            manufacturer=manufacturer,
            equipment="",
            model_code="",
            tags=[],
            header=header,
            options=[],
        )
        return wb

    # Group by (equipment, model_code)
    groups: dict[tuple[str, str], dict[str, Any]] = {}

    for ln in lines:
        model_code = ln.get("model_code", "") or ""
        tagging = ln.get("tagging", "") or ""
        tags = ln.get("tags") or ([tagging] if tagging else [""])
        options = ln.get("options", []) or []

        # You can refine equipment logic here later if needed
        equipment = "RTU" if tagging else ""

        key = (equipment, model_code)
        if key not in groups:
            groups[key] = {
                "equipment": equipment,
                "model_code": model_code,
                "tags": [],
                "options": options,
            }
        # Add tags to this group
        groups[key]["tags"].extend(tags)

    # Deduplicate tags per group (preserve order)
    for _, info in groups.items():
        seen: set[str] = set()
        deduped_tags: list[str] = []
        for t in info["tags"]:
            if t and t not in seen:
                seen.add(t)
                deduped_tags.append(t)
        info["tags"] = deduped_tags

    # Build sheets: ONE per (equipment, model_code)
    sheet_names_seen: set[str] = set()

    for (equipment, model_code), info in groups.items():
        base_name = _safe_sheet_name(model_code or equipment or "KCC Import")
        name = base_name
        n = 2
        while name in sheet_names_seen:
            suffix = f"_{n}"
            name = _safe_sheet_name(base_name[: 31 - len(suffix)] + suffix)
            n += 1
        sheet_names_seen.add(name)

        ws = wb.create_sheet(title=name)

        _write_one_sheet(
            ws,
            manufacturer=manufacturer,
            equipment=equipment,
            model_code=model_code,
            tags=info["tags"],
            header=header,
            options=info["options"],
        )

    return wb


# -----------------------------
# RUNNER (PDF bytes -> XLSX bytes)
# -----------------------------

import logging
logger = logging.getLogger(__name__)

def convert_kcc_pdf_to_xlsx_bytes(
    pdf_bytes: bytes,
    job_name: str | None = None,
    output_type: str = "all_in_one",
) -> tuple[bytes, str]:
    """
    output_type:
      - "all_in_one" (default)  -> this workbook format (Base Unit/Warranty in Description, quote meta in Notes)
      - "shopping_list"         -> calls your shopping list builder if present (you can wire it later)
    """
    job = (job_name or "").strip() or "job"
    parsed = parse_kcc_pdf_bytes(pdf_bytes)

    out_type = (output_type or "all_in_one").strip().lower()

    if out_type == "all_in_one":
        wb = write_kcc_template_workbook(parsed)

    elif out_type == "shopping_list":
        # If you already have the shopping list workbook writer in this file, name it:
        #   write_kcc_shopping_list_workbook(parsed) -> Workbook
        if "write_kcc_shopping_list_workbook" in globals():
            wb = globals()["write_kcc_shopping_list_workbook"](parsed)
        else:
            raise RuntimeError(
                "Shopping list output requested, but write_kcc_shopping_list_workbook(parsed) "
                "was not found in kcc_parser.py. Paste that function into this module (or tell me "
                "what file/function it lives in) and I’ll wire it in cleanly."
            )

    else:
        raise ValueError(f"Unsupported output_type: {output_type}")

    # DEBUG: log sheet names so we can confirm grouping
    sheet_names = [ws.title for ws in wb.worksheets]
    logger.info("KCC DEBUG: output_type=%s workbook sheets=%s", out_type, sheet_names)

    out = io.BytesIO()
    wb.save(out)
    xlsx_bytes = out.getvalue()
    out.close()

    filename = f"{job}_{out_type}_template_output.xlsx"
    return xlsx_bytes, filename






