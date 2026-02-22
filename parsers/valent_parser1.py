# valent_parser1.py  (DROP-IN MODULE)
from __future__ import annotations

import io
import re
import logging
from typing import Any

import fitz  # PyMuPDF
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

logger = logging.getLogger(__name__)

# -----------------------------
# Helpers / regex
# -----------------------------

DOLLAR_ANY_RE = re.compile(r"\$\s*([0-9]{1,3}(?:,[0-9]{3})*(?:\.[0-9]{2})?)")

def _norm(s: str) -> str:
    return re.sub(r"\s+", " ", (s or "")).strip()

def _money_to_float(s: str) -> float:
    return float(s.replace(",", "").strip())

def _safe_sheet_name(name: str) -> str:
    bad = r'[:\\/?*\[\]]'
    name = re.sub(bad, "-", (name or "")).strip()
    return (name[:31] or "Sheet1")


# -----------------------------
# Valent page parsing helpers
# -----------------------------

def _find_mark_and_model_on_page(page_text: str) -> tuple[str, str]:
    m_tag = re.search(r"Mark:\s*([^\n]+)", page_text, flags=re.IGNORECASE)
    m_model = re.search(r"Model:\s*([^\n]+)", page_text, flags=re.IGNORECASE)
    tag = _norm(m_tag.group(1)) if m_tag else ""
    model = _norm(m_model.group(1)) if m_model else ""
    return tag, model

def _get_blocks(page) -> list[tuple[float, float, float, float, str]]:
    out: list[tuple[float, float, float, float, str]] = []
    for b in page.get_text("blocks"):
        x0, y0, x1, y1, text = b[0], b[1], b[2], b[3], b[4]
        if text and _norm(text):
            out.append((x0, y0, x1, y1, text))
    return out

def _find_desc_block_flexible(blocks, header: str):
    """
    Handles:
      A) one big block starting with 'Accessories\\n...'
      B) 'Accessories' header block + body block below it
    Returns (x0,y0,x1,y1,text) for the BODY block
    """
    header_l = header.lower()

    # Case A: combined block starts with 'Accessories\n...'
    for (x0, y0, x1, y1, text) in blocks:
        t = (text or "").strip()
        if t.lower().startswith(header_l + "\n"):
            return (x0, y0, x1, y1, text)

    # Case B: header is its own block; body is the next block beneath it
    header_blocks = []
    for (x0, y0, x1, y1, text) in blocks:
        if _norm(text).lower() == header_l:
            header_blocks.append((x0, y0, x1, y1))

    if not header_blocks:
        return None

    hx0, hy0, hx1, hy1 = header_blocks[0]

    candidates = []
    for (x0, y0, x1, y1, text) in blocks:
        if y0 <= hy1:
            continue
        # same column band
        if abs(x0 - hx0) <= 40:
            candidates.append((y0, x0, y1, x1, text))

    if not candidates:
        return None

    candidates.sort(key=lambda t: t[0])  # nearest below
    y0, x0, y1, x1, text = candidates[0]
    return (x0, y0, x1, y1, text)

def _find_money_block_to_right(blocks, desc_bbox):
    """
    Select best money block to the right:
    - to the right of desc
    - vertically overlapping
    - highest count of $ matches; tie -> closest
    """
    dx0, dy0, dx1, dy1 = desc_bbox

    best = None
    best_score = -1
    best_dist = None

    for (x0, y0, x1, y1, text) in blocks:
        if "$" not in (text or ""):
            continue
        if x0 <= dx1:
            continue

        overlap = min(dy1, y1) - max(dy0, y0)
        if overlap <= 0:
            continue

        score = len(DOLLAR_ANY_RE.findall(text))
        dist = x0 - dx1

        if score > best_score or (score == best_score and (best_dist is None or dist < best_dist)):
            best = (x0, y0, x1, y1, text)
            best_score = score
            best_dist = dist

    return best

def _extract_dollar_spans_in_column(page, *, x_min: float, y_top: float, y_bottom: float) -> list[float]:
    """
    Fallback extractor: scan spans for $ amounts by geometry.
    Returns floats ordered top-to-bottom.
    """
    d = page.get_text("dict")
    hits: list[tuple[float, float]] = []
    for block in d.get("blocks", []):
        if block.get("type") != 0:
            continue
        for line in block.get("lines", []):
            for span in line.get("spans", []):
                txt = span.get("text", "")
                if "$" not in txt:
                    continue
                x0, y0, x1, y1 = span.get("bbox", (None, None, None, None))
                if x0 is None:
                    continue
                if x0 < x_min:
                    continue
                if y0 < y_top or y0 > y_bottom:
                    continue

                m = DOLLAR_ANY_RE.search(txt)
                if not m:
                    continue
                hits.append((y0, _money_to_float(m.group(1))))

    hits.sort(key=lambda t: t[0])
    return [p for _, p in hits]

def _split_table_lines(desc_text: str, header: str) -> list[str]:
    lines = [ln.rstrip() for ln in (desc_text or "").splitlines() if ln.strip()]
    if lines and _norm(lines[0]).lower() == header.lower():
        lines = lines[1:]
    return lines

def _extract_table_prices(page, blocks, desc_blk) -> list[float]:
    x0, y0, x1, y1, _ = desc_blk
    money_blk = _find_money_block_to_right(blocks, (x0, y0, x1, y1))
    if money_blk:
        amounts = DOLLAR_ANY_RE.findall(money_blk[4] or "")
        return [_money_to_float(a) for a in amounts]

    return _extract_dollar_spans_in_column(
        page,
        x_min=x1 + 5,
        y_top=y0 - 5,
        y_bottom=y1 + 5
    )

def _render_table_as_text(page, blocks, header: str, total_label: str) -> tuple[str, float | None]:
    """
    Render table WITHOUT prices in text, but still detect total value.
    """
    desc_blk = _find_desc_block_flexible(blocks, header)
    if not desc_blk:
        return "", None

    _, _, _, _, desc_text = desc_blk
    desc_lines = _split_table_lines(desc_text, header=header)

    prices = _extract_table_prices(page, blocks, desc_blk)

    total_val: float | None = None
    if prices and any(total_label.lower() in (ln or "").lower() for ln in desc_lines):
        total_val = prices[-1]

    out_lines = [f"{header}:"]
    item_lines = [ln for ln in desc_lines if total_label.lower() not in (ln or "").lower()]

    for ln in item_lines:
        out_lines.append(f"- {ln}")

    # include the total label but no $ amount
    if any(total_label.lower() in (ln or "").lower() for ln in desc_lines):
        out_lines.append(total_label)

    return "\n".join(out_lines).strip(), total_val


def _extract_main_feature_sections(page_text: str) -> str:
    """
    Capture from the earliest occurrence of any of:
      - General Product Features:
      - General Purpose Features:
      - Cooling System:
    Stop before Pricing (exclusive) if present.
    """
    t = page_text or ""
    headings = [
        "General Product Features:",
        "General Purpose Features:",
        "Cooling System:",
    ]

    starts = []
    for h in headings:
        idx = t.lower().find(h.lower())
        if idx != -1:
            starts.append(idx)

    if not starts:
        return ""

    start_idx = min(starts)

    pricing_idx = t.lower().find("\npricing")
    if pricing_idx == -1:
        pricing_idx = t.lower().find("pricing")

    if pricing_idx != -1 and pricing_idx > start_idx:
        return t[start_idx:pricing_idx].strip()

    return t[start_idx:].strip()


def _extract_general_purpose_plus_next_paragraphs(page_text: str) -> str:
    """
    Capture:
      - "General Purpose Features:" section
      - the next paragraph(s)
    Stop before "Pricing" (exclusive).
    """
    t = page_text or ""
    m = re.search(
        r"(General Purpose Features:\s*.*?)(?:\n\s*Pricing\b)",
        t,
        flags=re.IGNORECASE | re.DOTALL
    )
    if m:
        return m.group(1).strip()

    # Fallback: take just the features blob
    m2 = re.search(
        r"(General Purpose Features:\s*.*?)(?:\n\s*\n|\Z)",
        t,
        flags=re.IGNORECASE | re.DOTALL
    )
    if m2:
        return m2.group(1).strip()

    return ""


# -----------------------------
# VALENT PARSER (PDF bytes -> structured dict)
# -----------------------------

def parse_valent_pdf_bytes(pdf_bytes: bytes) -> dict[str, Any]:
    """
    Returns:
      {
        "manufacturer": "Valent",
        "header": {...},  # (keep empty for now - Valent sample doesn't show quote meta consistently)
        "lines": [
           {
             "model_code": "...",
             "tagging": "...",   # Mark
             "tags": [...],      # [Mark]
             "qty": 1,
             "list_each": <Unit Total or 0>,
             "total_list": <Unit Total or 0>,
             "options": [
                 {"feature": "Equipment", "desc": "...", "add_price": <Unit Total>},
                 {"feature": "Include", "desc": "...", "add_price": 0},
                 {"feature": "Include", "desc": "...", "add_price": 0},
             ]
           },
           ...
        ]
      }
    """
    doc = fitz.open(stream=pdf_bytes, filetype="pdf")

    header: dict[str, Any] = {}  # Valent PDFs vary; you can extend later

    lines_out: list[dict[str, Any]] = []

    for pi in range(len(doc)):
        page = doc[pi]
        page_text = page.get_text("text") or ""

        if "Mark:" not in page_text:
            continue

        mark, model = _find_mark_and_model_on_page(page_text)
        if not mark:
            continue

        blocks = _get_blocks(page)

        # 1) Features block (General Purpose + next paragraph(s))
        features_block = _extract_main_feature_sections(page_text).strip()

        # 2) Unit table rendered + Unit Total (used for main line Option Price)
        unit_text, unit_total = _render_table_as_text(
            page=page, blocks=blocks, header="Unit", total_label="Unit Total"
        )

        # 3) Accessories include (optional)
        acc_text, _ = _render_table_as_text(
            page=page, blocks=blocks, header="Accessories", total_label="Accessories Total"
        )

        # 4) Warranties include (optional)
        war_text, _ = _render_table_as_text(
            page=page, blocks=blocks, header="Warranties", total_label="Warranty Total"
        )

        # Build Main Equipment Description:
        #   - General Purpose Features + next paragraph(s)
        #   - everything under Unit
        main_desc_parts: list[str] = []
        if features_block:
            main_desc_parts.append(features_block)
        if unit_text.strip():
            main_desc_parts.append(unit_text.strip())
        main_desc = "\n\n".join([p for p in main_desc_parts if p]).strip()

        # Options payload (to feed sheet writer consistently)
        opts: list[dict[str, Any]] = []

        # Main equipment line (always)
        opts.append({
            "feature": "Equipment",
            "desc": main_desc if main_desc else f"{mark} | {model}".strip(" |"),
            "add_price": float(unit_total or 0.0),
        })

        # Accessories include line (only if present)
        if acc_text.strip():
            opts.append({
                "feature": "Include",
                "desc": acc_text.strip(),
                "add_price": 0.0,
            })

        # Warranties include line (only if present)
        if war_text.strip():
            opts.append({
                "feature": "Include",
                "desc": war_text.strip(),
                "add_price": 0.0,
            })

        # This parser treats each Mark as a “line”
        lines_out.append({
            "model_code": model,
            "tagging": mark,
            "tags": [mark],
            "qty": 1,
            "list_each": float(unit_total or 0.0),
            "total_list": float(unit_total or 0.0),
            "options": opts
        })

    return {
        "manufacturer": "Valent",
        "header": header,
        "lines": lines_out
    }


# -----------------------------
# TEMPLATE WRITER (same pattern as KCC)
# -----------------------------

def _write_one_sheet(
    ws,
    *,
    manufacturer: str,
    equipment: str,
    model_code: str,
    tags: list[str],
    header: dict[str, Any],
    options: list[dict[str, Any]],
    output_type: str = "shopping_list",
) -> None:
    """
    Same workbook feel as KCC:
      - Row 1 headers A-F
      - Row 2 values
      - Row 3 option headers starting at C (no gap row)
      - Rows 4+: option rows (shopping_list) OR empty (all_in_one)

    For Valent:
      - shopping_list: we still output option rows, but ONLY 1 Equipment row + optional Include rows
      - all_in_one: place the main equipment description in E2; and also append include descriptions below it.
                   (table rows left empty)
    """
    output_type = (output_type or "shopping_list").strip().lower()

    # Row 1
    top_headers = [
        "Equipment",
        "Manufacturer",
        "Model",
        "Part Number",
        "Description (Not Overwritten)",
        "Notes (Not Overwritten)",
    ]
    for col, h in enumerate(top_headers, start=1):
        ws.cell(row=1, column=col, value=h)

    # Row 2
    model_short = model_code.split("-")[0] if model_code else ""
    ws.cell(row=2, column=1, value=equipment)
    ws.cell(row=2, column=2, value=manufacturer)
    ws.cell(row=2, column=3, value=model_short)
    ws.cell(row=2, column=4, value=model_code)

    tags_label = ", ".join(t for t in (tags or []) if t) if tags else ""

    if output_type == "all_in_one":
    # E2 = tags + MAIN equipment description only (no include text)
        desc_parts: list[str] = []
        if tags_label:
            desc_parts.append(tags_label)
            desc_parts.append("")
    
        main_desc = ""
        for opt in (options or []):
            if (opt.get("feature") or "").strip().lower() == "equipment":
                main_desc = (opt.get("desc") or "").strip()
                break
    
        if main_desc:
            desc_parts.append(main_desc)
    
        ws.cell(row=2, column=5, value="\n".join(desc_parts).strip())
        ws.cell(row=2, column=6, value="")
    
    else:
        # shopping_list mode: keep E2 simple like KCC (UNCHANGED)
        ws.cell(row=2, column=5, value=tags_label)
        ws.cell(row=2, column=6, value="")

    # Row 3 headers starting at C
    opt_headers = [
        "Tag", "Part Number", "Feature", "Description",
        "Qty", "List Price", "LP Ext.", "Buy Mult.", "Net Price", "Markup", "Margin", "Sell Price",
        "Weight", "Freight", "Fr. Multi.", "Alignment", "Subtotal", "Option Price"
    ]
    start_col = 3
    for i, h in enumerate(opt_headers):
        ws.cell(row=3, column=start_col + i, value=h)

    def z() -> int:
        return 0

    option_qty = 1
    row_idx = 4

    if output_type == "all_in_one":
        # ✅ all_in_one: write ONLY include rows (Accessories/Warranties) in the lower table
        for tag in (tags or [""]):
            for opt in (options or []):
                feat = (opt.get("feature") or "").strip().lower()
                if feat != "include":
                    continue
    
                desc = (opt.get("desc") or "").strip()
    
                option_price = 0.0   # force $0 for includes
                lp_ext = option_price * option_qty
    
                ws.cell(row=row_idx, column=3, value=tag)            # Tag
                ws.cell(row=row_idx, column=4, value="")             # Part Number
                ws.cell(row=row_idx, column=5, value="Include")      # Feature
                ws.cell(row=row_idx, column=6, value=desc)           # Description
    
                ws.cell(row=row_idx, column=7, value=option_qty)     # Qty
                ws.cell(row=row_idx, column=8, value=option_price)   # List Price
                ws.cell(row=row_idx, column=9, value=lp_ext)         # LP Ext.
                ws.cell(row=row_idx, column=10, value=z())
                ws.cell(row=row_idx, column=11, value=z())
                ws.cell(row=row_idx, column=12, value=z())
                ws.cell(row=row_idx, column=13, value=z())
                ws.cell(row=row_idx, column=14, value=z())
                ws.cell(row=row_idx, column=15, value=z())
                ws.cell(row=row_idx, column=16, value=z())
                ws.cell(row=row_idx, column=17, value=z())
                ws.cell(row=row_idx, column=18, value=z())
                ws.cell(row=row_idx, column=19, value=z())
                ws.cell(row=row_idx, column=20, value=option_price)
    
                row_idx += 1

    else:
        # ✅ shopping_list: DO NOT CHANGE ANYTHING (your existing behavior)
        for tag in (tags or [""]):
            for opt in (options or []):
                feature = (opt.get("feature") or "Include").strip() or "Include"
                desc = (opt.get("desc") or "").strip()
                option_price = float(opt.get("add_price") or 0.0)
                lp_ext = option_price * option_qty
    
                ws.cell(row=row_idx, column=3, value=tag)            # Tag
                ws.cell(row=row_idx, column=4, value="")             # Part Number
                ws.cell(row=row_idx, column=5, value=feature)        # Feature (Equipment or Include)
                ws.cell(row=row_idx, column=6, value=desc)           # Description
    
                ws.cell(row=row_idx, column=7, value=option_qty)     # Qty
                ws.cell(row=row_idx, column=8, value=option_price)   # List Price
                ws.cell(row=row_idx, column=9, value=lp_ext)         # LP Ext.
                ws.cell(row=row_idx, column=10, value=z())           # Buy Mult.
                ws.cell(row=row_idx, column=11, value=z())           # Net Price
                ws.cell(row=row_idx, column=12, value=z())           # Markup
                ws.cell(row=row_idx, column=13, value=z())           # Margin
                ws.cell(row=row_idx, column=14, value=z())           # Sell Price
                ws.cell(row=row_idx, column=15, value=z())           # Weight
                ws.cell(row=row_idx, column=16, value=z())           # Freight
                ws.cell(row=row_idx, column=17, value=z())           # Fr. Multi.
                ws.cell(row=row_idx, column=18, value=z())           # Alignment
                ws.cell(row=row_idx, column=19, value=z())           # Subtotal
                ws.cell(row=row_idx, column=20, value=option_price)  # Option Price
    
                row_idx += 1


    # Formatting (same vibe as KCC)
    header_fill = PatternFill("solid", fgColor="D9EAD3")
    top_header_font = Font(bold=True, italic=True)
    opt_header_font = Font(bold=True)

    for c in range(1, 7):
        cell = ws.cell(row=1, column=c)
        cell.font = top_header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    for c in range(3, 21):
        cell = ws.cell(row=3, column=c)
        cell.font = opt_header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    for r in range(2, max(row_idx, 4)):
        for c in range(1, 21):
            ws.cell(row=r, column=c).alignment = Alignment(vertical="top", wrap_text=True)

    widths = {
        1: 18, 2: 18, 3: 16, 4: 28, 5: 40, 6: 55,
        7: 8, 8: 12, 9: 12, 10: 10, 11: 12, 12: 10, 13: 10, 14: 12,
        15: 10, 16: 10, 17: 10, 18: 10, 19: 12, 20: 12
    }
    for col, w in widths.items():
        ws.column_dimensions[get_column_letter(col)].width = w

    ws.freeze_panes = "A4"


def write_valent_template_workbook(parsed: dict[str, Any], output_type: str = "shopping_list") -> Workbook:
    """
    Creates ONE workbook and ONE sheet per (equipment, model_code) – same grouping behavior as KCC.
    Valent lines are typically one Mark per page, but this supports multiples.
    """
    wb = Workbook()
    wb.remove(wb.active)

    manufacturer = parsed.get("manufacturer", "Valent")
    header = parsed.get("header", {})
    lines = parsed.get("lines", []) or []

    if not lines:
        ws = wb.create_sheet("Valent Import")
        _write_one_sheet(
            ws,
            manufacturer=manufacturer,
            equipment="",
            model_code="",
            tags=[],
            header=header,
            options=[],
            output_type=output_type,
        )
        return wb

    # Group by (equipment, model_code)
    groups: dict[tuple[str, str], dict[str, Any]] = {}

    for ln in lines:
        model_code = ln.get("model_code", "") or ""
        tagging = ln.get("tagging", "") or ""
        tags = ln.get("tags") or ([tagging] if tagging else [""])
        options = ln.get("options", []) or []

        equipment = "RTU" if tagging else ""

        key = (equipment, model_code)
        if key not in groups:
            groups[key] = {
                "equipment": equipment,
                "model_code": model_code,
                "tags": [],
                # For Valent, options are per Mark; but sheet is per model.
                # We'll APPEND options blocks so you don't lose content.
                "options": [],
            }

        groups[key]["tags"].extend(tags)
        groups[key]["options"].extend(options)

    # Deduplicate tags per group (preserve order)
    for _, info in groups.items():
        seen: set[str] = set()
        deduped_tags: list[str] = []
        for t in info["tags"]:
            if t and t not in seen:
                seen.add(t)
                deduped_tags.append(t)
        info["tags"] = deduped_tags

    # De-dupe options by (feature, desc, price)
    for _, info in groups.items():
        seen_opt: set[tuple[str, str, float]] = set()
        uniq_opts: list[dict[str, Any]] = []
        for opt in info["options"]:
            feat = _norm(str(opt.get("feature") or ""))
            desc = _norm(str(opt.get("desc") or ""))
            price = float(opt.get("add_price") or 0.0)
            key = (feat, desc, round(price, 2))
            if not desc:
                continue
            if key in seen_opt:
                continue
            seen_opt.add(key)
            uniq_opts.append({"feature": feat or "Include", "desc": opt.get("desc") or "", "add_price": price})
        info["options"] = uniq_opts

    # Build sheets
    sheet_names_seen: set[str] = set()

    for (equipment, model_code), info in groups.items():
        base_name = _safe_sheet_name(model_code or equipment or "Valent Import")
        name = base_name
        n = 2
        while name in sheet_names_seen:
            suffix = f"_{n}"
            name = _safe_sheet_name(base_name[: 31 - len(suffix)] + suffix)
            n += 1
        sheet_names_seen.add(name)

        ws = wb.create_sheet(title=name)

        _write_one_sheet(
            ws,
            manufacturer=manufacturer,
            equipment=equipment,
            model_code=model_code,
            tags=info["tags"],
            header=header,
            options=info["options"],
            output_type=output_type,
        )

    return wb


# -----------------------------
# RUNNER (PDF bytes -> XLSX bytes)
# -----------------------------

def convert_valent_pdf_to_xlsx_bytes(
    pdf_bytes: bytes,
    job_name: str | None = None,
    output_type: str = "all_in_one",
) -> tuple[bytes, str]:
    """
    output_type:
      - "all_in_one"     -> puts main + include blocks into E2; leaves table empty
      - "shopping_list"  -> writes option rows (Equipment + Include + Include) like KCC tables
    """
    job = (job_name or "").strip() or "job"
    parsed = parse_valent_pdf_bytes(pdf_bytes)

    out_type = (output_type or "all_in_one").strip().lower()
    if out_type not in ("all_in_one", "shopping_list"):
        raise ValueError(f"Unsupported output_type: {output_type}")

    wb = write_valent_template_workbook(parsed, output_type=out_type)

    sheet_names = [ws.title for ws in wb.worksheets]
    logger.info("VALENT DEBUG: output_type=%s workbook sheets=%s", out_type, sheet_names)

    out = io.BytesIO()
    wb.save(out)
    xlsx_bytes = out.getvalue()
    out.close()

    filename = f"{job}_{out_type}_template_output.xlsx"
    return xlsx_bytes, filename




