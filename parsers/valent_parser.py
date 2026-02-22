import re
import fitz  # PyMuPDF
from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font, Alignment, PatternFill


# -----------------------------
# Helpers
# -----------------------------

DOLLAR_ANY_RE = re.compile(r"\$\s*([0-9]{1,3}(?:,[0-9]{3})*(?:\.[0-9]{2})?)")

def _norm(s: str) -> str:
    return re.sub(r"\s+", " ", (s or "")).strip()

def _money_to_float(s: str) -> float:
    return float(s.replace(",", "").strip())

def _safe_sheet_name(name: str) -> str:
    bad = r'[:\\/?*\[\]]'
    name = re.sub(bad, "-", (name or "")).strip()
    return (name[:31] or "Sheet1")


# -----------------------------
# Core Valent extraction (block pairing + span fallback)
# -----------------------------

def _find_mark_and_model_on_page(page_text: str):
    m_tag = re.search(r"Mark:\s*([^\n]+)", page_text, flags=re.IGNORECASE)
    m_model = re.search(r"Model:\s*([^\n]+)", page_text, flags=re.IGNORECASE)
    tag = _norm(m_tag.group(1)) if m_tag else ""
    model = _norm(m_model.group(1)) if m_model else ""
    return tag, model

def _get_blocks(page):
    # blocks: (x0, y0, x1, y1, text, block_no, block_type)
    out = []
    for b in page.get_text("blocks"):
        x0, y0, x1, y1, text = b[0], b[1], b[2], b[3], b[4]
        if text and _norm(text):
            out.append((x0, y0, x1, y1, text))
    return out

def _find_desc_block_flexible(blocks, header: str):
    """
    Handles both layouts:
      A) one big block: 'Accessories\\n...'
      B) header block 'Accessories' + body block below it
    Returns (x0, y0, x1, y1, text) for the BODY block that contains lines.
    """
    header_l = header.lower()

    # Case A: combined block starts with 'Accessories\\n...'
    for (x0, y0, x1, y1, text) in blocks:
        t = (text or "").strip()
        if t.lower().startswith(header_l + "\n"):
            return (x0, y0, x1, y1, text)

    # Case B: header is its own block; body is the next block beneath it
    header_blocks = []
    for (x0, y0, x1, y1, text) in blocks:
        if _norm(text).lower() == header_l:
            header_blocks.append((x0, y0, x1, y1))

    if not header_blocks:
        return None

    # Take the first matching header block
    hx0, hy0, hx1, hy1 = header_blocks[0]

    # Find the nearest block BELOW the header in the same column band
    candidates = []
    for (x0, y0, x1, y1, text) in blocks:
        if y0 <= hy1:
            continue
        # same column: close x0 to header x0
        if abs(x0 - hx0) <= 40:
            candidates.append((y0, x0, y1, x1, text))

    if not candidates:
        return None

    candidates.sort(key=lambda t: t[0])  # nearest below
    y0, x0, y1, x1, text = candidates[0]
    return (x0, y0, x1, y1, text)

def _find_money_block_to_right(blocks, desc_bbox):
    """
    Select the best money block to the right:
    - must be to the right of desc
    - must overlap vertically
    - choose the one with the MOST '$' amounts (robust vs wrong blocks)
    """
    dx0, dy0, dx1, dy1 = desc_bbox

    best = None
    best_score = -1
    best_dist = None

    for (x0, y0, x1, y1, text) in blocks:
        if "$" not in (text or ""):
            continue

        # to the right
        if x0 <= dx1:
            continue

        # vertical overlap
        overlap = min(dy1, y1) - max(dy0, y0)
        if overlap <= 0:
            continue

        # scoring: how many dollar values are in this block?
        score = len(DOLLAR_ANY_RE.findall(text))
        dist = x0 - dx1

        # pick highest score; tie-breaker: closest distance
        if score > best_score or (score == best_score and (best_dist is None or dist < best_dist)):
            best = (x0, y0, x1, y1, text)
            best_score = score
            best_dist = dist

    return best

def _extract_lines_and_prices(desc_text: str, money_text: str, header: str, total_label: str):
    """
    desc_text: block with header + description lines + Total line
    money_text: block with $ values (including total)
    We:
      - remove header line (Unit / Accessories)
      - remove total line (Unit Total / Accessories Total)
      - find all $ amounts in money_text
      - drop the last amount as the total (best assumption for these tables)
      - zip desc_lines to amounts
    """
    desc_lines = [ln.strip() for ln in desc_text.splitlines() if ln.strip()]
    if desc_lines and desc_lines[0].lower() == header.lower():
        desc_lines = desc_lines[1:]

    desc_lines = [ln for ln in desc_lines if ln.lower() != total_label.lower() and total_label.lower() not in ln.lower()]

    amounts = DOLLAR_ANY_RE.findall(money_text or "")
    prices = [_money_to_float(a) for a in amounts]

    prices_no_total = prices[:-1] if prices else []
    n = min(len(desc_lines), len(prices_no_total))
    return [(desc_lines[i], prices_no_total[i]) for i in range(n)]

def _extract_dollar_spans_in_column(page, *, x_min: float, y_top: float, y_bottom: float):
    """
    Fallback extractor: pull $ amounts from PDF spans (not blocks), by geometry.
    Returns list of floats sorted by y (top-to-bottom).
    """
    d = page.get_text("dict")
    hits = []
    for block in d.get("blocks", []):
        if block.get("type") != 0:
            continue
        for line in block.get("lines", []):
            for span in line.get("spans", []):
                txt = span.get("text", "")
                if "$" not in txt:
                    continue
                x0, y0, x1, y1 = span.get("bbox", (None, None, None, None))
                if x0 is None:
                    continue
                if x0 < x_min:
                    continue
                if y0 < y_top or y0 > y_bottom:
                    continue

                m = DOLLAR_ANY_RE.search(txt)
                if not m:
                    continue
                hits.append((y0, _money_to_float(m.group(1))))

    hits.sort(key=lambda t: t[0])
    return [p for _, p in hits]

def _extract_lines_and_prices_from_desc_and_prices(desc_text: str, prices: list[float], header: str, total_label: str):
    """
    Same as _extract_lines_and_prices(), but uses an already-extracted list of prices.
    Includes a couple guardrails for mismatched totals.
    """
    desc_lines = [ln.strip() for ln in desc_text.splitlines() if ln.strip()]
    if desc_lines and desc_lines[0].lower() == header.lower():
        desc_lines = desc_lines[1:]

    desc_lines = [ln for ln in desc_lines if ln.lower() != total_label.lower() and total_label.lower() not in ln.lower()]

    # If prices look like they include a total, drop last one when it helps alignment.
    if len(prices) == len(desc_lines) + 1:
        prices = prices[:-1]
    elif len(prices) > len(desc_lines) + 1:
        # overly chatty span capture (rare): keep the first N that match desc lines
        prices = prices[:len(desc_lines)]

    n = min(len(desc_lines), len(prices))
    return [(desc_lines[i], prices[i]) for i in range(n)]

def _get_table_rows_with_fallback(page, blocks, header: str, total_label: str):
    """
    KEEP EXISTING behavior:
      1) Try block-paired money extraction (current script behavior)
      2) If that yields nothing, fallback to span-based $ extraction (fixes RTU-1-2 edge case)
    Returns list of (desc, price).
    """
    desc_blk = _find_desc_block_flexible(blocks, header)
    if not desc_blk:
        return []

    x0, y0, x1, y1, text = desc_blk

    # 1) current approach: money block to right
    money_blk = _find_money_block_to_right(blocks, (x0, y0, x1, y1))
    if money_blk:
        rows = _extract_lines_and_prices(
            desc_text=text,
            money_text=money_blk[4],
            header=header,
            total_label=total_label
        )
        if rows:
            return rows

    # 2) fallback: get $ from spans to the right of the desc block
    prices = _extract_dollar_spans_in_column(
        page,
        x_min=x1 + 5,
        y_top=y0 - 5,
        y_bottom=y1 + 5
    )
    if not prices:
        return []

    return _extract_lines_and_prices_from_desc_and_prices(
        desc_text=text,
        prices=prices,
        header=header,
        total_label=total_label
    )


def parse_valent_pdf(pdf_path: str):
    doc = fitz.open(pdf_path)
    units = []

    for pi in range(len(doc)):
        page = doc[pi]
        page_text = page.get_text("text")

        if "Mark:" not in page_text:
            continue

        tag, model = _find_mark_and_model_on_page(page_text)
        if not tag:
            continue

        blocks = _get_blocks(page)

        # --- UNIT table (Feature=Equipment) ---
        unit_rows = _get_table_rows_with_fallback(
            page=page,
            blocks=blocks,
            header="Unit",
            total_label="Unit Total"
        )

        # --- ACCESSORIES table (Feature=Include) ---
        acc_rows = _get_table_rows_with_fallback(
            page=page,
            blocks=blocks,
            header="Accessories",
            total_label="Accessories Total"
        )

        # Build template rows
        rows = []
        for desc, price in unit_rows:
            rows.append({
                "tag": tag,
                "feature": "Equipment",
                "description": desc,
                "option_price": float(price or 0)
            })

        for desc, price in acc_rows:
            rows.append({
                "tag": tag,
                "feature": "Include",
                "description": desc,
                "option_price": float(price or 0)
            })

        # De-dupe
        seen = set()
        uniq = []
        for r in rows:
            key = (r["tag"], r["feature"], r["description"], round(r["option_price"], 2))
            if key in seen:
                continue
            seen.add(key)
            uniq.append(r)

        units.append({
            "tag": tag,
            "model": model,
            "part_number": model,
            "rows": uniq
        })

    return units


# -----------------------------
# Write workbook (one sheet per Mark)
# -----------------------------

def write_valent_workbook(units, out_xlsx_path: str):
    wb = Workbook()
    wb.remove(wb.active)

    header_fill = PatternFill("solid", fgColor="D9EAD3")
    top_header_font = Font(bold=True, italic=True)
    opt_header_font = Font(bold=True)

    for u in units:
        ws = wb.create_sheet(title=_safe_sheet_name(u["tag"] or u["model"] or "Unit"))

        # Row 1 headers A-F
        top_headers = [
            "Equipment", "Manufacturer", "Model", "Part Number",
            "Description (Not Overwritten)", "Notes (Not Overwritten)"
        ]
        for col, h in enumerate(top_headers, start=1):
            ws.cell(row=1, column=col, value=h)

        # Row 2 values
        equipment = "RTU"
        manufacturer = "Valent"
        model_code = u.get("model", "")
        model_short = model_code.split("-")[0] if model_code else ""

        ws.cell(row=2, column=1, value=equipment)
        ws.cell(row=2, column=2, value=manufacturer)
        ws.cell(row=2, column=3, value=model_short)
        ws.cell(row=2, column=4, value=model_code)
        ws.cell(row=2, column=5, value=f"{u.get('tag','')} | {model_code}".strip(" |"))
        ws.cell(row=2, column=6, value="")

        # Row 3 option headers (NO GAP ROW) start at C
        opt_headers = [
            "Tag", "Part Number", "Feature", "Description",
            "Qty", "List Price", "LP Ext.", "Buy Mult.", "Net Price", "Markup", "Margin", "Sell Price",
            "Weight", "Freight", "Fr. Multi.", "Alignment", "Subtotal", "Option Price"
        ]
        start_col = 3
        for i, h in enumerate(opt_headers):
            ws.cell(row=3, column=start_col + i, value=h)

        def z(): return 0
        qty = 1
        row_idx = 4

        for r in u.get("rows", []):
            tag = r.get("tag", "")
            feature = r.get("feature", "Include")
            desc = r.get("description", "")
            price = float(r.get("option_price") or 0)

            ws.cell(row=row_idx, column=3, value=tag)
            ws.cell(row=row_idx, column=4, value="")
            ws.cell(row=row_idx, column=5, value=feature)
            ws.cell(row=row_idx, column=6, value=desc)

            ws.cell(row=row_idx, column=7, value=qty)
            ws.cell(row=row_idx, column=8, value=price)
            ws.cell(row=row_idx, column=9, value=price)

            ws.cell(row=row_idx, column=10, value=z())
            ws.cell(row=row_idx, column=11, value=z())
            ws.cell(row=row_idx, column=12, value=z())
            ws.cell(row=row_idx, column=13, value=z())
            ws.cell(row=row_idx, column=14, value=z())
            ws.cell(row=row_idx, column=15, value=z())
            ws.cell(row=row_idx, column=16, value=z())
            ws.cell(row=row_idx, column=17, value=z())
            ws.cell(row=row_idx, column=18, value=z())
            ws.cell(row=row_idx, column=19, value=z())
            ws.cell(row=row_idx, column=20, value=price)

            row_idx += 1

        # Minimal formatting (NO borders)
        for c in range(1, 7):
            cell = ws.cell(row=1, column=c)
            cell.font = top_header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

        for c in range(3, 21):
            cell = ws.cell(row=3, column=c)
            cell.font = opt_header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

        for rr in range(2, row_idx):
            for cc in range(1, 21):
                ws.cell(row=rr, column=cc).alignment = Alignment(vertical="top", wrap_text=True)

        widths = {
            1: 18, 2: 18, 3: 16, 4: 28, 5: 20, 6: 55,
            7: 8, 8: 12, 9: 12, 10: 10, 11: 12, 12: 10, 13: 10, 14: 12,
            15: 10, 16: 10, 17: 10, 18: 10, 19: 12, 20: 12
        }
        for col, w in widths.items():
            ws.column_dimensions[get_column_letter(col)].width = w

        ws.freeze_panes = "A4"

    wb.save(out_xlsx_path)


def convert_valent_pdf_to_template(pdf_path: str, out_xlsx_path: str):
    units = parse_valent_pdf(pdf_path)
    write_valent_workbook(units, out_xlsx_path)


if __name__ == "__main__":
    pdf_path = r"Valent - Test Configuration.pdf"
    out_path = r"Valent_template_output.xlsx"
    convert_valent_pdf_to_template(pdf_path, out_path)
    print(f"Wrote: {out_path}")



