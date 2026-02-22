# helpers_schedule_textract_0_1.py
from __future__ import annotations

import os
import re
import time
import json
import base64
import shutil
import zipfile
import logging
import mimetypes
import threading
from typing import Any, Dict, List, Tuple, Optional
from datetime import datetime

import boto3
from botocore.config import Config as BotoConfig
from werkzeug.utils import secure_filename

from openpyxl import Workbook
from openpyxl.utils import get_column_letter

from concurrent.futures import ThreadPoolExecutor


# =========================================================
# Env / Config
# =========================================================
BASE_DIR = os.path.dirname(os.path.abspath(__file__))

AWS_REGION = os.getenv("AWS_REGION", "us-east-1")
S3_BUCKET = os.getenv("S3_BUCKET", "ays-artifacts-betatwo")

S3_SCH_UPLOAD_PREFIX = os.getenv("S3_SCH_UPLOAD_PREFIX", "sch_upload")
S3_SCH_PROCESSED_PREFIX = os.getenv("S3_SCH_PROCESSED_PREFIX", "sch_processed")

APP_PUBLIC_BASE = os.getenv("APP_PUBLIC_BASE", "").rstrip("/")

SCH_UPLOAD_LOCAL = os.getenv("SCH_UPLOAD_LOCAL", os.path.join(BASE_DIR, "sch_uploads"))
SCH_PROCESSED_LOCAL = os.getenv("SCH_PROCESSED_LOCAL", os.path.join(BASE_DIR, "sch_processed"))

os.makedirs(SCH_UPLOAD_LOCAL, exist_ok=True)
os.makedirs(SCH_PROCESSED_LOCAL, exist_ok=True)

# If you ever want raw textract json in the zip, set env to "1"
INCLUDE_RAW_TEXTRACT_JSON = os.getenv("SCH_INCLUDE_RAW_TEXTRACT_JSON", "0") == "1"

EXECUTOR = ThreadPoolExecutor(max_workers=max(os.cpu_count() or 4, 4))

JOBS: Dict[str, Dict[str, Any]] = {}
JOBS_LOCK = threading.Lock()


# =========================================================
# AWS Clients
# =========================================================
_s3 = boto3.client(
    "s3",
    region_name=AWS_REGION,
    config=BotoConfig(
        s3={"addressing_style": "virtual"},
        retries={"max_attempts": 10, "mode": "standard"},
    ),
)

_textract = boto3.client(
    "textract",
    region_name=AWS_REGION,
    config=BotoConfig(retries={"max_attempts": 10, "mode": "standard"}),
)


# =========================================================
# S3 Utilities
# =========================================================
def s3_key(*parts) -> str:
    return "/".join(str(p).strip("/\\") for p in parts if p)


def s3_upload_file(local_path: str, key: str) -> None:
    _s3.upload_file(local_path, S3_BUCKET, key)


def s3_upload_bytes(data: bytes, key: str) -> None:
    ct, _ = mimetypes.guess_type(key)
    _s3.put_object(
        Bucket=S3_BUCKET,
        Key=key,
        Body=data,
        ContentType=ct or "application/octet-stream",
    )


def s3_presign_get(key: str, expires: int = 3600, extra: dict | None = None) -> str:
    params = {"Bucket": S3_BUCKET, "Key": key}
    if extra:
        params.update(extra)
    return _s3.generate_presigned_url("get_object", Params=params, ExpiresIn=expires)


def s3_ensure_prefix(prefix: str) -> None:
    prefix = (prefix or "").strip().strip("/\\")
    if not prefix:
        return
    key = prefix + "/"
    try:
        _s3.put_object(Bucket=S3_BUCKET, Key=key, Body=b"")
    except Exception:
        # not fatal
        pass


# =========================================================
# General Helpers
# =========================================================
def slugify(s: str, max_length: int = 80) -> str:
    s = (s or "").strip()
    if not s:
        return "project"
    s = re.sub(r"[^\w\s-]", "", s, flags=re.UNICODE)
    s = re.sub(r"[-\s]+", "_", s).strip("_")
    return (s or "project")[:max_length]


def make_project_id(subject: str) -> str:
    base = slugify(subject or "Project")
    stamp = datetime.now().strftime("%Y%m%d-%H%M%S")
    return f"{base}_SCH-{stamp}"


def build_internal_doc_slug(attachment_name: str, submitted_at: str | None = None) -> str:
    if not attachment_name:
        attachment_name = "schedule.pdf"

    base, _ = os.path.splitext(attachment_name)
    base = (base or "schedule").strip()
    slug_base = slugify(base, max_length=80) or "schedule"

    try:
        if submitted_at:
            dt = datetime.fromisoformat(submitted_at.replace("Z", ""))
        else:
            dt = datetime.utcnow()
    except Exception:
        dt = datetime.utcnow()

    ts = dt.strftime("%Y%m%d-%H%M%S")
    return f"{slug_base}_{ts}"


def _safe_rmtree(path: str) -> None:
    try:
        if path and os.path.isdir(path):
            shutil.rmtree(path, ignore_errors=True)
    except Exception:
        pass


# =========================================================
# Job Manager
# =========================================================
def submit_job(fn, *args, **kwargs) -> str:
    import uuid
    job_id = str(uuid.uuid4())
    with JOBS_LOCK:
        JOBS[job_id] = {"state": "QUEUED", "created": time.time(), "info": {}}
    fut = EXECUTOR.submit(fn, job_id, *args, **kwargs)
    with JOBS_LOCK:
        JOBS[job_id]["future"] = fut
    return job_id


def set_job(job_id: str, **fields) -> None:
    with JOBS_LOCK:
        if job_id in JOBS:
            JOBS[job_id].update(fields)


def get_job(job_id: str) -> Dict[str, Any]:
    with JOBS_LOCK:
        return JOBS.get(job_id, {})


# =========================================================
# Textract Async (multi-page safe)
# =========================================================
def textract_start_tables_job_s3(*, bucket: str, key: str) -> str:
    resp = _textract.start_document_analysis(
        DocumentLocation={"S3Object": {"Bucket": bucket, "Name": key}},
        FeatureTypes=["TABLES"],
    )
    return resp["JobId"]


def textract_get_all_pages(*, job_id: str, timeout_s: int = 900, poll_s: float = 1.5) -> List[dict]:
    """
    Poll until SUCCEEDED/FAILED, then paginate all results.
    Minimal logging; no geometry printing.
    """
    t0 = time.time()

    # poll status
    last_status = None
    while True:
        if time.time() - t0 > timeout_s:
            raise TimeoutError(f"Textract job timed out after {timeout_s}s (job_id={job_id})")

        r = _textract.get_document_analysis(JobId=job_id)
        status = r.get("JobStatus", "")

        if status != last_status:
            last_status = status  # only step-change style
        if status == "SUCCEEDED":
            break
        if status == "FAILED":
            msg = r.get("StatusMessage", "")
            raise RuntimeError(f"Textract FAILED job_id={job_id}. {msg}")

        time.sleep(poll_s)

    # paginate
    pages: List[dict] = []
    next_token: Optional[str] = None
    first = True
    while True:
        if first:
            resp = _textract.get_document_analysis(JobId=job_id)
            first = False
        else:
            if not next_token:
                break
            resp = _textract.get_document_analysis(JobId=job_id, NextToken=next_token)

        pages.append(resp)
        next_token = resp.get("NextToken")
        if not next_token:
            break

    return pages


# =========================================================
# Textract Table Parsing
# =========================================================
def _build_block_maps(pages: List[dict]) -> Dict[str, dict]:
    block_map: Dict[str, dict] = {}
    for p in pages:
        for b in p.get("Blocks", []) or []:
            bid = b.get("Id")
            if bid:
                block_map[bid] = b
    return block_map


def _get_text_for_cell(cell_block: dict, block_map: Dict[str, dict]) -> str:
    text_parts: List[str] = []
    for rel in cell_block.get("Relationships", []) or []:
        if rel.get("Type") != "CHILD":
            continue
        for cid in rel.get("Ids", []) or []:
            cb = block_map.get(cid)
            if not cb:
                continue
            bt = cb.get("BlockType")
            if bt == "WORD":
                t = cb.get("Text", "")
                if t:
                    text_parts.append(t)
            elif bt == "SELECTION_ELEMENT":
                if cb.get("SelectionStatus") == "SELECTED":
                    text_parts.append("☑")
    return " ".join(text_parts).strip()


def _table_to_grid(table_block: dict, block_map: Dict[str, dict]) -> Tuple[List[List[str]], int]:
    page_num = int(table_block.get("Page", 0) or 0)
    cells: List[dict] = []

    for rel in table_block.get("Relationships", []) or []:
        if rel.get("Type") != "CHILD":
            continue
        for cid in rel.get("Ids", []) or []:
            cb = block_map.get(cid)
            if cb and cb.get("BlockType") == "CELL":
                cells.append(cb)

    if not cells:
        return [[]], page_num

    max_row = max(int(c.get("RowIndex", 1) or 1) for c in cells)
    max_col = max(int(c.get("ColumnIndex", 1) or 1) for c in cells)

    grid: List[List[str]] = [["" for _ in range(max_col)] for __ in range(max_row)]

    for c in cells:
        r = int(c.get("RowIndex", 1) or 1) - 1
        col = int(c.get("ColumnIndex", 1) or 1) - 1
        txt = _get_text_for_cell(c, block_map)

        rs = int(c.get("RowSpan", 1) or 1)
        cs = int(c.get("ColumnSpan", 1) or 1)

        for rr in range(r, min(r + rs, max_row)):
            for cc in range(col, min(col + cs, max_col)):
                if not grid[rr][cc]:
                    grid[rr][cc] = txt

    def row_has_data(row: List[str]) -> bool:
        return any((x or "").strip() for x in row)

    while grid and not row_has_data(grid[-1]):
        grid.pop()

    if grid:
        last_col = len(grid[0]) - 1
        while last_col >= 0:
            if any((grid[r][last_col] or "").strip() for r in range(len(grid))):
                break
            last_col -= 1
        if last_col >= 0:
            grid = [row[: last_col + 1] for row in grid]

    return grid, page_num


def extract_tables_by_page(pages: List[dict]) -> Dict[int, List[List[List[str]]]]:
    block_map = _build_block_maps(pages)

    tables_by_page: Dict[int, List[List[List[str]]]] = {}

    table_blocks = [b for b in block_map.values() if b.get("BlockType") == "TABLE"]

    def _geom_y(b: dict) -> float:
        try:
            return float(b.get("Geometry", {}).get("BoundingBox", {}).get("Top", 0.0))
        except Exception:
            return 0.0

    table_blocks.sort(key=lambda b: (int(b.get("Page", 0) or 0), _geom_y(b)))

    for tb in table_blocks:
        grid, page_num = _table_to_grid(tb, block_map)
        if page_num not in tables_by_page:
            tables_by_page[page_num] = []
        if grid and any(any((c or "").strip() for c in row) for row in grid):
            tables_by_page[page_num].append(grid)

    return tables_by_page


# =========================================================
# Excel Writer
# =========================================================
def write_tables_to_excel(
    tables_by_page: Dict[int, List[List[List[str]]]],
    out_path: str,
    *,
    max_sheet_name_len: int = 31,
) -> None:
    wb = Workbook()
    default = wb.active
    wb.remove(default)

    pages_sorted = sorted(tables_by_page.keys())

    if not pages_sorted:
        ws = wb.create_sheet("No Tables")
        ws["A1"] = "No tables were detected in this document."
        wb.save(out_path)
        return

    for page_num in pages_sorted:
        sheet_name = f"Page {page_num}"
        sheet_name = sheet_name[:max_sheet_name_len]

        ws = wb.create_sheet(sheet_name)
        row_cursor = 1
        tables = tables_by_page.get(page_num, []) or []

        if not tables:
            ws["A1"] = "No tables detected on this page."
            continue

        for grid in tables:
            for r, row in enumerate(grid, start=0):
                for c, val in enumerate(row, start=1):
                    ws.cell(row=row_cursor + r, column=c).value = val
            row_cursor += len(grid) + 2

        # very light column sizing
        try:
            max_cols = max((len(g[0]) if g and g[0] else 0) for g in tables) if tables else 0
            for col in range(1, max_cols + 1):
                letter = get_column_letter(col)
                max_len = 0
                for row in ws.iter_rows(min_row=1, max_col=col, min_col=col):
                    for cell in row:
                        if cell.value is None:
                            continue
                        s = str(cell.value)
                        if len(s) > max_len:
                            max_len = len(s)
                ws.column_dimensions[letter].width = min(max(10, max_len + 2), 60)
        except Exception:
            pass

    wb.save(out_path)


# =========================================================
# Preview builder (for HTML popup)
# =========================================================
def build_preview(
    tables_by_page: Dict[int, List[List[List[str]]]],
    *,
    max_pages: int = 25,
    max_tables_per_page: int = 5,
    max_rows: int = 40,
    max_cols: int = 20,
) -> dict:
    """
    Lightweight preview to drive an HTML table modal.
    """
    pages_sorted = sorted(tables_by_page.keys())[:max_pages]
    out_pages = []
    for p in pages_sorted:
        tables = tables_by_page.get(p, []) or []
        page_tables = []
        for i, grid in enumerate(tables[:max_tables_per_page], start=1):
            # clamp rows/cols
            clipped = [row[:max_cols] for row in (grid[:max_rows] if grid else [])]
            page_tables.append({
                "table_index": i,
                "grid": clipped,
            })
        out_pages.append({"page": p, "tables": page_tables})
    return {"pages": out_pages}


# =========================================================
# End-to-end Schedule Pipeline
# =========================================================
def run_schedule_pipeline_to_s3(
    job_id: str,
    payload: Dict[str, Any],
    *,
    upload_prefix: str = S3_SCH_UPLOAD_PREFIX,
    processed_prefix: str = S3_SCH_PROCESSED_PREFIX,
    timeout_s: int = 900,
) -> Dict[str, Any]:
    """
    S3 -> Textract -> Tables -> Excel -> S3
    Minimal logging. Updates job state + step for UI.
    """
    set_job(job_id, state="RUNNING", info={"step": "init"})

    original_subject = payload.get("Subject") or "Untitled Project"
    project_id = payload.get("ProjectID") or make_project_id(original_subject)

    display_name = (payload.get("DisplayName") or payload.get("AttachmentName") or "schedule.pdf").strip()
    submitted_at = payload.get("SubmittedAt")

    secure_name = secure_filename(payload.get("AttachmentName") or "schedule.pdf")
    doc_folder = build_internal_doc_slug(display_name, submitted_at=submitted_at)

    s3_ensure_prefix(s3_key(upload_prefix, project_id))
    s3_ensure_prefix(s3_key(processed_prefix, project_id))

    work_root = os.path.join(SCH_PROCESSED_LOCAL, f"sch-job-{job_id}")
    os.makedirs(work_root, exist_ok=True)

    project_slug = slugify(original_subject)
    raw_local = os.path.join(work_root, secure_name)
    excel_local = os.path.join(work_root, f"{project_slug}_{doc_folder}_schedule_tables.xlsx")
    zip_local = os.path.join(work_root, f"{project_slug}_{doc_folder}_schedule_pack.zip")

    upload_key = s3_key(upload_prefix, project_id, doc_folder, "original.pdf")
    processed_prefix_full = s3_key(processed_prefix, project_id, doc_folder)

    excel_s3_key = s3_key(processed_prefix_full, os.path.basename(excel_local))
    zip_s3_key = s3_key(processed_prefix_full, os.path.basename(zip_local))

    try:
        # decode pdf
        set_job(job_id, info={"step": "decode_pdf"})
        logging.warning("SCH[%s] step=decode_pdf", job_id)
        with open(raw_local, "wb") as f:
            f.write(base64.b64decode(payload["AttachmentContent"]))

        # upload original
        set_job(job_id, info={"step": "upload_original", "upload_key": upload_key})
        logging.warning("SCH[%s] step=upload_original", job_id)
        s3_upload_file(raw_local, upload_key)

        # start textract
        set_job(job_id, info={"step": "textract_start"})
        logging.warning("SCH[%s] step=textract_start", job_id)
        job_textract_id = textract_start_tables_job_s3(bucket=S3_BUCKET, key=upload_key)
        set_job(job_id, info={"step": "textract_running", "textract_job_id": job_textract_id})

        # wait + collect pages
        pages = textract_get_all_pages(job_id=job_textract_id, timeout_s=timeout_s, poll_s=1.5)
        set_job(job_id, info={"step": "textract_done", "pages_returned": len(pages)})
        logging.warning("SCH[%s] step=textract_done pages=%s", job_id, len(pages))

        # parse tables
        set_job(job_id, info={"step": "parse_tables"})
        logging.warning("SCH[%s] step=parse_tables", job_id)
        tables_by_page = extract_tables_by_page(pages)

        # build preview for UI
        preview = build_preview(tables_by_page)
        set_job(job_id, preview=preview)

        # write excel
        set_job(job_id, info={"step": "write_excel"})
        logging.warning("SCH[%s] step=write_excel", job_id)
        write_tables_to_excel(tables_by_page, excel_local)

        # zip pack
        set_job(job_id, info={"step": "zip_pack"})
        logging.warning("SCH[%s] step=zip_pack", job_id)
        with zipfile.ZipFile(zip_local, "w", compression=zipfile.ZIP_DEFLATED) as z:
            z.write(raw_local, arcname=os.path.basename(raw_local))
            z.write(excel_local, arcname=os.path.basename(excel_local))

            if INCLUDE_RAW_TEXTRACT_JSON:
                textract_json_local = os.path.join(work_root, "textract_raw.json")
                with open(textract_json_local, "w", encoding="utf-8") as f:
                    json.dump(pages, f)
                z.write(textract_json_local, arcname=os.path.basename(textract_json_local))

        # upload processed
        set_job(job_id, info={"step": "upload_processed"})
        logging.warning("SCH[%s] step=upload_processed", job_id)
        s3_upload_file(excel_local, excel_s3_key)
        s3_upload_file(zip_local, zip_s3_key)

        out = {
            "job_id": job_id,
            "project_id": project_id,
            "doc_folder": doc_folder,
            "upload_key": upload_key,
            "processed_prefix": processed_prefix_full,
            "textract_job_id": job_textract_id,
            "excel_key": excel_s3_key,
            "excel_url": s3_presign_get(excel_s3_key, expires=3600),
            "zip_key": zip_s3_key,
            "zip_url": s3_presign_get(zip_s3_key, expires=3600),
            "tables_pages": sorted(tables_by_page.keys()),
            "display_name": display_name,
            "submitted_at": submitted_at,
        }

        # IMPORTANT: store result where routes expect it
        set_job(job_id, state="DONE", result=out, info={"step": "complete"})
        logging.warning("SCH[%s] step=complete DONE", job_id)
        return out

    except Exception as e:
        logging.error("SCH[%s] FAILED step=%s error=%s", job_id, (get_job(job_id).get("info") or {}).get("step"), str(e))
        set_job(job_id, state="ERROR", info={"step": "error", "error": str(e)})
        raise

    finally:
        _safe_rmtree(work_root)


# =========================================================
# List schedule docs for a project
# =========================================================
def list_schedule_docs(project_id: str) -> List[dict]:
    prefix = s3_key(S3_SCH_PROCESSED_PREFIX, project_id) + "/"
    try:
        resp = _s3.list_objects_v2(Bucket=S3_BUCKET, Prefix=prefix, Delimiter="/")
    except Exception:
        return []

    folders = [cp.get("Prefix", "") for cp in resp.get("CommonPrefixes", []) if cp.get("Prefix")]
    out: List[dict] = []

    for f in folders:
        doc_folder = f.rstrip("/").split("/")[-1]
        try:
            resp2 = _s3.list_objects_v2(Bucket=S3_BUCKET, Prefix=f)
        except Exception:
            continue

        arts: Dict[str, Optional[str]] = {"excel": None, "zip": None, "original": None}

        for obj in resp2.get("Contents", []) or []:
            key = obj.get("Key", "")
            if not key or key.endswith("/"):
                continue
            name = key.rsplit("/", 1)[-1].lower()

            if name.endswith(".xlsx") and "schedule_tables" in name:
                arts["excel"] = key
            elif name.endswith(".zip"):
                arts["zip"] = key

        original_key = s3_key(S3_SCH_UPLOAD_PREFIX, project_id, doc_folder, "original.pdf")
        try:
            _s3.head_object(Bucket=S3_BUCKET, Key=original_key)
            arts["original"] = original_key
        except Exception:
            arts["original"] = None

        out.append({"doc_folder": doc_folder, "artifacts": arts})

    out.sort(key=lambda d: d["doc_folder"])
    return out


# alias to match your route call
def list_schedule_project_docs(project_id: str) -> List[dict]:
    return list_schedule_docs(project_id)


