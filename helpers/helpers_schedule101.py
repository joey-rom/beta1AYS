"""
helpers_schedule.py
==========================
HVAC Schedule Extraction using Claude Vision API

All the heavy lifting for extracting equipment schedules from mechanical 
drawing PDFs. Flask routes should be thin wrappers around these functions.
"""


import anthropic
import base64
import json
import os
import re
import time
import uuid
import filelock
from dataclasses import dataclass, field
from typing import List, Dict, Any, Optional, Callable
from pathlib import Path
from io import BytesIO
from datetime import datetime
from concurrent.futures import ThreadPoolExecutor

# PDF to image conversion
try:
    from pdf2image import convert_from_path, convert_from_bytes
    PDF2IMAGE_AVAILABLE = True
except ImportError:
    PDF2IMAGE_AVAILABLE = False

# Excel generation
try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    OPENPYXL_AVAILABLE = True
except ImportError:
    OPENPYXL_AVAILABLE = False

# PIL for image handling
try:
    from PIL import Image
    PIL_AVAILABLE = True
except ImportError:
    PIL_AVAILABLE = False


# =============================================================================
# CONFIGURATION
# =============================================================================

class Config:
    """Configuration for schedule extraction."""
    MODEL = "claude-sonnet-4-20250514"
    MAX_TOKENS = 8000
    PDF_DPI = 150
    MAX_IMAGE_DIMENSION = 4096
    MAX_CONCURRENT_PAGES = 3
    INPUT_TOKEN_COST = 3.00   # per 1M tokens
    OUTPUT_TOKEN_COST = 15.00  # per 1M tokens


# =============================================================================
# DATA STRUCTURES
# =============================================================================

@dataclass
class ScheduleTable:
    """Represents a single extracted schedule table."""
    table_id: str
    schedule_type: str
    title: str
    headers: List[str]
    rows: List[List[str]]
    page_number: int
    confidence: float = 1.0
    notes: str = ""
    
    @property
    def row_count(self) -> int:
        return len(self.rows)
    
    @property
    def col_count(self) -> int:
        return len(self.headers)
    
    def to_dict(self) -> Dict:
        return {
            "table_id": self.table_id,
            "schedule_type": self.schedule_type,
            "title": self.title,
            "headers": self.headers,
            "rows": self.rows,
            "row_count": self.row_count,
            "col_count": self.col_count,
            "page_number": self.page_number,
            "confidence": self.confidence,
            "notes": self.notes
        }


@dataclass
class ExtractionResult:
    """Complete result of schedule extraction from a PDF."""
    job_id: str
    filename: str
    page_count: int
    tables: List[ScheduleTable] = field(default_factory=list)
    processing_time: float = 0.0
    api_cost_estimate: float = 0.0
    input_tokens: int = 0
    output_tokens: int = 0
    errors: List[str] = field(default_factory=list)
    status: str = "pending"
    progress: int = 0
    created_at: str = field(default_factory=lambda: datetime.utcnow().isoformat())
    
    @property
    def total_rows(self) -> int:
        return sum(t.row_count for t in self.tables)
    
    @property
    def total_equipment(self) -> int:
        return self.total_rows
    
    def to_dict(self) -> Dict:
        return {
            "job_id": self.job_id,
            "filename": self.filename,
            "page_count": self.page_count,
            "tables": [t.to_dict() for t in self.tables],
            "processing_time": round(self.processing_time, 2),
            "api_cost_estimate": round(self.api_cost_estimate, 4),
            "input_tokens": self.input_tokens,
            "output_tokens": self.output_tokens,
            "total_schedules": len(self.tables),
            "total_equipment": self.total_equipment,
            "errors": self.errors,
            "status": self.status,
            "progress": self.progress,
            "created_at": self.created_at
        }


# =============================================================================
# JOB MANAGEMENT (File-based for multi-worker support)
# =============================================================================

_executor = ThreadPoolExecutor(max_workers=4)

# File-based job storage
JOBS_DIR = Path("/tmp/schedule_jobs")
JOBS_DIR.mkdir(exist_ok=True)


def _job_file(job_id: str) -> Path:
    """Get path to job JSON file."""
    return JOBS_DIR / f"{job_id}.json"


def _save_job(job_id: str, data: Dict):
    """Save job data to file."""
    job_file = _job_file(job_id)
    lock_file = job_file.with_suffix('.lock')
    with filelock.FileLock(lock_file, timeout=10):
        with open(job_file, 'w') as f:
            json.dump(data, f)


def _load_job(job_id: str) -> Optional[Dict]:
    """Load job data from file."""
    job_file = _job_file(job_id)
    if not job_file.exists():
        return None
    lock_file = job_file.with_suffix('.lock')
    try:
        with filelock.FileLock(lock_file, timeout=10):
            with open(job_file, 'r') as f:
                return json.load(f)
    except Exception:
        return None


def create_job(filename: str) -> str:
    """Create a new extraction job and return its ID."""
    job_id = str(uuid.uuid4())[:8]
    job_data = {
        "job_id": job_id,
        "filename": filename,
        "page_count": 0,
        "tables": [],
        "processing_time": 0.0,
        "api_cost_estimate": 0.0,
        "input_tokens": 0,
        "output_tokens": 0,
        "errors": [],
        "status": "pending",
        "progress": 0,
        "created_at": datetime.utcnow().isoformat()
    }
    _save_job(job_id, job_data)
    return job_id


def get_job(job_id: str) -> Optional[ExtractionResult]:
    """Get job by ID, returns ExtractionResult object."""
    data = _load_job(job_id)
    if not data:
        return None
    
    # Reconstruct ExtractionResult from dict
    tables = []
    for t in data.get("tables", []):
        tables.append(ScheduleTable(
            table_id=t.get("table_id", ""),
            schedule_type=t.get("schedule_type", ""),
            title=t.get("title", ""),
            headers=t.get("headers", []),
            rows=t.get("rows", []),
            page_number=t.get("page_number", 1),
            confidence=t.get("confidence", 1.0),
            notes=t.get("notes", "")
        ))
    
    return ExtractionResult(
        job_id=data.get("job_id", job_id),
        filename=data.get("filename", ""),
        page_count=data.get("page_count", 0),
        tables=tables,
        processing_time=data.get("processing_time", 0.0),
        api_cost_estimate=data.get("api_cost_estimate", 0.0),
        input_tokens=data.get("input_tokens", 0),
        output_tokens=data.get("output_tokens", 0),
        errors=data.get("errors", []),
        status=data.get("status", "pending"),
        progress=data.get("progress", 0),
        created_at=data.get("created_at", "")
    )


def update_job(job_id: str, **kwargs):
    """Update job attributes."""
    data = _load_job(job_id)
    if not data:
        return
    
    for key, value in kwargs.items():
        if key == "tables":
            # Convert ScheduleTable objects to dicts
            data["tables"] = [t.to_dict() if hasattr(t, 'to_dict') else t for t in value]
        else:
            data[key] = value
    
    _save_job(job_id, data)


def get_job_status(job_id: str) -> Optional[Dict]:
    """Get job status as dict (for API responses)."""
    job = get_job(job_id)
    if job:
        return {
            "job_id": job.job_id,
            "status": job.status,
            "progress": job.progress,
            "filename": job.filename,
            "page_count": job.page_count,
            "tables_found": len(job.tables),
            "total_equipment": job.total_equipment,
            "processing_time": round(job.processing_time, 2),
            "errors": job.errors
        }
    return None


# =============================================================================
# CLAUDE API PROMPT
# =============================================================================

EXTRACTION_PROMPT = """You are an expert at reading mechanical engineering drawings and extracting equipment schedules.

Analyze this mechanical drawing page and extract ALL equipment schedules you can find.

Common schedule types: AHU, RTU, FCU, VAV, VRF, PUMP, FAN, BOILER, CHILLER, UNIT HEATER, CONVECTOR, CABINET HEATER, DUCT COIL, AIR DEVICE, EXPANSION TANK, AIR SEPARATOR, CONDENSING UNIT, SPLIT SYSTEM, HEAT PUMP, ERV/HRV, MAU, COIL.

Return ONLY valid JSON in this exact format:
{
    "schedules": [
        {
            "title": "EXACT SCHEDULE TITLE FROM DRAWING",
            "type": "SCHEDULE_TYPE",
            "headers": ["COL1", "COL2", "COL3"],
            "rows": [
                ["val1", "val2", "val3"],
                ["val1", "", "val3"]
            ],
            "notes": ""
        }
    ]
}

CRITICAL RULES:
- Extract EVERY schedule visible on the page
- Include EVERY row, even if some cells are empty
- Use "" for empty cells, never skip columns
- Each row array must have exactly the same length as headers array
- Preserve exact formatting (fractions like "1-1/4", units, model numbers)
- Include manufacturer names and model numbers exactly as shown
- If you cannot read a value clearly, use "?" 
- Return ONLY the JSON object, no markdown, no explanation"""


# =============================================================================
# SCHEDULE TYPE NORMALIZATION
# =============================================================================

SCHEDULE_TYPE_MAPPING = {
    'AIR HANDLING UNIT': 'AHU', 'AIR HANDLER': 'AHU', 'AHU': 'AHU', 'DOAS': 'AHU',
    'ROOFTOP UNIT': 'RTU', 'RTU': 'RTU', 'PACKAGED UNIT': 'RTU',
    'FAN COIL': 'FCU', 'FAN COIL UNIT': 'FCU', 'FCU': 'FCU',
    'VAV': 'VAV', 'VARIABLE AIR VOLUME': 'VAV',
    'VRF': 'VRF', 'VARIABLE REFRIGERANT': 'VRF', 'VRV': 'VRF',
    'PUMP': 'PUMP', 'PUMPS': 'PUMP',
    'FAN': 'FAN', 'EXHAUST FAN': 'FAN', 'SUPPLY FAN': 'FAN', 'EF': 'FAN',
    'BOILER': 'BOILER', 'CONDENSING BOILER': 'BOILER',
    'CHILLER': 'CHILLER',
    'UNIT HEATER': 'UNIT_HEATER', 'UH': 'UNIT_HEATER',
    'CONVECTOR': 'CONVECTOR', 'CABINET HEATER': 'CONVECTOR',
    'COIL': 'COIL', 'DUCT COIL': 'COIL', 'REHEAT COIL': 'COIL',
    'AIR DEVICE': 'AIR_DEVICE', 'DIFFUSER': 'AIR_DEVICE', 'GRILLE': 'AIR_DEVICE',
    'EXPANSION TANK': 'EXPANSION_TANK',
    'AIR SEPARATOR': 'AIR_SEPARATOR',
    'CONDENSING UNIT': 'CONDENSING_UNIT', 'CONDENSER': 'CONDENSING_UNIT', 'CU': 'CONDENSING_UNIT',
    'SPLIT SYSTEM': 'SPLIT_SYSTEM', 'MINI SPLIT': 'SPLIT_SYSTEM',
    'HEAT PUMP': 'HEAT_PUMP', 'HP': 'HEAT_PUMP', 'WSHP': 'HEAT_PUMP',
    'ERV': 'ERV', 'HRV': 'ERV', 'ENERGY RECOVERY': 'ERV',
    'MAU': 'MAU', 'MAKEUP AIR': 'MAU',
}


def normalize_schedule_type(raw_type: str) -> str:
    """Normalize schedule type to standard abbreviation."""
    if not raw_type:
        return "OTHER"
    upper = raw_type.upper().strip()
    if upper in SCHEDULE_TYPE_MAPPING:
        return SCHEDULE_TYPE_MAPPING[upper]
    for key, value in SCHEDULE_TYPE_MAPPING.items():
        if key in upper or upper in key:
            return value
    return "OTHER"


# =============================================================================
# CORE EXTRACTION FUNCTIONS
# =============================================================================

def _resize_image_if_needed(image: 'Image.Image', max_dim: int = Config.MAX_IMAGE_DIMENSION) -> 'Image.Image':
    """Resize image if it exceeds max dimensions."""
    width, height = image.size
    if width > max_dim or height > max_dim:
        ratio = min(max_dim / width, max_dim / height)
        new_size = (int(width * ratio), int(height * ratio))
        return image.resize(new_size, Image.Resampling.LANCZOS)
    return image


def _image_to_base64(image: 'Image.Image') -> str:
    """Convert PIL image to base64 string."""
    buffer = BytesIO()
    image.save(buffer, format='PNG', optimize=True)
    return base64.standard_b64encode(buffer.getvalue()).decode('utf-8')


def _extract_page(client: anthropic.Anthropic, image: 'Image.Image', page_num: int, model: str = Config.MODEL) -> Dict:
    """Extract schedules from a single page image."""
    image = _resize_image_if_needed(image)
    image_b64 = _image_to_base64(image)
    
    response = client.messages.create(
        model=model,
        max_tokens=Config.MAX_TOKENS,
        messages=[{
            "role": "user",
            "content": [
                {"type": "image", "source": {"type": "base64", "media_type": "image/png", "data": image_b64}},
                {"type": "text", "text": EXTRACTION_PROMPT}
            ]
        }]
    )
    
    response_text = response.content[0].text
    
    try:
        cleaned = re.sub(r'^```json\s*', '', response_text.strip())
        cleaned = re.sub(r'\s*```$', '', cleaned)
        json_match = re.search(r'\{[\s\S]*\}', cleaned)
        if json_match:
            data = json.loads(json_match.group())
        else:
            data = {"schedules": [], "error": "No JSON found in response"}
    except json.JSONDecodeError as e:
        data = {"schedules": [], "error": f"JSON parse error: {str(e)}"}
    
    return {
        "page_number": page_num,
        "schedules": data.get("schedules", []),
        "input_tokens": response.usage.input_tokens,
        "output_tokens": response.usage.output_tokens,
        "error": data.get("error")
    }


def extract_schedules_from_pdf(
    pdf_path: str,
    job_id: Optional[str] = None,
    api_key: Optional[str] = None,
    model: str = Config.MODEL,
    max_pages: Optional[int] = None,
    dpi: int = Config.PDF_DPI,
    progress_callback: Optional[Callable[[int, str], None]] = None
) -> ExtractionResult:
    """
    Extract equipment schedules from a PDF using Claude Vision API.
    
    Args:
        pdf_path: Path to the PDF file
        job_id: Optional job ID for tracking
        api_key: Anthropic API key (uses env var if not provided)
        model: Claude model to use
        max_pages: Maximum pages to process (None for all)
        dpi: Resolution for PDF conversion
        progress_callback: Optional callback(progress_pct, message)
        
    Returns:
        ExtractionResult with all extracted schedules
    """
    start_time = time.time()
    filename = Path(pdf_path).name
    job_id = job_id or create_job(filename)
    
    result = get_job(job_id) or ExtractionResult(job_id=job_id, filename=filename, page_count=0)
    result.status = "processing"
    result.filename = filename
    
    def update_progress(pct: int, msg: str = ""):
        result.progress = pct
        update_job(job_id, progress=pct, status="processing")
        if progress_callback:
            progress_callback(pct, msg)
    
    # Validate dependencies
    if not PDF2IMAGE_AVAILABLE:
        result.errors.append("pdf2image not installed. Run: pip install pdf2image")
        result.status = "failed"
        return result
    
    if not PIL_AVAILABLE:
        result.errors.append("Pillow not installed. Run: pip install pillow")
        result.status = "failed"
        return result
    
    api_key = api_key or os.environ.get("ANTHROPIC_API_KEY")
    if not api_key:
        result.errors.append("ANTHROPIC_API_KEY not set")
        result.status = "failed"
        return result
    
    client = anthropic.Anthropic(api_key=api_key)
    
    # Convert PDF to images
    update_progress(5, "Converting PDF to images...")
    try:
        images = convert_from_path(pdf_path, dpi=dpi)
        result.page_count = len(images)
        if max_pages:
            images = images[:max_pages]
    except Exception as e:
        result.errors.append(f"PDF conversion error: {str(e)}")
        result.status = "failed"
        return result
    
    update_job(job_id, page_count=result.page_count)
    
    # Process each page
    total_input_tokens = 0
    total_output_tokens = 0
    table_counter = 0
    
    for idx, image in enumerate(images):
        page_num = idx + 1
        progress_pct = 10 + int((idx / len(images)) * 80)
        update_progress(progress_pct, f"Processing page {page_num}/{len(images)}...")
        
        try:
            page_result = _extract_page(client, image, page_num, model)
            
            total_input_tokens += page_result["input_tokens"]
            total_output_tokens += page_result["output_tokens"]
            
            if page_result.get("error"):
                result.errors.append(f"Page {page_num}: {page_result['error']}")
            
            for schedule_data in page_result["schedules"]:
                table_counter += 1
                
                headers = schedule_data.get("headers", [])
                rows = schedule_data.get("rows", [])
                header_count = len(headers)
                
                # Fix rows that don't match header count
                fixed_rows = []
                for row in rows:
                    if len(row) < header_count:
                        row = row + [""] * (header_count - len(row))
                    elif len(row) > header_count:
                        row = row[:header_count]
                    fixed_rows.append(row)
                
                table = ScheduleTable(
                    table_id=f"schedule_{table_counter}",
                    schedule_type=normalize_schedule_type(schedule_data.get("type", "")),
                    title=schedule_data.get("title", f"Schedule {table_counter}"),
                    headers=headers,
                    rows=fixed_rows,
                    page_number=page_num,
                    notes=schedule_data.get("notes", "")
                )
                result.tables.append(table)
                
        except Exception as e:
            result.errors.append(f"Page {page_num} extraction error: {str(e)}")
    
    # Calculate costs
    result.input_tokens = total_input_tokens
    result.output_tokens = total_output_tokens
    result.api_cost_estimate = (
        (total_input_tokens / 1_000_000) * Config.INPUT_TOKEN_COST +
        (total_output_tokens / 1_000_000) * Config.OUTPUT_TOKEN_COST
    )
    
    result.processing_time = time.time() - start_time
    result.status = "completed" if not any("error" in e.lower() for e in result.errors) else "completed_with_errors"
    result.progress = 100
    
    update_job(
        job_id,
        tables=result.tables,
        processing_time=result.processing_time,
        api_cost_estimate=result.api_cost_estimate,
        input_tokens=result.input_tokens,
        output_tokens=result.output_tokens,
        status=result.status,
        progress=100,
        errors=result.errors
    )
    
    return result


def start_async_extraction(pdf_path: str, **kwargs) -> str:
    """Start extraction in background thread. Returns job_id."""
    filename = Path(pdf_path).name
    job_id = create_job(filename)
    
    def run():
        extract_schedules_from_pdf(pdf_path, job_id=job_id, **kwargs)
    
    _executor.submit(run)
    return job_id


# =============================================================================
# EXCEL GENERATION
# =============================================================================

def generate_schedule_excel(result: ExtractionResult, output_path: str) -> str:
    """Generate Excel file with extracted schedules."""
    if not OPENPYXL_AVAILABLE:
        raise ImportError("openpyxl not installed. Run: pip install openpyxl")
    
    wb = Workbook()
    
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="2C5282", end_color="2C5282", fill_type="solid")
    header_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
    thin_border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
    
    # Summary sheet
    summary = wb.active
    summary.title = "Summary"
    summary_data = [
        ["HVAC Schedule Extraction Summary"], [""],
        ["Filename", result.filename], ["Pages", result.page_count],
        ["Schedules Found", len(result.tables)], ["Total Equipment", result.total_equipment],
        [""], ["Schedules by Type:"]
    ]
    
    type_counts = {}
    for t in result.tables:
        type_counts[t.schedule_type] = type_counts.get(t.schedule_type, 0) + t.row_count
    for stype, count in sorted(type_counts.items()):
        summary_data.append([f"  {stype}", count])
    
    for row_idx, row_data in enumerate(summary_data, 1):
        for col_idx, value in enumerate(row_data, 1):
            cell = summary.cell(row=row_idx, column=col_idx, value=value)
            if row_idx == 1:
                cell.font = Font(bold=True, size=14)
    summary.column_dimensions['A'].width = 25
    summary.column_dimensions['B'].width = 40
    
    # Sheet for each schedule
    for table in result.tables:
        sheet_name = re.sub(r'[\\/*?:\[\]]', '', table.title)[:30]
        if sheet_name in [ws.title for ws in wb.worksheets]:
            sheet_name = f"{sheet_name[:27]}_{table.table_id[-3:]}"
        
        ws = wb.create_sheet(title=sheet_name)
        ws.cell(row=1, column=1, value=table.title).font = Font(bold=True, size=12)
        ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(table.headers) or 1)
        
        for col_idx, header in enumerate(table.headers, 1):
            cell = ws.cell(row=3, column=col_idx, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_align
            cell.border = thin_border
        
        for row_idx, row_data in enumerate(table.rows, 4):
            for col_idx, value in enumerate(row_data, 1):
                cell = ws.cell(row=row_idx, column=col_idx, value=value if value else "")
                cell.border = thin_border
        
        for col_idx, header in enumerate(table.headers, 1):
            max_length = len(str(header))
            for row in table.rows:
                if col_idx <= len(row):
                    max_length = max(max_length, len(str(row[col_idx - 1] or "")))
            ws.column_dimensions[get_column_letter(col_idx)].width = min(max_length + 2, 50)
        
        ws.freeze_panes = 'A4'
    
    wb.save(output_path)
    return output_path


# =============================================================================
# JSON EXPORT
# =============================================================================

def generate_schedule_json(result: ExtractionResult) -> str:
    """Generate JSON string of extraction result."""
    return json.dumps(result.to_dict(), indent=2)


# =============================================================================
# CONVENIENCE FUNCTIONS FOR FLASK ROUTES
# =============================================================================

def process_upload(file_storage, async_mode: bool = True) -> Dict[str, Any]:
    """
    Process an uploaded PDF file (for Flask routes).
    
    Args:
        file_storage: Flask request.files['file'] object
        async_mode: If True, returns job_id immediately
        
    Returns:
        Dict with job_id and status info
    """
    import tempfile
    
    filename = file_storage.filename or "upload.pdf"
    
    with tempfile.NamedTemporaryFile(suffix='.pdf', delete=False) as tmp:
        file_storage.save(tmp)
        tmp_path = tmp.name
    
    if async_mode:
        job_id = start_async_extraction(tmp_path)
        return {
            "success": True,
            "job_id": job_id,
            "filename": filename,
            "status": "processing",
            "message": "Extraction started. Poll /api/schedule/status/<job_id> for progress."
        }
    else:
        result = extract_schedules_from_pdf(tmp_path)
        os.unlink(tmp_path)
        return {"success": True, "job_id": result.job_id, "result": result.to_dict()}


def get_result_excel_bytes(job_id: str) -> Optional[bytes]:
    """Get Excel file bytes for a completed job."""
    import tempfile
    job = get_job(job_id)
    if job and job.status in ("completed", "completed_with_errors"):
        with tempfile.NamedTemporaryFile(suffix='.xlsx', delete=False) as tmp:
            generate_schedule_excel(job, tmp.name)
            with open(tmp.name, 'rb') as f:
                data = f.read()
            os.unlink(tmp.name)
            return data
    return None


def get_result_json(job_id: str) -> Optional[str]:
    """Get JSON output for a completed job."""
    job = get_job(job_id)
    if job and job.status in ("completed", "completed_with_errors"):
        return generate_schedule_json(job)
    return None